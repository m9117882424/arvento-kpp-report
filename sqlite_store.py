from __future__ import annotations

import csv
import sqlite3
from contextlib import contextmanager
from pathlib import Path
from typing import Iterable, Iterator

from openpyxl import load_workbook

from arvento_io import Point, as_datetime, as_float, detect_columns, is_header_row

BATCH_SIZE = 10_000


def detect_csv_encoding(path: Path) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp1251", "cp1254"):
        try:
            with path.open("r", encoding=encoding, newline="") as stream:
                stream.read(256 * 1024)
            return encoding
        except UnicodeDecodeError:
            continue
    raise ValueError("Не удалось определить кодировку CSV")


def detect_csv_delimiter(path: Path, encoding: str) -> str:
    with path.open("r", encoding=encoding, newline="") as stream:
        sample = stream.read(64 * 1024)
    try:
        return csv.Sniffer().sniff(sample, delimiters=";,\t|").delimiter
    except csv.Error:
        return ";"


def row_to_point(row: list[object], columns: dict[str, int]) -> Point | None:
    if not row or len(row) <= max(columns.values()):
        return None
    plate = str(row[columns["plate"]] or "").strip()
    timestamp = as_datetime(row[columns["time"]])
    lat = as_float(row[columns["lat"]])
    lon = as_float(row[columns["lon"]])
    if not plate or timestamp is None or lat is None or lon is None:
        return None
    return Point(
        plate=plate,
        time=timestamp,
        lat=lat,
        lon=lon,
        odometer=as_float(row[columns["odometer"]]) if "odometer" in columns else None,
        source_distance=as_float(row[columns["distance"]]) if "distance" in columns else None,
        speed=as_float(row[columns["speed"]]) if "speed" in columns else None,
        region=str(row[columns["region"]] or "").strip() if "region" in columns else "",
        address=str(row[columns["address"]] or "").strip() if "address" in columns else "",
    )


def iter_csv_points(path: Path) -> Iterator[Point]:
    encoding = detect_csv_encoding(path)
    delimiter = detect_csv_delimiter(path, encoding)
    with path.open("r", encoding=encoding, newline="") as stream:
        reader = csv.reader(stream, delimiter=delimiter)
        headers: list[object] | None = None
        columns: dict[str, int] | None = None
        for row in reader:
            if headers is None:
                if is_header_row(row):
                    headers = list(row)
                    columns = detect_columns(headers)
                continue
            point = row_to_point(list(row), columns or {})
            if point is not None:
                yield point


def iter_excel_points(path: Path) -> Iterator[Point]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.worksheets[0]
        headers: list[object] | None = None
        columns: dict[str, int] | None = None
        for row in sheet.iter_rows(values_only=True):
            values = list(row)
            if headers is None:
                if is_header_row(values):
                    headers = values
                    columns = detect_columns(headers)
                continue
            point = row_to_point(values, columns or {})
            if point is not None:
                yield point
    finally:
        workbook.close()


def iter_points(path: Path) -> Iterator[Point]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        yield from iter_csv_points(path)
        return
    if suffix in (".xlsx", ".xlsm"):
        yield from iter_excel_points(path)
        return
    raise ValueError("Поддерживаются только XLSX, XLSM и CSV")


def create_database(db_path: Path) -> sqlite3.Connection:
    connection = sqlite3.connect(db_path)
    connection.execute("PRAGMA journal_mode=WAL")
    connection.execute("PRAGMA synchronous=NORMAL")
    connection.execute("PRAGMA temp_store=FILE")
    connection.execute("PRAGMA cache_size=-65536")
    connection.execute(
        """
        CREATE TABLE IF NOT EXISTS points (
            plate TEXT NOT NULL,
            ts TEXT NOT NULL,
            day TEXT NOT NULL,
            lat REAL NOT NULL,
            lon REAL NOT NULL,
            odometer REAL,
            source_distance REAL,
            speed REAL,
            region TEXT,
            address TEXT
        )
        """
    )
    connection.commit()
    return connection


def import_source_to_sqlite(source: Path, db_path: Path) -> dict[str, int]:
    connection = create_database(db_path)
    loaded = 0
    skipped = 0
    batch: list[tuple[object, ...]] = []
    try:
        for point in iter_points(source):
            batch.append(
                (
                    point.plate,
                    point.time.isoformat(sep=" "),
                    point.time.date().isoformat(),
                    point.lat,
                    point.lon,
                    point.odometer,
                    point.source_distance,
                    point.speed,
                    point.region,
                    point.address,
                )
            )
            loaded += 1
            if len(batch) >= BATCH_SIZE:
                connection.executemany(
                    "INSERT INTO points VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                    batch,
                )
                connection.commit()
                batch.clear()
            if loaded % 500_000 == 0:
                print(f"Импортировано в SQLite: {loaded:,}".replace(",", " "))
        if batch:
            connection.executemany(
                "INSERT INTO points VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                batch,
            )
            connection.commit()
        print("Создание индекса SQLite...")
        connection.execute("CREATE INDEX IF NOT EXISTS idx_points_day_plate_ts ON points(day, plate, ts)")
        connection.execute("CREATE INDEX IF NOT EXISTS idx_points_plate_ts ON points(plate, ts)")
        connection.commit()
        return {"loaded": loaded, "skipped": skipped}
    finally:
        connection.close()


def list_days(db_path: Path) -> list[str]:
    with sqlite3.connect(db_path) as connection:
        return [row[0] for row in connection.execute("SELECT DISTINCT day FROM points ORDER BY day")]


def list_plates_for_day(db_path: Path, day: str) -> list[str]:
    with sqlite3.connect(db_path) as connection:
        return [
            row[0]
            for row in connection.execute(
                "SELECT DISTINCT plate FROM points WHERE day=? ORDER BY plate",
                (day,),
            )
        ]


def load_vehicle_day_points(db_path: Path, day: str, plate: str) -> list[Point]:
    with sqlite3.connect(db_path) as connection:
        rows = connection.execute(
            """
            SELECT plate, ts, lat, lon, odometer, source_distance, speed, region, address
            FROM points
            WHERE day=? AND plate=?
            ORDER BY ts
            """,
            (day, plate),
        )
        return [
            Point(
                plate=row[0],
                time=as_datetime(row[1]),
                lat=float(row[2]),
                lon=float(row[3]),
                odometer=row[4],
                source_distance=row[5],
                speed=row[6],
                region=row[7] or "",
                address=row[8] or "",
            )
            for row in rows
        ]
