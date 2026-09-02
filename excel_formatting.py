#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Explicit workbook formatting shared by report writers."""
from __future__ import annotations

from os import PathLike
from typing import BinaryIO

from openpyxl.styles.numbers import is_date_format
from openpyxl.workbook.workbook import Workbook


COORDINATE_HEADER_TOKENS = (
    "координат",
    "latitude",
    "longitude",
    "широт",
    "долгот",
    "latitudine",
    "enlem",
    "boylam",
)


def detect_header_row(sheet) -> int:
    upper = min(max(sheet.max_row, 1), 10)
    candidates = []
    for row_number in range(1, upper + 1):
        populated = sum(
            1 for cell in sheet[row_number] if cell.value not in (None, "")
        )
        candidates.append((populated, row_number))
    maximum = max((value for value, _row in candidates), default=0)
    return min(
        (row for value, row in candidates if value == maximum),
        default=1,
    )


def apply_one_decimal_metrics(workbook: Workbook) -> None:
    """Round measurable floats while preserving dates, times, and coordinates."""
    for sheet in workbook.worksheets:
        header_row = detect_header_row(sheet)
        headers = {
            column: str(sheet.cell(header_row, column).value or "").strip().casefold()
            for column in range(1, sheet.max_column + 1)
        }
        for row in sheet.iter_rows(min_row=header_row + 1):
            for cell in row:
                if not isinstance(cell.value, float):
                    continue
                header = headers.get(cell.column, "")
                if any(token in header for token in COORDINATE_HEADER_TOKENS):
                    continue
                number_format = str(cell.number_format or "General")
                lowered_format = number_format.casefold()
                if is_date_format(number_format) or any(
                    token in lowered_format
                    for token in ("yy", "dd", "hh", "ss", "[h]")
                ):
                    continue
                if "%" in number_format:
                    cell.number_format = "0.0%"
                else:
                    cell.value = round(cell.value, 1)
                    cell.number_format = "0.0"


def save_report_workbook(
    workbook: Workbook,
    target: str | PathLike[str] | BinaryIO,
) -> None:
    apply_one_decimal_metrics(workbook)
    workbook.save(target)


__all__ = [
    "apply_one_decimal_metrics",
    "detect_header_row",
    "save_report_workbook",
]
