#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Prepare the user-facing consolidated Excel workbook.

Internal workbooks retain diagnostics and parameter sheets while calculations
and cache writes are in progress. The downloadable workbook is finalized only
at the portal boundary: it keeps the report sheet, appends roster attributes on
the right and removes every service sheet.
"""
from __future__ import annotations

from copy import copy
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Mapping, Sequence

import psycopg
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from responsible_roster_fields import ensure_schema
from roster_registry import normalize_plate

REPORT_SHEET = "Сводный отчет"
EXPORT_HEADERS = (
    "Имя водителя",
    "Грейд",
    "Дирекция",
    "Ответственный",
)


@dataclass(frozen=True, slots=True)
class RosterExportDetails:
    driver_name: str = ""
    grade: str = ""
    directorate: str = ""
    responsible: str = ""


def _as_day(value) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if value is None:
        return None
    text = str(value).strip()
    for parser in (
        lambda: date.fromisoformat(text),
        lambda: datetime.strptime(text, "%d.%m.%Y").date(),
    ):
        try:
            return parser()
        except ValueError:
            continue
    return None


def _select_roster_day(cursor, report_day: date) -> date | None:
    cursor.execute(
        """
        SELECT roster_day
        FROM consolidated_roster_snapshots
        ORDER BY
            CASE WHEN roster_day <= %s THEN 0 ELSE 1 END,
            CASE WHEN roster_day <= %s THEN roster_day END DESC,
            CASE WHEN roster_day > %s THEN roster_day END ASC
        LIMIT 1
        """,
        (report_day, report_day, report_day),
    )
    row = cursor.fetchone()
    return row[0] if row else None


def load_roster_export_details(
    database_url: str,
    report_days: Sequence[date],
) -> dict[tuple[date, str], RosterExportDetails]:
    """Load central roster fields for every requested report date and plate."""
    days = sorted(set(report_days))
    if not days:
        return {}

    result: dict[tuple[date, str], RosterExportDetails] = {}
    entries_by_roster_day: dict[date, dict[str, RosterExportDetails]] = {}

    with psycopg.connect(database_url) as connection:
        ensure_schema(connection)
        connection.commit()
        with connection.cursor() as cursor:
            for report_day in days:
                roster_day = _select_roster_day(cursor, report_day)
                if roster_day is None:
                    continue

                if roster_day not in entries_by_roster_day:
                    cursor.execute(
                        """
                        SELECT
                            normalized_plate,
                            user_name,
                            grade,
                            directorate,
                            responsible
                        FROM consolidated_roster_entries
                        WHERE roster_day=%s
                        """,
                        (roster_day,),
                    )
                    entries_by_roster_day[roster_day] = {
                        str(normalized_plate): RosterExportDetails(
                            driver_name=str(user_name or ""),
                            grade=str(grade or ""),
                            directorate=str(directorate or ""),
                            responsible=str(responsible or ""),
                        )
                        for (
                            normalized_plate,
                            user_name,
                            grade,
                            directorate,
                            responsible,
                        ) in cursor.fetchall()
                    }

                for normalized_plate, details in entries_by_roster_day[roster_day].items():
                    result[(report_day, normalized_plate)] = details

    return result


def _copy_header_style(source, target) -> None:
    if source.has_style:
        target._style = copy(source._style)
    if source.number_format:
        target.number_format = source.number_format
    target.font = copy(source.font)
    target.fill = copy(source.fill)
    target.border = copy(source.border)
    target.alignment = copy(source.alignment)
    target.protection = copy(source.protection)


def finalize_consolidated_workbook(
    output_path: Path,
    roster_details: Mapping[tuple[date, str], RosterExportDetails],
) -> dict[str, int]:
    """Append roster columns and leave exactly one report sheet."""
    workbook = load_workbook(output_path)
    try:
        if REPORT_SHEET not in workbook.sheetnames:
            raise RuntimeError(f"В книге отсутствует лист «{REPORT_SHEET}»")

        sheet = workbook[REPORT_SHEET]
        headers = {
            str(cell.value or "").strip(): cell.column
            for cell in sheet[1]
        }
        day_column = headers.get("Дата")
        plate_column = next(
            (
                headers[name]
                for name in ("Госномер / Plaka", "Госномер", "Номерной знак")
                if name in headers
            ),
            None,
        )
        source_driver_column = next(
            (
                headers[name]
                for name in ("ПОЛЬЗОВАТЕЛЬ / KULLANICI", "Пользователь", "Водитель")
                if name in headers
            ),
            None,
        )
        source_grade_column = next(
            (
                headers[name]
                for name in ("Грейд / SCALA", "Грейд", "SCALA", "Grade")
                if name in headers
            ),
            None,
        )
        if day_column is None or plate_column is None:
            raise RuntimeError("В сводном отчёте не найдены столбцы даты или госномера")

        original_last_column = sheet.max_column
        header_template = sheet.cell(1, original_last_column)
        export_columns: dict[str, int] = {}
        next_column = original_last_column + 1

        for header in EXPORT_HEADERS:
            existing = headers.get(header)
            column = existing or next_column
            if existing is None:
                next_column += 1
            export_columns[header] = column
            cell = sheet.cell(1, column, header)
            _copy_header_style(header_template, cell)
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        widths = {
            "Имя водителя": 38,
            "Грейд": 12,
            "Дирекция": 52,
            "Ответственный": 36,
        }
        enriched_rows = 0
        for row_index in range(2, sheet.max_row + 1):
            report_day = _as_day(sheet.cell(row_index, day_column).value)
            normalized_plate = normalize_plate(sheet.cell(row_index, plate_column).value)
            details = (
                roster_details.get((report_day, normalized_plate))
                if report_day is not None and normalized_plate
                else None
            )

            fallback_driver = (
                str(sheet.cell(row_index, source_driver_column).value or "")
                if source_driver_column is not None
                else ""
            )
            fallback_grade = (
                str(sheet.cell(row_index, source_grade_column).value or "")
                if source_grade_column is not None
                else ""
            )
            values = {
                "Имя водителя": details.driver_name if details and details.driver_name else fallback_driver,
                "Грейд": details.grade if details and details.grade else fallback_grade,
                "Дирекция": details.directorate if details else "",
                "Ответственный": details.responsible if details else "",
            }

            if details is not None:
                enriched_rows += 1
            for header, value in values.items():
                cell = sheet.cell(row_index, export_columns[header], value)
                cell.alignment = Alignment(vertical="center", wrap_text=True)

        for header, column in export_columns.items():
            sheet.column_dimensions[get_column_letter(column)].width = widths[header]

        sheet.auto_filter.ref = sheet.dimensions
        sheet.sheet_view.showGridLines = False

        for sheet_name in list(workbook.sheetnames):
            if sheet_name != REPORT_SHEET:
                del workbook[sheet_name]
        workbook.active = 0
        workbook.save(output_path)

        return {
            "rows": max(0, sheet.max_row - 1),
            "enriched_rows": enriched_rows,
            "sheets": 1,
            "columns": sheet.max_column,
        }
    finally:
        workbook.close()


def finalize_consolidated_export(
    output_path: Path,
    database_url: str,
) -> dict[str, int]:
    """Load required central-roster data and finalize one downloadable workbook."""
    workbook = load_workbook(output_path, read_only=True, data_only=True)
    try:
        if REPORT_SHEET not in workbook.sheetnames:
            raise RuntimeError(f"В книге отсутствует лист «{REPORT_SHEET}»")
        sheet = workbook[REPORT_SHEET]
        headers = {
            str(cell.value or "").strip(): cell.column
            for cell in sheet[1]
        }
        day_column = headers.get("Дата")
        if day_column is None:
            raise RuntimeError("В сводном отчёте не найден столбец «Дата»")
        report_days = [
            day
            for row_index in range(2, sheet.max_row + 1)
            if (day := _as_day(sheet.cell(row_index, day_column).value)) is not None
        ]
    finally:
        workbook.close()

    details = load_roster_export_details(database_url, report_days)
    return finalize_consolidated_workbook(output_path, details)


__all__ = [
    "EXPORT_HEADERS",
    "REPORT_SHEET",
    "RosterExportDetails",
    "finalize_consolidated_export",
    "finalize_consolidated_workbook",
    "load_roster_export_details",
]
