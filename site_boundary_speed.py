#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""Authoritative Akkuyu site classification for speed violations.

The overall site polygon marked with ``purpose=site_boundary`` in geozones.json
is authoritative. This avoids errors when a vehicle starts the day inside the
site or when a KPP crossing is missing from the GPS sequence.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Sequence

from openpyxl import load_workbook
from excel_formatting import save_report_workbook

from arvento_io import Point
from geozone_registry import (
    Registry,
    find_site_boundary,
    point_in_zone,
    suppress_speed_in_exclusions,
)
from speed_violation_report import (
    DEFAULT_OUTSIDE_SPEED_THRESHOLD_KMH,
    DEFAULT_SITE_SPEED_THRESHOLD_KMH,
    MAX_SPEED_EVENT_GAP_SECONDS,
    OUTSIDE_SPEED_LIMIT_KMH,
    SITE_SPEED_LIMIT_KMH,
    SpeedViolation,
    _event_is_smooth,
    _valid_speed,
    validate_speed_thresholds,
)


def _as_point(point: Any) -> Point:
    return Point(
        plate=point.plate,
        time=point.timestamp,
        lat=float(point.lat),
        lon=float(point.lon),
        speed=float(point.speed) if point.speed is not None else None,
        address=point.address or "",
    )


def classify_site_state_by_polygon(
    track: Sequence[Any],
    registry: Registry,
) -> list[tuple[Any, bool]]:
    """Classify GPS points by the authoritative site polygon.

    A single point with the opposite state between two equal neighbours is
    treated as boundary GPS jitter and receives the neighbours' state.
    """
    if not track:
        return []

    boundary = find_site_boundary(registry)
    raw_states = [
        point_in_zone(float(point.lat), float(point.lon), boundary)
        for point in track
    ]
    stable_states = list(raw_states)

    for index in range(1, len(raw_states) - 1):
        if raw_states[index - 1] == raw_states[index + 1] != raw_states[index]:
            stable_states[index] = raw_states[index - 1]

    return list(zip(track, stable_states))


def detect_speed_violations_by_polygon(
    track: Sequence[Any],
    registry: Registry,
    site_threshold_kmh: float = DEFAULT_SITE_SPEED_THRESHOLD_KMH,
    outside_threshold_kmh: float = DEFAULT_OUTSIDE_SPEED_THRESHOLD_KMH,
) -> tuple[list[SpeedViolation], list[SpeedViolation]]:
    """Detect smooth speed events using polygon-based site classification."""
    site_threshold, outside_threshold = validate_speed_thresholds(
        site_threshold_kmh,
        outside_threshold_kmh,
    )
    track = suppress_speed_in_exclusions(track, registry)
    site_violations: list[SpeedViolation] = []
    outside_violations: list[SpeedViolation] = []
    active: dict[str, Any] | None = None

    def close_active() -> None:
        nonlocal active
        if active is None:
            return
        points = active["points"]
        if _event_is_smooth(points):
            max_point = max(points, key=lambda point: float(point.speed))
            violation = SpeedViolation(
                plate=points[0].plate,
                start_point=points[0],
                finish_point=points[-1],
                max_point=max_point,
                point_count=len(points),
                speed_limit_kmh=active["limit"],
                threshold_kmh=active["threshold"],
                on_site=active["on_site"],
            )
            (site_violations if violation.on_site else outside_violations).append(violation)
        active = None

    for point, on_site in classify_site_state_by_polygon(track, registry):
        limit = SITE_SPEED_LIMIT_KMH if on_site else OUTSIDE_SPEED_LIMIT_KMH
        threshold = site_threshold if on_site else outside_threshold
        speed = _valid_speed(point.speed)
        exceeds = speed is not None and speed > threshold

        if active is not None:
            previous = active["points"][-1]
            gap = (point.timestamp - previous.timestamp).total_seconds()
            if (
                active["on_site"] != on_site
                or gap <= 0
                or gap > MAX_SPEED_EVENT_GAP_SECONDS
                or not exceeds
            ):
                close_active()

        if not exceeds:
            continue

        if active is None:
            active = {
                "points": [point],
                "limit": limit,
                "threshold": threshold,
                "on_site": on_site,
            }
        else:
            active["points"].append(point)

    close_active()
    return site_violations, outside_violations


def write_site_boundary_metadata(workbook_path: Path, registry: Registry) -> None:
    """Record the authoritative boundary used for site/outside classification."""
    boundary = find_site_boundary(registry)
    workbook = load_workbook(workbook_path)
    try:
        if "Параметры" not in workbook.sheetnames:
            return
        sheet = workbook["Параметры"]
        updated = False
        for row in range(1, sheet.max_row + 1):
            if sheet.cell(row, 1).value == "Разделение площадка/вне площадки":
                sheet.cell(row, 2).value = f"по полигону «{boundary.name}»"
                updated = True
                break
        if not updated:
            sheet.append(["Разделение площадка/вне площадки", f"по полигону «{boundary.name}»"])
        sheet.append(["Назначение геозоны", "site_boundary"])
        sheet.append(["Точек в полигоне площадки", len(boundary.points or [])])
        save_report_workbook(workbook, workbook_path)
    finally:
        workbook.close()
