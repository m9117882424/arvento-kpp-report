#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Persistent cache for ready-made consolidated report rows.

The cache stores the final daily vehicle metrics, archived roster snapshots and
refresh metadata. Scheduled refreshes replace complete calendar days inside one
transaction, so overlapping runs safely update late GPS and fuel data.
"""
from __future__ import annotations

import tempfile
from collections import defaultdict
from dataclasses import replace
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any, Iterable, Sequence
from zoneinfo import ZoneInfo

import psycopg
from openpyxl import Workbook, load_workbook

import consolidated_report as core
from consolidated_multi_report import load_rosters
from fuel_enriched_consolidated_report import FUEL_HEADER, add_fuel_column
from roster_registry import normalize_plate

TZ = ZoneInfo("Europe/Istanbul")
CACHE_SOURCE = "PostgreSQL consolidated_report_cache"


def day_range(start_day: date, end_day: date) -> list[date]:
    return [start_day + timedelta(days=offset) for offset in range((end_day - start_day).days + 1)]


def ensure_schema(connection: psycopg.Connection) -> None:
    statements = (
        """
        CREATE TABLE IF NOT EXISTS consolidated_roster_snapshots (
            roster_day DATE PRIMARY KEY,
            source_filename TEXT NOT NULL,
            entry_count INTEGER NOT NULL DEFAULT 0,
            loaded_at TIMESTAMPTZ NOT NULL DEFAULT now()
        )
        """,
        """
        CREATE TABLE IF NOT EXISTS consolidated_roster_entries (
            roster_day DATE NOT NULL REFERENCES consolidated_roster_snapshots(roster_day) ON DELETE CASCADE,
            normalized_plate TEXT NOT NULL,
            plate TEXT NOT NULL,
            company TEXT NOT NULL DEFAULT '',
            user_name TEXT NOT NULL DEFAULT '',
            grade TEXT NOT NULL DEFAULT '',
            PRIMARY KEY (roster_day, normalized_plate)
        )
        """,
        """
        CREATE INDEX IF NOT EXISTS ix_consolidated_roster_entries_plate
        ON consolidated_roster_entries (normalized_plate, roster_day)
        """,
        """
        CREATE TABLE IF NOT EXISTS consolidated_cache_runs (
            id BIGSERIAL PRIMARY KEY,
            period_start DATE NOT NULL,
            period_end DATE NOT NULL,
            trigger_name TEXT NOT NULL,
            status TEXT NOT NULL DEFAULT 'RUNNING',
            rows_written INTEGER NOT NULL DEFAULT 0,
            started_at TIMESTAMPTZ NOT NULL DEFAULT now(),
            finished_at TIMESTAMPTZ,
            error_message TEXT
        )
        """,
        """
        CREATE TABLE IF NOT EXISTS consolidated_report_cache (
            report_day DATE NOT NULL,
            normalized_plate TEXT NOT NULL,
            company TEXT NOT NULL DEFAULT '',
            plate TEXT NOT NULL,
            user_name TEXT NOT NULL DEFAULT '',
            grade TEXT NOT NULL DEFAULT '',
            max_speed DOUBLE PRECISION,
            route_max_speed DOUBLE PRECISION,
            site_max_speed DOUBLE PRECISION,
            total_km DOUBLE PRECISION NOT NULL DEFAULT 0,
            inside_km DOUBLE PRECISION NOT NULL DEFAULT 0,
            outside_km DOUBLE PRECISION NOT NULL DEFAULT 0,
            distance_difference_km DOUBLE PRECISION NOT NULL DEFAULT 0,
            inside_percent DOUBLE PRECISION NOT NULL DEFAULT 0,
            outside_percent DOUBLE PRECISION NOT NULL DEFAULT 0,
            percent_difference DOUBLE PRECISION NOT NULL DEFAULT 0,
            departure TIMESTAMP,
            arrival TIMESTAMP,
            weekday TEXT NOT NULL DEFAULT '',
            entry_time TIME,
            exit_time TIME,
            worked_hours DOUBLE PRECISION,
            boundary_violation INTEGER NOT NULL DEFAULT 0,
            personal_use INTEGER NOT NULL DEFAULT 0,
            weekend_work INTEGER NOT NULL DEFAULT 0,
            night_work INTEGER NOT NULL DEFAULT 0,
            fuel_liters DOUBLE PRECISION NOT NULL DEFAULT 0,
            in_roster BOOLEAN NOT NULL DEFAULT FALSE,
            raw_points INTEGER NOT NULL DEFAULT 0,
            retained_points INTEGER NOT NULL DEFAULT 0,
            valid_speed_points INTEGER NOT NULL DEFAULT 0,
            max_distance_from_site_km DOUBLE PRECISION NOT NULL DEFAULT 0,
            roster_day DATE,
            roster_filename TEXT NOT NULL DEFAULT '',
            refresh_run_id BIGINT REFERENCES consolidated_cache_runs(id),
            computed_at TIMESTAMPTZ NOT NULL DEFAULT now(),
            PRIMARY KEY (report_day, normalized_plate)
        )
        """,
        """
        ALTER TABLE consolidated_report_cache
            ALTER COLUMN worked_hours DROP NOT NULL
        """,
        """
        CREATE INDEX IF NOT EXISTS ix_consolidated_report_cache_plate_day
        ON consolidated_report_cache (normalized_plate, report_day)
        """,
        """
        CREATE INDEX IF NOT EXISTS ix_consolidated_report_cache_company_day
        ON consolidated_report_cache (company, report_day)
        """,
        """
        CREATE TABLE IF NOT EXISTS consolidated_cache_days (
            report_day DATE PRIMARY KEY,
            status TEXT NOT NULL,
            row_count INTEGER NOT NULL DEFAULT 0,
            gps_max_event_time TIMESTAMPTZ,
            refresh_run_id BIGINT REFERENCES consolidated_cache_runs(id),
            refreshed_at TIMESTAMPTZ NOT NULL DEFAULT now()
        )
        """,
    )
    with connection.cursor() as cursor:
        for statement in statements:
            cursor.execute(statement)


def save_roster_uploads(database_url: str, uploads: Sequence[tuple[str, bytes]]) -> int:
    """Parse uploaded roster workbooks and store authoritative dated snapshots."""
    if not uploads:
        return 0
    with tempfile.TemporaryDirectory(prefix="arvento_roster_archive_") as temp_name:
        temp_dir = Path(temp_name)
        paths: list[Path] = []
        original_names: dict[Path, str] = {}
        for index, (filename, content) in enumerate(uploads, start=1):
            target = temp_dir / f"{index:02d}_{Path(filename).name}"
            target.write_bytes(content)
            paths.append(target)
            original_names[target.resolve()] = Path(filename).name
        rosters = load_rosters(paths)

        with psycopg.connect(database_url) as connection:
            ensure_schema(connection)
            with connection.cursor() as cursor:
                for roster in rosters:
                    source_name = original_names.get(roster.path.resolve(), roster.path.name)
                    cursor.execute(
                        """
                        INSERT INTO consolidated_roster_snapshots(roster_day, source_filename, entry_count, loaded_at)
                        VALUES (%s,%s,%s,now())
                        ON CONFLICT (roster_day) DO UPDATE SET
                            source_filename=EXCLUDED.source_filename,
                            entry_count=EXCLUDED.entry_count,
                            loaded_at=now()
                        """,
                        (roster.day, source_name, len(roster.vehicles)),
                    )
                    cursor.execute(
                        "DELETE FROM consolidated_roster_entries WHERE roster_day=%s",
                        (roster.day,),
                    )
                    for normalized, vehicle in roster.vehicles.items():
                        cursor.execute(
                            """
                            INSERT INTO consolidated_roster_entries(
                                roster_day, normalized_plate, plate, company, user_name, grade
                            ) VALUES (%s,%s,%s,%s,%s,%s)
                            """,
                            (
                                roster.day,
                                normalized,
                                vehicle.plate,
                                vehicle.company,
                                vehicle.user,
                                vehicle.grade,
                            ),
                        )
            connection.commit()
    return len(rosters)


def export_stored_rosters(database_url: str, target_dir: Path) -> list[Path]:
    """Create temporary XLSX roster files from archived database snapshots."""
    with psycopg.connect(database_url) as connection:
        ensure_schema(connection)
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT s.roster_day, s.source_filename,
                       e.plate, e.company, e.user_name, e.grade
                FROM consolidated_roster_snapshots s
                JOIN consolidated_roster_entries e USING (roster_day)
                ORDER BY s.roster_day, e.normalized_plate
                """
            )
            rows = cursor.fetchall()

    grouped: defaultdict[date, list[tuple[str, str, str, str]]] = defaultdict(list)
    for roster_day, _source, plate, company, user_name, grade in rows:
        grouped[roster_day].append((plate, company, user_name, grade))
    if not grouped:
        raise ValueError(
            "В базе нет сохранённых разнарядок. Сначала сформируйте сводный отчёт с загруженной разнарядкой."
        )

    result: list[Path] = []
    for roster_day in sorted(grouped):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Разнарядка"
        sheet.append(["Гос рег знак", "Компания или фирма", "ПОЛЬЗОВАТЕЛЬ", "Грейд"])
        for values in grouped[roster_day]:
            sheet.append(values)
        path = target_dir / f"roster_{roster_day.isoformat()}.xlsx"
        workbook.save(path)
        workbook.close()
        result.append(path)
    return result


def _header_map(sheet) -> dict[str, int]:
    return {str(cell.value or "").strip(): cell.column for cell in sheet[1]}


def _cell(sheet, row_index: int, headers: dict[str, int], name: str, default: Any = None) -> Any:
    column = headers.get(name)
    return sheet.cell(row_index, column).value if column else default


def _as_day(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if value is None:
        return None
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _as_float(value: Any, default: float = 0.0) -> float:
    try:
        return float(value) if value is not None and value != "" else default
    except (TypeError, ValueError):
        return default


def _as_int(value: Any, default: int = 0) -> int:
    try:
        return int(value) if value is not None and value != "" else default
    except (TypeError, ValueError):
        return default


def _as_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    return str(value or "").strip().casefold() in {"1", "true", "да", "yes"}


def _diagnostics(workbook) -> dict[tuple[date, str], dict[str, Any]]:
    if "Диагностика" not in workbook.sheetnames:
        return {}
    sheet = workbook["Диагностика"]
    headers = _header_map(sheet)
    result: dict[tuple[date, str], dict[str, Any]] = {}
    for row_index in range(2, sheet.max_row + 1):
        report_day = _as_day(_cell(sheet, row_index, headers, "Дата"))
        plate = str(_cell(sheet, row_index, headers, "Госномер", "") or "")
        normalized = normalize_plate(plate)
        if report_day is None or not normalized:
            continue
        result[(report_day, normalized)] = {
            "in_roster": _as_bool(_cell(sheet, row_index, headers, "В разнарядке")),
            "raw_points": _as_int(_cell(sheet, row_index, headers, "Исходных GPS-точек")),
            "retained_points": _as_int(_cell(sheet, row_index, headers, "Точек после удаления координатных выбросов")),
            "valid_speed_points": _as_int(_cell(sheet, row_index, headers, "Точек для расчёта максимальной скорости")),
            "max_distance": _as_float(_cell(sheet, row_index, headers, "Максимальное удаление от АЭС, км")),
            "roster_day": _as_day(_cell(sheet, row_index, headers, "Дата разнарядки")),
            "roster_filename": str(_cell(sheet, row_index, headers, "Файл разнарядки", "") or ""),
        }
    return result


def upsert_cache_from_workbook(
    database_url: str,
    workbook_path: Path,
    start_day: date,
    end_day: date,
    *,
    trigger_name: str,
) -> dict[str, Any]:
    """Atomically replace cached rows for the requested calendar days."""
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        sheet = workbook["Сводный отчет"]
        headers = _header_map(sheet)
        diagnostics = _diagnostics(workbook)
        records: list[tuple[Any, ...]] = []
        counts: defaultdict[date, int] = defaultdict(int)

        for row_index in range(2, sheet.max_row + 1):
            report_day = _as_day(_cell(sheet, row_index, headers, "Дата"))
            plate = str(_cell(sheet, row_index, headers, "Госномер / Plaka", "") or "").strip()
            normalized = normalize_plate(plate)
            if report_day is None or not normalized or not (start_day <= report_day <= end_day):
                continue
            diag = diagnostics.get((report_day, normalized), {})
            records.append((
                report_day,
                normalized,
                str(_cell(sheet, row_index, headers, "Компания или фирма", "") or ""),
                plate,
                str(_cell(sheet, row_index, headers, "ПОЛЬЗОВАТЕЛЬ / KULLANICI", "") or ""),
                str(_cell(sheet, row_index, headers, "Грейд / SCALA", "") or ""),
                _cell(sheet, row_index, headers, "Максимальная скорость / Maximum Hız"),
                _cell(sheet, row_index, headers, "Максимальная скорость только на маршруте Аккую - Ташуджу / Maximum Hız Rota Akkuyu-Taşucu"),
                _cell(sheet, row_index, headers, "Максимальная скорость на АЭС Аккую"),
                _as_float(_cell(sheet, row_index, headers, "Пробег общий, км")),
                _as_float(_cell(sheet, row_index, headers, "Пробег внутри АЭС АККУЮ, км")),
                _as_float(_cell(sheet, row_index, headers, "Пробег вне площадки, км")),
                _as_float(_cell(sheet, row_index, headers, "Разница между пробегами, км")),
                _as_float(_cell(sheet, row_index, headers, "Пробег внутри АЭС АККУЮ, %")),
                _as_float(_cell(sheet, row_index, headers, "Пробег вне площадки, %")),
                _as_float(_cell(sheet, row_index, headers, "Разница между пробегами, %")),
                _cell(sheet, row_index, headers, "Дата выезда"),
                _cell(sheet, row_index, headers, "Дата прибытия"),
                str(_cell(sheet, row_index, headers, "День недели", "") or ""),
                _cell(sheet, row_index, headers, "Прибыл / Giriş"),
                _cell(sheet, row_index, headers, "Убыл / Çıkış"),
                _as_float(_cell(sheet, row_index, headers, "Всего отработано часов")),
                _as_int(_cell(sheet, row_index, headers, "Нарушение геозоны / Sınır ihlali")),
                _as_int(_cell(sheet, row_index, headers, "Использование ТС в личных целях")),
                _as_int(_cell(sheet, row_index, headers, "Hafta sonu çalışmaları / Работа в выходные дни")),
                _as_int(_cell(sheet, row_index, headers, "Gece vardiyasında çalışmalar / Работа ночью")),
                _as_float(_cell(sheet, row_index, headers, FUEL_HEADER)),
                bool(diag.get("in_roster", True)),
                int(diag.get("raw_points", 0)),
                int(diag.get("retained_points", 0)),
                int(diag.get("valid_speed_points", 0)),
                float(diag.get("max_distance", 0.0)),
                diag.get("roster_day"),
                str(diag.get("roster_filename", "")),
            ))
            counts[report_day] += 1
    finally:
        workbook.close()

    insert_sql = """
        INSERT INTO consolidated_report_cache(
            report_day, normalized_plate, company, plate, user_name, grade,
            max_speed, route_max_speed, site_max_speed,
            total_km, inside_km, outside_km, distance_difference_km,
            inside_percent, outside_percent, percent_difference,
            departure, arrival, weekday, entry_time, exit_time, worked_hours,
            boundary_violation, personal_use, weekend_work, night_work,
            fuel_liters, in_roster, raw_points, retained_points,
            valid_speed_points, max_distance_from_site_km,
            roster_day, roster_filename, refresh_run_id, computed_at
        ) VALUES (
            %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,
            %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,now()
        )
        ON CONFLICT (report_day, normalized_plate) DO UPDATE SET
            company=EXCLUDED.company,
            plate=EXCLUDED.plate,
            user_name=EXCLUDED.user_name,
            grade=EXCLUDED.grade,
            max_speed=EXCLUDED.max_speed,
            route_max_speed=EXCLUDED.route_max_speed,
            site_max_speed=EXCLUDED.site_max_speed,
            total_km=EXCLUDED.total_km,
            inside_km=EXCLUDED.inside_km,
            outside_km=EXCLUDED.outside_km,
            distance_difference_km=EXCLUDED.distance_difference_km,
            inside_percent=EXCLUDED.inside_percent,
            outside_percent=EXCLUDED.outside_percent,
            percent_difference=EXCLUDED.percent_difference,
            departure=EXCLUDED.departure,
            arrival=EXCLUDED.arrival,
            weekday=EXCLUDED.weekday,
            entry_time=EXCLUDED.entry_time,
            exit_time=EXCLUDED.exit_time,
            worked_hours=EXCLUDED.worked_hours,
            boundary_violation=EXCLUDED.boundary_violation,
            personal_use=EXCLUDED.personal_use,
            weekend_work=EXCLUDED.weekend_work,
            night_work=EXCLUDED.night_work,
            fuel_liters=EXCLUDED.fuel_liters,
            in_roster=EXCLUDED.in_roster,
            raw_points=EXCLUDED.raw_points,
            retained_points=EXCLUDED.retained_points,
            valid_speed_points=EXCLUDED.valid_speed_points,
            max_distance_from_site_km=EXCLUDED.max_distance_from_site_km,
            roster_day=EXCLUDED.roster_day,
            roster_filename=EXCLUDED.roster_filename,
            refresh_run_id=EXCLUDED.refresh_run_id,
            computed_at=now()
    """

    run_id: int | None = None
    try:
        with psycopg.connect(database_url) as connection:
            ensure_schema(connection)
            with connection.cursor() as cursor:
                cursor.execute(
                    """
                    INSERT INTO consolidated_cache_runs(period_start, period_end, trigger_name)
                    VALUES (%s,%s,%s) RETURNING id
                    """,
                    (start_day, end_day, trigger_name),
                )
                run_id = cursor.fetchone()[0]
            connection.commit()

            with connection.cursor() as cursor:
                cursor.execute(
                    "DELETE FROM consolidated_report_cache WHERE report_day BETWEEN %s AND %s",
                    (start_day, end_day),
                )
                for record in records:
                    cursor.execute(insert_sql, (*record, run_id))

                for report_day in day_range(start_day, end_day):
                    start_at = datetime.combine(report_day, time.min, TZ)
                    finish_at = start_at + timedelta(days=1)
                    cursor.execute(
                        "SELECT MAX(event_time) FROM gps_points WHERE event_time >= %s AND event_time < %s",
                        (start_at, finish_at),
                    )
                    gps_max = cursor.fetchone()[0]
                    cursor.execute(
                        """
                        INSERT INTO consolidated_cache_days(
                            report_day, status, row_count, gps_max_event_time,
                            refresh_run_id, refreshed_at
                        ) VALUES (%s,'SUCCESS',%s,%s,%s,now())
                        ON CONFLICT (report_day) DO UPDATE SET
                            status='SUCCESS',
                            row_count=EXCLUDED.row_count,
                            gps_max_event_time=EXCLUDED.gps_max_event_time,
                            refresh_run_id=EXCLUDED.refresh_run_id,
                            refreshed_at=now()
                        """,
                        (report_day, counts.get(report_day, 0), gps_max, run_id),
                    )
                cursor.execute(
                    """
                    UPDATE consolidated_cache_runs
                    SET status='SUCCESS', rows_written=%s, finished_at=now()
                    WHERE id=%s
                    """,
                    (len(records), run_id),
                )
            connection.commit()
    except Exception as exc:
        if run_id is not None:
            try:
                with psycopg.connect(database_url) as connection:
                    ensure_schema(connection)
                    with connection.cursor() as cursor:
                        cursor.execute(
                            """
                            UPDATE consolidated_cache_runs
                            SET status='FAILED', finished_at=now(), error_message=%s
                            WHERE id=%s
                            """,
                            (str(exc)[:4000], run_id),
                        )
                    connection.commit()
            except Exception:
                pass
        raise

    return {"run_id": run_id, "rows": len(records), "days": len(day_range(start_day, end_day))}


def cache_complete(database_url: str, start_day: date, end_day: date) -> bool:
    days = day_range(start_day, end_day)
    with psycopg.connect(database_url) as connection:
        ensure_schema(connection)
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT report_day
                FROM consolidated_cache_days
                WHERE report_day BETWEEN %s AND %s AND status='SUCCESS'
                """,
                (start_day, end_day),
            )
            existing = {row[0] for row in cursor.fetchall()}
    return all(day in existing for day in days)


def load_cached_rows(database_url: str, start_day: date, end_day: date) -> tuple[list[dict[str, Any]], datetime | None]:
    with psycopg.connect(database_url, row_factory=psycopg.rows.dict_row) as connection:
        ensure_schema(connection)
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT *
                FROM consolidated_report_cache
                WHERE report_day BETWEEN %s AND %s
                ORDER BY report_day, company, normalized_plate
                """,
                (start_day, end_day),
            )
            rows = list(cursor.fetchall())
            cursor.execute(
                """
                SELECT MAX(refreshed_at)
                FROM consolidated_cache_days
                WHERE report_day BETWEEN %s AND %s
                """,
                (start_day, end_day),
            )
            refreshed_at = cursor.fetchone()["max"]
    return rows, refreshed_at


def write_cached_workbook(
    database_url: str,
    output_path: Path,
    start_day: date,
    end_day: date,
    *,
    route_kml: Path = core.DEFAULT_ROUTE_KML,
    geozones: Path = core.DEFAULT_GEOZONES,
) -> dict[str, Any]:
    rows, refreshed_at = load_cached_rows(database_url, start_day, end_day)
    if not rows:
        raise ValueError("В базе готовых данных нет строк за выбранный период")

    report_rows: list[core.ReportRow] = []
    fuel_totals: dict[tuple[date, str], float] = {}
    for item in rows:
        report_rows.append(core.ReportRow(
            day=item["report_day"],
            company=item["company"],
            plate=item["plate"],
            user=item["user_name"],
            grade=item["grade"],
            max_speed=item["max_speed"],
            route_max_speed=item["route_max_speed"],
            site_max_speed=item["site_max_speed"],
            total_km=item["total_km"],
            inside_km=item["inside_km"],
            outside_km=item["outside_km"],
            distance_difference_km=item["distance_difference_km"],
            inside_percent=item["inside_percent"],
            outside_percent=item["outside_percent"],
            percent_difference=item["percent_difference"],
            departure=item["departure"],
            arrival=item["arrival"],
            weekday=item["weekday"],
            entry_time=item["entry_time"],
            exit_time=item["exit_time"],
            worked_hours=item["worked_hours"],
            boundary_violation=item["boundary_violation"],
            personal_use=item["personal_use"],
            weekend_work=item["weekend_work"],
            night_work=item["night_work"],
            in_roster=item["in_roster"],
            raw_points=item["raw_points"],
            retained_points=item["retained_points"],
            valid_speed_points=item["valid_speed_points"],
            max_distance_from_site_km=item["max_distance_from_site_km"],
        ))
        fuel_totals[(item["report_day"], item["normalized_plate"])] = item["fuel_liters"]

    registry = core.load_registry(geozones)
    site_zone = core.find_site_boundary(registry)
    core.save_report(
        output_path,
        report_rows,
        Path("Архив разнарядок PostgreSQL"),
        None,
        route_kml,
        site_zone,
        start_day,
        end_day,
        CACHE_SOURCE,
    )
    add_fuel_column(output_path, fuel_totals, configured=True)

    workbook = load_workbook(output_path)
    try:
        parameters = workbook["Параметры"]
        parameters.append([])
        parameters.append(["Режим формирования", "готовые строки из consolidated_report_cache без повторного расчёта GPS"])
        parameters.append(["Последнее обновление кэша", refreshed_at.astimezone(TZ).replace(tzinfo=None) if refreshed_at else "не определено"])
        workbook.save(output_path)
    finally:
        workbook.close()

    return {
        "rows": len(report_rows),
        "refreshed_at": refreshed_at,
        "fuel_liters": round(sum(fuel_totals.values()), 1),
    }


def recent_status(database_url: str, limit: int = 10) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    with psycopg.connect(database_url, row_factory=psycopg.rows.dict_row) as connection:
        ensure_schema(connection)
        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT * FROM consolidated_cache_days ORDER BY report_day DESC LIMIT %s",
                (limit,),
            )
            days = list(cursor.fetchall())
            cursor.execute(
                "SELECT * FROM consolidated_cache_runs ORDER BY id DESC LIMIT %s",
                (limit,),
            )
            runs = list(cursor.fetchall())
    return days, runs


__all__ = [
    "cache_complete",
    "ensure_schema",
    "export_stored_rosters",
    "load_cached_rows",
    "recent_status",
    "save_roster_uploads",
    "upsert_cache_from_workbook",
    "write_cached_workbook",
]
