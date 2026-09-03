#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Fast export path for cached consolidated reports.

The historical cached path used to reopen/save the same XLSX several times:
fuel enrichment, cache metadata, roster export, mileage-review annotation and
browser preview. This module keeps the existing workbook contract but combines
those lightweight export stages into fewer workbook I/O passes.
"""
from __future__ import annotations

from copy import copy
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any, Mapping

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import consolidated_cache as cache
import consolidated_report as core
from consolidated_export_layout import (
    EXPORT_HEADERS,
    REPORT_SHEET,
    RosterExportDetails,
    load_roster_export_details,
)
from excel_formatting import save_report_workbook
from fuel_enriched_consolidated_report import FUEL_HEADER, FUEL_SOURCE
from mileage_review_policy import (
    MileageReviewCandidate,
    REVIEW_HEADER,
    REVIEW_SHEET,
    REVIEW_VALUE,
    load_mileage_review_candidates,
)
from roster_registry import normalize_plate


PREVIEW_ROWS = 300
_PATCHED = False


def _as_day(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if value is None:
        return None
    text = str(value).strip()
    for parser in (
        lambda: date.fromisoformat(text),
        lambda: datetime.strptime(text, "%d.%m.%Y").date(),
    ):
        try:
            return parser()
        except ValueError:
            continue
    return None


def _json_cell(value: Any, *, date_only: bool = False) -> Any:
    if value is None:
        return ""
    if isinstance(value, datetime):
        if date_only:
            return value.strftime("%d.%m.%Y")
        return value.strftime("%d.%m.%Y %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%d.%m.%Y")
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    if isinstance(value, timedelta):
        return str(value)
    if isinstance(value, float):
        return round(value, 6)
    return value


def _copy_header_style(source, target) -> None:
    if source.has_style:
        target._style = copy(source._style)
    target.font = copy(source.font)
    target.fill = copy(source.fill)
    target.border = copy(source.border)
    target.alignment = copy(source.alignment)
    target.protection = copy(source.protection)
    target.number_format = source.number_format


def _add_cached_fuel_and_parameters(
    workbook,
    totals: Mapping[tuple[date, str], float],
    refreshed_at: datetime | None,
) -> tuple[float, int]:
    """Apply cached fuel values and cache metadata without an extra file reopen."""
    sheet = workbook[REPORT_SHEET]
    headers = {str(cell.value or "").strip(): cell.column for cell in sheet[1]}
    day_column = headers.get("Дата")
    plate_column = next(
        (
            headers[name]
            for name in ("Госномер / Plaka", "Госномер", "Номерной знак")
            if name in headers
        ),
        None,
    )
    if day_column is None or plate_column is None:
        raise RuntimeError("В сводном отчёте не найдены столбцы даты или госномера")

    fuel_column = headers.get(FUEL_HEADER) or (sheet.max_column + 1)
    header = sheet.cell(1, fuel_column, FUEL_HEADER)
    header.font = Font(bold=True, color="FFFFFF")
    header.fill = PatternFill("solid", fgColor="1F4E78")
    header.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.column_dimensions[get_column_letter(fuel_column)].width = 16

    total_liters = 0.0
    matched_rows = 0
    for row_index in range(2, sheet.max_row + 1):
        report_day = _as_day(sheet.cell(row_index, day_column).value)
        normalized_plate = normalize_plate(sheet.cell(row_index, plate_column).value)
        cell = sheet.cell(row_index, fuel_column)
        if report_day is None or not normalized_plate:
            cell.value = None
        else:
            value = round(float(totals.get((report_day, normalized_plate), 0.0)), 1)
            cell.value = value
            cell.number_format = "0.0"
            total_liters += value
            if value > 0:
                matched_rows += 1
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    sheet.auto_filter.ref = sheet.dimensions

    if "Параметры" in workbook.sheetnames:
        parameters = workbook["Параметры"]
        parameters.append([])
        parameters.append(["Источник заправок", FUEL_SOURCE])
        parameters.append([
            "Расчёт заправок",
            "сумма fuel_events.liters по календарной дате и нормализованному госномеру",
        ])
        parameters.append([])
        parameters.append([
            "Режим формирования",
            "готовые строки из consolidated_report_cache без повторного расчёта GPS",
        ])
        parameters.append([
            "Последнее обновление кэша",
            refreshed_at.astimezone(cache.TZ).replace(tzinfo=None)
            if refreshed_at
            else "не определено",
        ])
        for row in parameters.iter_rows(min_row=max(1, parameters.max_row - 6)):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    return round(total_liters, 1), matched_rows


def write_cached_workbook_fast(
    database_url: str,
    output_path: Path,
    start_day: date,
    end_day: date,
    *,
    route_kml: Path = core.DEFAULT_ROUTE_KML,
    geozones: Path = core.DEFAULT_GEOZONES,
) -> dict[str, Any]:
    """Write a cached workbook with one post-save enrichment pass instead of two."""
    rows, refreshed_at = cache.load_cached_rows(database_url, start_day, end_day)
    if not rows:
        raise ValueError("В базе готовых данных нет строк за выбранный период")

    report_rows: list[core.ReportRow] = []
    fuel_totals: dict[tuple[date, str], float] = {}
    for item in rows:
        report_rows.append(
            core.ReportRow(
                day=item["report_day"],
                company=item["company"],
                plate=item["plate"],
                user=item["user_name"],
                grade=item["grade"],
                max_speed=item["max_speed"],
                route_max_speed=item["route_max_speed"],
                site_max_speed=item["site_max_speed"],
                total_km=item["total_km"],
                inside_km=item["inside_km"],
                outside_km=item["outside_km"],
                distance_difference_km=item["distance_difference_km"],
                inside_percent=item["inside_percent"],
                outside_percent=item["outside_percent"],
                percent_difference=item["percent_difference"],
                departure=item["departure"],
                arrival=item["arrival"],
                weekday=item["weekday"],
                entry_time=item["entry_time"],
                exit_time=item["exit_time"],
                worked_hours=item["worked_hours"],
                boundary_violation=item["boundary_violation"],
                personal_use=item["personal_use"],
                weekend_work=item["weekend_work"],
                night_work=item["night_work"],
                in_roster=item["in_roster"],
                raw_points=item["raw_points"],
                retained_points=item["retained_points"],
                valid_speed_points=item["valid_speed_points"],
                max_distance_from_site_km=item["max_distance_from_site_km"],
            )
        )
        fuel_totals[(item["report_day"], item["normalized_plate"])] = float(
            item["fuel_liters"] or 0.0
        )

    registry = core.load_registry(geozones)
    site_zone = core.find_site_boundary(registry)
    core.save_report(
        output_path,
        report_rows,
        Path("Архив разнарядок PostgreSQL"),
        None,
        route_kml,
        site_zone,
        start_day,
        end_day,
        cache.CACHE_SOURCE,
    )

    workbook = load_workbook(output_path)
    try:
        _add_cached_fuel_and_parameters(workbook, fuel_totals, refreshed_at)
        save_report_workbook(workbook, output_path)
    finally:
        workbook.close()

    return {
        "rows": len(report_rows),
        "refreshed_at": refreshed_at,
        "fuel_liters": round(sum(fuel_totals.values()), 1),
    }


def _apply_roster_layout(
    workbook,
    roster_details: Mapping[tuple[date, str], RosterExportDetails],
) -> dict[str, int]:
    """Apply the established roster export contract to an already-open workbook."""
    if REPORT_SHEET not in workbook.sheetnames:
        raise RuntimeError(f"В книге отсутствует лист «{REPORT_SHEET}»")
    sheet = workbook[REPORT_SHEET]
    headers = {str(cell.value or "").strip(): cell.column for cell in sheet[1]}
    day_column = headers.get("Дата")
    plate_column = next(
        (
            headers[name]
            for name in ("Госномер / Plaka", "Госномер", "Номерной знак")
            if name in headers
        ),
        None,
    )
    source_driver_column = next(
        (
            headers[name]
            for name in ("ПОЛЬЗОВАТЕЛЬ / KULLANICI", "Пользователь", "Водитель")
            if name in headers
        ),
        None,
    )
    source_grade_column = next(
        (
            headers[name]
            for name in ("Грейд / SCALA", "Грейд", "SCALA", "Grade")
            if name in headers
        ),
        None,
    )
    if day_column is None or plate_column is None:
        raise RuntimeError("В сводном отчёте не найдены столбцы даты или госномера")

    original_last_column = sheet.max_column
    header_template = sheet.cell(1, original_last_column)
    export_columns: dict[str, int] = {}
    next_column = original_last_column + 1
    for header_name in EXPORT_HEADERS:
        existing = headers.get(header_name)
        column = existing or next_column
        if existing is None:
            next_column += 1
        export_columns[header_name] = column
        cell = sheet.cell(1, column, header_name)
        _copy_header_style(header_template, cell)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = {
        "Имя водителя": 38,
        "Грейд": 12,
        "Дирекция": 52,
        "Ответственный": 36,
    }
    enriched_rows = 0
    for row_index in range(2, sheet.max_row + 1):
        report_day = _as_day(sheet.cell(row_index, day_column).value)
        normalized_plate = normalize_plate(sheet.cell(row_index, plate_column).value)
        details = (
            roster_details.get((report_day, normalized_plate))
            if report_day is not None and normalized_plate
            else None
        )
        fallback_driver = (
            str(sheet.cell(row_index, source_driver_column).value or "")
            if source_driver_column is not None
            else ""
        )
        fallback_grade = (
            str(sheet.cell(row_index, source_grade_column).value or "")
            if source_grade_column is not None
            else ""
        )
        values = {
            "Имя водителя": details.driver_name
            if details and details.driver_name
            else fallback_driver,
            "Грейд": details.grade if details and details.grade else fallback_grade,
            "Дирекция": details.directorate if details else "",
            "Ответственный": details.responsible if details else "",
        }
        if details is not None:
            enriched_rows += 1
        for header_name, value in values.items():
            cell = sheet.cell(row_index, export_columns[header_name], value)
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    for header_name, column in export_columns.items():
        sheet.column_dimensions[get_column_letter(column)].width = widths[header_name]
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.showGridLines = False

    for sheet_name in list(workbook.sheetnames):
        if sheet_name != REPORT_SHEET:
            del workbook[sheet_name]
    workbook.active = 0
    return {
        "rows": max(0, sheet.max_row - 1),
        "enriched_rows": enriched_rows,
        "sheets": 1,
        "columns": sheet.max_column,
    }


def _apply_mileage_review(
    workbook,
    candidates: Mapping[tuple[date, str], MileageReviewCandidate],
) -> dict[str, int]:
    """Apply the established mileage-review sheet without reopening the file."""
    sheet = workbook[REPORT_SHEET]
    headers = {str(cell.value or "").strip(): cell.column for cell in sheet[1]}
    day_column = headers.get("Дата")
    plate_column = next(
        (
            headers[name]
            for name in ("Госномер / Plaka", "Госномер", "Номерной знак")
            if name in headers
        ),
        None,
    )
    company_column = next(
        (
            headers[name]
            for name in ("Компания или фирма", "Компания", "Фирма")
            if name in headers
        ),
        None,
    )
    driver_column = next(
        (
            headers[name]
            for name in (
                "Имя водителя",
                "ПОЛЬЗОВАТЕЛЬ / KULLANICI",
                "Пользователь",
                "Водитель",
            )
            if name in headers
        ),
        None,
    )
    if day_column is None or plate_column is None:
        raise RuntimeError("В сводном отчёте не найдены столбцы даты или госномера")

    review_column = headers.get(REVIEW_HEADER) or (sheet.max_column + 1)
    header = sheet.cell(1, review_column, REVIEW_HEADER)
    template = sheet.cell(1, max(1, review_column - 1))
    _copy_header_style(template, header)
    header.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.column_dimensions[get_column_letter(review_column)].width = 18

    row_details: dict[tuple[date, str], tuple[str, str, str]] = {}
    flagged_rows = 0
    review_fill = PatternFill("solid", fgColor="FDE68A")
    review_font = Font(bold=True, color="9C0006")
    for row_index in range(2, sheet.max_row + 1):
        report_day = _as_day(sheet.cell(row_index, day_column).value)
        normalized = normalize_plate(sheet.cell(row_index, plate_column).value)
        if report_day is None or not normalized:
            continue
        key = (report_day, normalized)
        display_plate = str(sheet.cell(row_index, plate_column).value or normalized)
        company = (
            str(sheet.cell(row_index, company_column).value or "")
            if company_column is not None
            else ""
        )
        driver = (
            str(sheet.cell(row_index, driver_column).value or "")
            if driver_column is not None
            else ""
        )
        row_details[key] = (display_plate, company, driver)
        cell = sheet.cell(row_index, review_column)
        if key in candidates:
            cell.value = REVIEW_VALUE
            cell.fill = review_fill
            cell.font = review_font
            flagged_rows += 1
        else:
            cell.value = None
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    if REVIEW_SHEET in workbook.sheetnames:
        del workbook[REVIEW_SHEET]
    review = workbook.create_sheet(REVIEW_SHEET, 1)
    review.append([
        "Дата",
        "Госномер",
        "Компания",
        "Водитель",
        "Пробег VehicleDistanceReport, км",
        "Пробег по координатам, км",
        "Разница, км",
        "Расхождение, %",
        "Статус",
        "Причина",
    ])
    for item in sorted(
        candidates.values(),
        key=lambda value: (value.report_day, value.normalized_plate),
    ):
        display_plate, company, driver = row_details.get(
            (item.report_day, item.normalized_plate),
            (item.normalized_plate, "", ""),
        )
        review.append([
            item.report_day,
            display_plate,
            company,
            driver,
            round(item.authoritative_km, 3),
            round(item.coordinate_km, 3),
            round(item.gap_km, 3),
            round(item.gap_percent, 1) if item.gap_percent is not None else None,
            REVIEW_VALUE,
            item.reason,
        ])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in review[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    review.freeze_panes = "A2"
    review.auto_filter.ref = review.dimensions
    review.sheet_view.showGridLines = False
    review.row_dimensions[1].height = 48
    widths = (12, 18, 26, 36, 28, 24, 16, 18, 16, 48)
    for index, width in enumerate(widths, 1):
        review.column_dimensions[get_column_letter(index)].width = width
    for row in review.iter_rows(min_row=2):
        row[0].number_format = "dd.mm.yyyy"
        for index in (4, 5, 6):
            row[index].number_format = "0.000"
        row[7].number_format = "0.0"
        row[8].fill = review_fill
        row[8].font = review_font
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    sheet.auto_filter.ref = sheet.dimensions
    return {
        "candidates": len(candidates),
        "flagged_rows": flagged_rows,
        "sheets": len(workbook.sheetnames),
    }


def _preview_loaded(workbook) -> tuple[list[str], list[list[Any]], int]:
    sheet = workbook.worksheets[0]
    iterator = sheet.iter_rows(values_only=True)
    columns = [str(value or "") for value in next(iterator, tuple())]
    date_indexes = {
        index for index, name in enumerate(columns) if name.strip() == "Дата"
    }
    rows: list[list[Any]] = []
    total = 0
    for row in iterator:
        if not any(value not in (None, "") for value in row):
            continue
        total += 1
        if len(rows) < PREVIEW_ROWS:
            rows.append([
                _json_cell(value, date_only=index in date_indexes)
                for index, value in enumerate(row)
            ])
    return columns, rows, total


def finalize_consolidated_output_fast(
    output_path: Path,
    database_url: str,
    start_day: date,
    end_day: date,
) -> dict[str, Any]:
    """Finalize roster, mileage review, save and preview in one workbook pass."""
    report_days = [
        start_day + timedelta(days=offset)
        for offset in range((end_day - start_day).days + 1)
    ]
    roster_details = load_roster_export_details(database_url, report_days)
    candidates = load_mileage_review_candidates(database_url, start_day, end_day)

    workbook = load_workbook(output_path)
    try:
        layout_stats = _apply_roster_layout(workbook, roster_details)
        review_stats = _apply_mileage_review(workbook, candidates)
        save_report_workbook(workbook, output_path)
        columns, rows, total_rows = _preview_loaded(workbook)
    finally:
        workbook.close()

    return {
        "layout": layout_stats,
        "review": review_stats,
        "columns": columns,
        "rows": rows,
        "total_rows": total_rows,
    }


def apply_cached_workbook_optimization() -> None:
    """Install the faster cached workbook writer used by the portal cache path."""
    global _PATCHED
    if _PATCHED:
        return
    import consolidated_cache_portal as cache_portal

    cache_portal.write_cached_workbook = write_cached_workbook_fast
    _PATCHED = True


__all__ = [
    "apply_cached_workbook_optimization",
    "finalize_consolidated_output_fast",
    "write_cached_workbook_fast",
]
