#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from __future__ import annotations

import base64
import csv
import os
import subprocess
import sys
import tempfile
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

import psycopg
from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.responses import HTMLResponse
from openpyxl import load_workbook

from business_rules import (
    DEFAULT_OUTSIDE_SPEED_THRESHOLD_KMH,
    DEFAULT_SITE_SPEED_THRESHOLD_KMH,
)

APP_DIR = Path(__file__).resolve().parent
TZ = ZoneInfo("Europe/Istanbul")
MAX_ROSTER_BYTES = 25 * 1024 * 1024
MAX_REPORT_DAYS = 31
PREVIEW_ROWS = 300

CANONICAL_REPORT_SCRIPTS = {
    "kpp": APP_DIR / "generate_first_entry_report.py",
    "efficiency": APP_DIR / "generate_kpp_summary_report.py",
    "violation": APP_DIR / "generate_prohibited_left_turn_report.py",
}

app = FastAPI(title="Arvento Report Portal")

HTML = r"""<!doctype html>
<html lang="ru">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Отчёты Arvento</title>
  <style>
    :root { color-scheme: light; font-family: Inter, Segoe UI, Arial, sans-serif; }
    body { margin: 0; background: #f3f5f8; color: #172033; }
    .wrap { max-width: 1500px; margin: 0 auto; padding: 24px; }
    .card { background: #fff; border: 1px solid #dfe4ec; border-radius: 14px; box-shadow: 0 4px 18px rgba(31,45,61,.06); padding: 20px; margin-bottom: 18px; }
    h1 { margin: 0 0 6px; font-size: 26px; }
    .muted, .note { color: #667085; font-size: 14px; }
    .grid { display: grid; grid-template-columns: repeat(4, minmax(180px, 1fr)); gap: 14px; margin-top: 18px; }
    label { display: block; font-size: 13px; font-weight: 650; margin-bottom: 6px; }
    input, select, button { width: 100%; box-sizing: border-box; border-radius: 9px; border: 1px solid #cfd6e1; padding: 10px 11px; font: inherit; background: #fff; }
    input[type=checkbox] { width: auto; margin-right: 8px; }
    .check { display: flex; align-items: center; min-height: 42px; font-weight: 500; }
    .actions { display: flex; gap: 12px; margin-top: 18px; }
    button { cursor: pointer; border: 0; font-weight: 700; }
    .primary { background: #1663d6; color: #fff; max-width: 280px; }
    .secondary { background: #e9eef6; color: #172033; max-width: 280px; }
    button:disabled { opacity: .55; cursor: wait; }
    .status { margin-top: 14px; min-height: 22px; font-size: 14px; white-space: pre-wrap; }
    .error { color: #b42318; }
    .ok { color: #067647; }
    .stats { display: grid; grid-template-columns: repeat(4, minmax(150px, 1fr)); gap: 12px; margin-bottom: 16px; }
    .stat { border: 1px solid #e2e7ef; background: #f8fafc; border-radius: 10px; padding: 12px; }
    .stat .k { color: #667085; font-size: 12px; margin-bottom: 5px; }
    .stat .v { font-size: 20px; font-weight: 750; }
    .table-wrap { overflow: auto; max-height: 62vh; border: 1px solid #dfe4ec; border-radius: 10px; }
    table { border-collapse: separate; border-spacing: 0; width: 100%; font-size: 13px; }
    th, td { padding: 8px 10px; border-right: 1px solid #e5e9f0; border-bottom: 1px solid #e5e9f0; white-space: nowrap; text-align: left; }
    th { position: sticky; top: 0; z-index: 1; background: #eef3f9; font-weight: 750; }
    .hidden { display: none !important; }
    .note { margin-top: 10px; }
    @media (max-width: 900px) { .grid, .stats { grid-template-columns: 1fr 1fr; } }
    @media (max-width: 560px) { .grid, .stats { grid-template-columns: 1fr; } .actions { flex-direction: column; } .primary, .secondary { max-width: none; } }
  </style>
</head>
<body>
<div class="wrap">
  <div class="card">
    <h1>Отчёты Arvento</h1>
    <div class="muted">Данные берутся напрямую из PostgreSQL. Временные CSV и Excel после формирования удаляются.</div>
    <form id="reportForm">
      <div class="grid">
        <div>
          <label for="reportType">Отчёт</label>
          <select id="reportType" name="report_type">
            <option value="kpp">Первый въезд через КПП</option>
            <option value="efficiency">Эффективность легкового транспорта</option>
            <option value="violation">Запрещённый поворот</option>
          </select>
        </div>
        <div>
          <label id="dateLabel" for="reportDate">Дата</label>
          <input id="reportDate" name="report_date" type="date" required>
        </div>
        <div class="efficiency-only hidden">
          <label for="reportEndDate">Дата до</label>
          <input id="reportEndDate" name="report_end_date" type="date">
        </div>
        <div id="rosterBox" style="grid-column: span 2">
          <label for="roster">Разнарядка XLSX/XLSM</label>
          <input id="roster" name="roster" type="file" accept=".xlsx,.xlsm">
        </div>
        <div class="kpp-only"><label for="gradeFrom">Грейд от</label><input id="gradeFrom" name="grade_from" value="7" placeholder="без фильтра"></div>
        <div class="kpp-only"><label for="gradeTo">Грейд до</label><input id="gradeTo" name="grade_to" value="14" placeholder="без фильтра"></div>
        <div class="kpp-only"><label for="timeFrom">Время от</label><input id="timeFrom" name="time_from" type="time" value="07:00"></div>
        <div class="kpp-only"><label for="timeTo">Время до</label><input id="timeTo" name="time_to" type="time" value="09:00"></div>
        <div class="kpp-only" style="grid-column: span 2">
          <label class="check"><input name="consider_previous_exits" type="checkbox" value="true" checked>Учитывать предыдущие выезды</label>
        </div>
      </div>
      <div class="actions">
        <button id="generateBtn" class="primary" type="submit">Сформировать отчёт</button>
        <button id="downloadBtn" class="secondary hidden" type="button">Выгрузить Excel</button>
      </div>
      <div id="status" class="status"></div>
    </form>
  </div>
  <div id="resultCard" class="card hidden">
    <div id="stats" class="stats"></div>
    <div class="table-wrap"><table id="resultTable"></table></div>
    <div id="previewNote" class="note"></div>
  </div>
</div>
<script>
const form = document.getElementById('reportForm');
const typeSelect = document.getElementById('reportType');
const dateInput = document.getElementById('reportDate');
const endDateInput = document.getElementById('reportEndDate');
const dateLabel = document.getElementById('dateLabel');
const roster = document.getElementById('roster');
const rosterBox = document.getElementById('rosterBox');
const generateBtn = document.getElementById('generateBtn');
const downloadBtn = document.getElementById('downloadBtn');
const statusBox = document.getElementById('status');
const resultCard = document.getElementById('resultCard');
const statsBox = document.getElementById('stats');
const table = document.getElementById('resultTable');
const previewNote = document.getElementById('previewNote');
let excelBase64 = '';
let excelFilename = '';
let downloadUrl = '';

function esc(value) {
  return String(value ?? '').replace(/[&<>'"]/g, ch => ({'&':'&amp;','<':'&lt;','>':'&gt;',"'":'&#39;','"':'&quot;'}[ch]));
}
function toggleType() {
  const type = typeSelect.value;
  const isKpp = type === 'kpp';
  const isEfficiency = type === 'efficiency';
  const needsRoster = isKpp || isEfficiency;
  rosterBox.classList.toggle('hidden', !needsRoster);
  roster.required = needsRoster;
  document.querySelectorAll('.kpp-only').forEach(x => x.classList.toggle('hidden', !isKpp));
  document.querySelectorAll('.efficiency-only').forEach(x => x.classList.toggle('hidden', !isEfficiency));
  dateLabel.textContent = isEfficiency ? 'Дата от' : 'Дата';
  endDateInput.required = isEfficiency;
  if (isEfficiency && !endDateInput.value) endDateInput.value = dateInput.value;
}
typeSelect.addEventListener('change', toggleType);
dateInput.addEventListener('change', () => {
  if (typeSelect.value === 'efficiency' && (!endDateInput.value || endDateInput.value < dateInput.value)) endDateInput.value = dateInput.value;
});
toggleType();

async function loadDates() {
  try {
    const response = await fetch('/api/dates');
    const data = await response.json();
    if (data.dates && data.dates.length) {
      const latest = data.dates[0];
      const earliest = data.dates[data.dates.length - 1];
      for (const input of [dateInput, endDateInput]) {
        input.min = earliest;
        input.max = latest;
        input.value = latest;
      }
    }
  } catch (_) {}
}
loadDates();

function renderStats(summary) {
  statsBox.innerHTML = Object.entries(summary).map(([key, value]) => `<div class="stat"><div class="k">${esc(key)}</div><div class="v">${esc(value)}</div></div>`).join('');
}
function renderTable(columns, rows) {
  const head = `<thead><tr>${columns.map(value => `<th>${esc(value)}</th>`).join('')}</tr></thead>`;
  const body = `<tbody>${rows.map(row => `<tr>${row.map(value => `<td>${esc(value)}</td>`).join('')}</tr>`).join('')}</tbody>`;
  table.innerHTML = head + body;
}
async function readResponsePayload(response) {
  const text = await response.text();
  if (!text) return {};
  try {
    return JSON.parse(text);
  } catch (_) {
    return {detail: text};
  }
}
form.addEventListener('submit', async event => {
  event.preventDefault();
  excelBase64 = '';
  excelFilename = '';
  downloadUrl = '';
  downloadBtn.classList.add('hidden');
  resultCard.classList.add('hidden');
  statusBox.className = 'status';
  statusBox.textContent = 'Формирование отчёта. Это может занять несколько минут…';
  generateBtn.disabled = true;
  const data = new FormData(form);
  if (!data.has('consider_previous_exits')) data.append('consider_previous_exits', 'false');
  try {
    const response = await fetch('/api/generate', {method: 'POST', body: data});
    const payload = await readResponsePayload(response);
    if (!response.ok) throw new Error(payload.detail || 'Ошибка формирования отчёта');
    excelBase64 = payload.excel_base64 || '';
    excelFilename = payload.filename;
    downloadUrl = payload.download_url || '';
    renderStats(payload.summary);
    renderTable(payload.columns, payload.rows);
    previewNote.textContent = payload.preview_truncated ? `Показаны первые ${payload.rows.length} строк. Полный результат находится в Excel.` : `Показано строк: ${payload.rows.length}.`;
    resultCard.classList.remove('hidden');
    downloadBtn.classList.remove('hidden');
    statusBox.className = 'status ok';
    statusBox.textContent = 'Отчёт сформирован.';
  } catch (error) {
    statusBox.className = 'status error';
    statusBox.textContent = error.message;
  } finally {
    generateBtn.disabled = false;
  }
});
downloadBtn.addEventListener('click', () => {
  if (downloadUrl) {
    const link = document.createElement('a');
    link.href = downloadUrl;
    link.download = excelFilename;
    document.body.appendChild(link);
    link.click();
    link.remove();
    return;
  }
  if (!excelBase64) return;
  const raw = atob(excelBase64);
  const bytes = new Uint8Array(raw.length);
  for (let index = 0; index < raw.length; index++) bytes[index] = raw.charCodeAt(index);
  const blob = new Blob([bytes], {type: 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'});
  const url = URL.createObjectURL(blob);
  const link = document.createElement('a');
  link.href = url;
  link.download = excelFilename;
  document.body.appendChild(link);
  link.click();
  link.remove();
  URL.revokeObjectURL(url);
});
</script>
</body>
</html>"""


def db_url() -> str:
    value = os.environ.get("DATABASE_URL", "").strip()
    if not value:
        raise RuntimeError("DATABASE_URL is not configured")
    return value


def parse_report_date(value: str) -> date:
    try:
        return date.fromisoformat(value)
    except ValueError as exc:
        raise ValueError("Некорректная дата") from exc


def export_period_to_csv(start_day: date, end_day: date, path: Path) -> int:
    start = datetime.combine(start_day, time.min, tzinfo=TZ)
    end = datetime.combine(end_day + timedelta(days=1), time.min, tzinfo=TZ)
    query = """
        SELECT
            COALESCE(NULLIF(plate, ''), normalized_plate),
            event_time AT TIME ZONE 'Europe/Istanbul',
            latitude,
            longitude,
            speed_kmh,
            distance_km,
            COALESCE(region_name, ''),
            COALESCE(address, '')
        FROM gps_points
        WHERE event_time >= %s AND event_time < %s
        ORDER BY normalized_plate, event_time
    """
    count = 0
    with psycopg.connect(db_url()) as connection:
        with connection.cursor(name="report_portal_export") as cursor:
            cursor.itersize = 10_000
            cursor.execute(query, (start, end))
            with path.open("w", encoding="utf-8-sig", newline="") as stream:
                writer = csv.writer(stream, delimiter=";")
                writer.writerow(["Номерной знак", "Дата / время", "Latitudine", "Долгота", "Скорость", "Расстояние (км)", "Область", "Адрес"])
                for row in cursor:
                    local_time = row[1]
                    writer.writerow([
                        row[0] or "",
                        local_time.strftime("%Y-%m-%d %H:%M:%S") if local_time else "",
                        row[2], row[3], row[4], row[5], row[6], row[7],
                    ])
                    count += 1
    return count


def run_command(command: list[str]) -> str:
    completed = subprocess.run(
        command,
        cwd=APP_DIR,
        capture_output=True,
        text=True,
        timeout=30 * 60,
        check=False,
    )
    output = "\n".join(part for part in (completed.stdout, completed.stderr) if part).strip()
    if completed.returncode != 0:
        raise RuntimeError(output[-6000:] or f"Команда завершилась с кодом {completed.returncode}")
    return output


def json_cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%d.%m.%Y")
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    if isinstance(value, timedelta):
        return str(value)
    if isinstance(value, float):
        return round(value, 6)
    return value


def workbook_preview(path: Path) -> tuple[list[str], list[list[Any]], int]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.worksheets[0]
        iterator = sheet.iter_rows(values_only=True)
        columns = [str(value or "") for value in next(iterator, tuple())]
        rows: list[list[Any]] = []
        total = 0
        for row in iterator:
            if not any(value not in (None, "") for value in row):
                continue
            total += 1
            if len(rows) < PREVIEW_ROWS:
                rows.append([json_cell(value) for value in row])
        return columns, rows, total
    finally:
        workbook.close()


def validate_canonical_scripts() -> None:
    missing = [path.name for path in CANONICAL_REPORT_SCRIPTS.values() if not path.is_file()]
    if missing:
        raise RuntimeError("Не найдены канонические скрипты отчётов: " + ", ".join(missing))


def roster_target(temp_dir: Path, filename: str, suffix: str) -> Path:
    safe_name = Path(filename or f"roster{suffix}").name
    if Path(safe_name).suffix.lower() not in {".xlsx", ".xlsm"}:
        safe_name = f"roster{suffix}"
    return temp_dir / safe_name


def generate_report(
    report_type: str,
    report_day: date,
    report_end_day: date | None,
    roster_bytes: bytes | None,
    roster_filename: str,
    roster_suffix: str,
    grade_from: str,
    grade_to: str,
    time_from: str,
    time_to: str,
    consider_previous_exits: bool,
) -> dict[str, Any]:
    validate_canonical_scripts()
    if report_type not in CANONICAL_REPORT_SCRIPTS:
        raise ValueError("Неизвестный тип отчёта")

    end_day = report_end_day if report_type == "efficiency" and report_end_day else report_day
    if end_day < report_day:
        raise ValueError("Дата окончания раньше даты начала")
    if (end_day - report_day).days + 1 > MAX_REPORT_DAYS:
        raise ValueError(f"Период отчёта не должен превышать {MAX_REPORT_DAYS} дней")
    if report_type != "efficiency" and end_day != report_day:
        end_day = report_day

    with tempfile.TemporaryDirectory(prefix="arvento_report_portal_") as temp_name:
        temp_dir = Path(temp_name)
        csv_path = temp_dir / f"gps_{report_day.isoformat()}_{end_day.isoformat()}.csv"
        gps_count = export_period_to_csv(report_day, end_day, csv_path)
        if gps_count == 0:
            raise ValueError("За выбранный период GPS-точки отсутствуют")

        if report_type == "violation":
            filename = f"Запрещенный_поворот_{report_day.isoformat()}.xlsx"
            output_path = temp_dir / filename
            log = run_command([sys.executable, str(CANONICAL_REPORT_SCRIPTS[report_type]), str(csv_path), str(output_path)])
            report_label = "Запрещённый поворот"

        elif report_type == "kpp":
            if not roster_bytes:
                raise ValueError("Для отчёта по КПП необходимо загрузить разнарядку")
            roster_path = roster_target(temp_dir, roster_filename, roster_suffix)
            roster_path.write_bytes(roster_bytes)
            filename = f"Первый_въезд_{report_day.isoformat()}.xlsx"
            output_path = temp_dir / filename
            command = [
                sys.executable, str(CANONICAL_REPORT_SCRIPTS[report_type]),
                str(csv_path), str(roster_path), str(output_path), "--no-filter-dialog",
            ]
            for option, value in (
                ("--grade-from", grade_from),
                ("--grade-to", grade_to),
                ("--time-from", time_from),
                ("--time-to", time_to),
            ):
                if value.strip():
                    command += [option, value.strip()]
            command.append("--consider-previous-exits" if consider_previous_exits else "--no-consider-previous-exits")
            log = run_command(command)
            report_label = "Первый въезд через КПП"

        else:
            if not roster_bytes:
                raise ValueError("Для отчёта эффективности необходимо загрузить разнарядку")
            roster_path = roster_target(temp_dir, roster_filename, roster_suffix)
            roster_path.write_bytes(roster_bytes)
            log = run_command([sys.executable, str(CANONICAL_REPORT_SCRIPTS[report_type]), str(csv_path)])
            output_path = csv_path.with_name(csv_path.stem + "_итоговая_сводка.xlsx")
            filename = f"Эффективность_легкового_транспорта_{report_day.isoformat()}_{end_day.isoformat()}.xlsx"
            report_label = "Эффективность легкового транспорта"

        if not output_path.exists():
            raise RuntimeError("Построитель не создал Excel-файл")

        columns, rows, total_rows = workbook_preview(output_path)
        period_text = report_day.strftime("%d.%m.%Y") if report_day == end_day else f"{report_day:%d.%m.%Y}–{end_day:%d.%m.%Y}"
        return {
            "filename": filename,
            "columns": columns,
            "rows": rows,
            "preview_truncated": total_rows > len(rows),
            "excel_base64": base64.b64encode(output_path.read_bytes()).decode("ascii"),
            "summary": {
                "Отчёт": report_label,
                "Период": period_text,
                "GPS-точек": gps_count,
                "Строк результата": total_rows,
            },
            "log": log,
        }


@app.get("/", response_class=HTMLResponse)
def index() -> str:
    return HTML


@app.get("/health")
def health() -> dict[str, str]:
    try:
        validate_canonical_scripts()
        with psycopg.connect(db_url(), connect_timeout=5) as connection:
            connection.execute("SELECT 1").fetchone()
    except Exception as exc:
        raise HTTPException(
            status_code=503,
            detail="Report portal is not ready",
        ) from exc
    return {"status": "ok", "database": "ok"}


@app.get("/api/dates")
def available_dates() -> dict[str, list[str]]:
    query = """
        SELECT DISTINCT (event_time AT TIME ZONE 'Europe/Istanbul')::date AS day
        FROM gps_points
        ORDER BY day DESC
        LIMIT 31
    """
    with psycopg.connect(db_url()) as connection:
        rows = connection.execute(query).fetchall()
    return {"dates": [row[0].isoformat() for row in rows if row[0] is not None]}


@app.post("/api/generate", deprecated=True)
async def api_generate(
    report_type: str = Form(...),
    report_date: str = Form(...),
    report_end_date: str = Form(default=""),
    roster: UploadFile | None = File(default=None),
    grade_from: str = Form(default=""),
    grade_to: str = Form(default=""),
    time_from: str = Form(default=""),
    time_to: str = Form(default=""),
    consider_previous_exits: bool = Form(default=False),
) -> dict[str, Any]:
    """Backward-compatible alias for the current v3 generation contract."""
    # Import at request time to avoid a circular import while the ASGI layers
    # are assembled. Production startup has already initialized v3 and all of
    # its central-roster/cache/final-export wrappers before requests arrive.
    import consolidated_portal as current_portal

    result = await current_portal.api_generate_v3(
        report_type=report_type,
        report_date=report_date,
        report_end_date=report_end_date,
        roster=roster,
        rosters=[roster] if roster is not None and roster.filename else None,
        grade_from=grade_from,
        grade_to=grade_to,
        time_from=time_from,
        time_to=time_to,
        consider_previous_exits=consider_previous_exits,
        site_speed_threshold=str(DEFAULT_SITE_SPEED_THRESHOLD_KMH),
        outside_speed_threshold=str(DEFAULT_OUTSIDE_SPEED_THRESHOLD_KMH),
    )
    from download_store import restore_legacy_base64

    return restore_legacy_base64(result)
