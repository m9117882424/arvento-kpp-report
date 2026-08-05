from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Iterator, Optional, TextIO

from openpyxl import load_workbook


ALIASES = {
    "plate": ["номерной знак", "госномер", "plaka", "license plate"],
    "time": ["дата / время", "дата/время", "tarih / saat", "date / time"],
    "odometer": ["одометр", "odometer"],
    "distance": ["расстояние (км)", "расстояние", "distance"],
    "speed": ["скорость", "speed"],
    "region": ["область", "регион", "region"],
    "address": ["адрес", "address"],
    "lat": ["latitudine", "широта", "latitude", "enlem"],
    "lon": ["долгота", "longitude", "boylam"],
}


@dataclass(slots=True)
class Point:
    plate: str
    time: datetime
    lat: float
    lon: float
    odometer: Optional[float] = None
    source_distance: Optional[float] = None
    prepared_distance: bool = False
    speed: Optional[float] = None
    region: str = ""
    address: str = ""


def norm(value: Any) -> str:
    return " ".join(str(value or "").replace("\n", " ").split()).strip().lower().replace("ё", "е")


def as_float(value: Any) -> Optional[float]:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip().replace(" ", "").replace(",", "."))
    except ValueError:
        return None


def as_datetime(value: Any) -> Optional[datetime]:
    if isinstance(value, datetime):
        return value
    if value in (None, ""):
        return None
    text = str(value).strip()
    for fmt in (
        "%d.%m.%Y %H:%M:%S",
        "%d.%m.%Y %H:%M",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%dT%H:%M:%S",
        "%m/%d/%Y %H:%M:%S",
    ):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def detect_columns(headers: list[Any]) -> dict[str, int]:
    normalized = [norm(value) for value in headers]
    result: dict[str, int] = {}
    for key, aliases in ALIASES.items():
        for index, header in enumerate(normalized):
            if any(norm(alias) == header or norm(alias) in header for alias in aliases):
                result[key] = index
                break
    return result


def is_header_row(row: Iterable[Any]) -> bool:
    columns = detect_columns(list(row))
    return all(key in columns for key in ("plate", "time", "lat", "lon"))


def row_to_point(row: list[Any], columns: dict[str, int]) -> Optional[Point]:
    if not row or len(row) <= max(columns.values()):
        return None

    plate = str(row[columns["plate"]] or "").strip()
    timestamp = as_datetime(row[columns["time"]])
    lat = as_float(row[columns["lat"]])
    lon = as_float(row[columns["lon"]])
    if not plate or timestamp is None or lat is None or lon is None:
        return None

    return Point(
        plate=plate,
        time=timestamp,
        lat=lat,
        lon=lon,
        odometer=as_float(row[columns["odometer"]]) if "odometer" in columns else None,
        source_distance=as_float(row[columns["distance"]]) if "distance" in columns else None,
        speed=as_float(row[columns["speed"]]) if "speed" in columns else None,
        region=str(row[columns["region"]] or "").strip() if "region" in columns else "",
        address=str(row[columns["address"]] or "").strip() if "address" in columns else "",
    )


def rows_to_points(headers: list[Any], rows: Iterable[Iterable[Any]]) -> tuple[list[Point], int]:
    columns = detect_columns(headers)
    missing = [key for key in ("plate", "time", "lat", "lon") if key not in columns]
    if missing:
        raise ValueError("Не найдены обязательные колонки: " + ", ".join(missing))

    points: list[Point] = []
    skipped = 0
    for source_row in rows:
        point = row_to_point(list(source_row), columns)
        if point is None:
            skipped += 1
        else:
            points.append(point)
    return points, skipped


def load_excel(path: Path) -> tuple[list[Point], dict[str, int]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    rows = sheet.iter_rows(values_only=True)

    headers: Optional[list[Any]] = None
    for index, row in enumerate(rows):
        if index >= 100:
            break
        candidate = list(row)
        if is_header_row(candidate):
            headers = candidate
            break

    if headers is None:
        workbook.close()
        raise ValueError("Не найдена строка заголовков в Excel")

    points, skipped = rows_to_points(headers, rows)
    workbook.close()
    return points, {"loaded": len(points), "skipped": skipped}


def detect_csv_encoding(path: Path) -> str:
    sample_size = 1024 * 1024
    raw = path.open("rb").read(sample_size)
    for encoding in ("utf-8-sig", "utf-8", "cp1251", "cp1254"):
        try:
            raw.decode(encoding)
            return encoding
        except UnicodeDecodeError:
            continue
    raise ValueError("Не удалось определить кодировку CSV")


def detect_csv_delimiter(path: Path, encoding: str) -> str:
    with path.open("r", encoding=encoding, newline="") as handle:
        sample = handle.read(100_000)
    try:
        return csv.Sniffer().sniff(sample, delimiters=";,\t|").delimiter
    except csv.Error:
        return ";"


def load_csv(path: Path) -> tuple[list[Point], dict[str, int]]:
    encoding = detect_csv_encoding(path)
    delimiter = detect_csv_delimiter(path, encoding)

    points: list[Point] = []
    skipped = 0
    columns: Optional[dict[str, int]] = None

    # CSV читается построчно. Файл целиком и список всех исходных строк в RAM не загружаются.
    with path.open("r", encoding=encoding, newline="", errors="strict") as handle:
        reader = csv.reader(handle, delimiter=delimiter)
        for row_number, row in enumerate(reader, start=1):
            if columns is None:
                if row_number <= 200 and is_header_row(row):
                    columns = detect_columns(row)
                elif row_number > 200:
                    raise ValueError("Не найдена строка заголовков в первых 200 строках CSV")
                continue

            point = row_to_point(row, columns)
            if point is None:
                skipped += 1
            else:
                points.append(point)

            if len(points) % 500_000 == 0:
                print(f"Загружено точек: {len(points):,}".replace(",", " "))

    if columns is None:
        raise ValueError("Не найдена строка заголовков в CSV")

    return points, {"loaded": len(points), "skipped": skipped}


def load_points(path: Path) -> tuple[list[Point], dict[str, int]]:
    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xlsm"):
        return load_excel(path)
    if suffix == ".csv":
        return load_csv(path)
    raise ValueError("Поддерживаются только XLSX, XLSM и CSV")
