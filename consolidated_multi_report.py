#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Multi-roster orchestration for the consolidated vehicle report.

A roster is selected independently for every report day:
1. exact roster date;
2. otherwise the latest roster dated before the report day;
3. otherwise the earliest uploaded roster.

The calculation engine and Excel layout remain in ``consolidated_report.py``.
"""
from __future__ import annotations

import argparse
import math
import os
from dataclasses import dataclass, replace
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Iterable, Sequence

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from excel_formatting import save_report_workbook

import consolidated_report as core
from arvento_io import Point
from roster_selection import select_effective_roster
from geozone_registry import suppress_speed_in_exclusions


@dataclass(frozen=True, slots=True)
class DatedRoster:
    path: Path
    day: date
    vehicles: dict[str, core.RosterVehicle]


def load_rosters(paths: Iterable[Path]) -> list[DatedRoster]:
    result: list[DatedRoster] = []
    seen_paths: set[Path] = set()
    for source in paths:
        path = Path(source).resolve()
        if path in seen_paths:
            continue
        seen_paths.add(path)
        if not path.exists():
            raise ValueError(f"Разнарядка не найдена: {path}")
        if path.suffix.lower() not in {".xlsx", ".xlsm"}:
            raise ValueError(f"Разнарядка должна быть XLSX/XLSM: {path.name}")
        vehicles, roster_day = core.load_roster(path)
        if roster_day is None:
            raise ValueError(
                f"Не удалось определить дату разнарядки «{path.name}». "
                "Укажите дату в имени файла или в верхней части Excel."
            )
        result.append(DatedRoster(path=path, day=roster_day, vehicles=vehicles))

    if not result:
        raise ValueError("Не загружены файлы разнарядки")

    # For duplicate dates the last uploaded file is authoritative.
    by_day: dict[date, DatedRoster] = {}
    for roster in result:
        by_day[roster.day] = roster
    return [by_day[day] for day in sorted(by_day)]


def select_roster(rosters: list[DatedRoster], report_day: date) -> DatedRoster:
    return select_effective_roster(rosters, report_day, lambda item: item.day)


def validate_period(start_day: date, end_day: date) -> None:
    if end_day < start_day:
        raise ValueError("Дата окончания раньше даты начала")
    days = (end_day - start_day).days + 1
    if days > core.MAX_REPORT_DAYS:
        raise ValueError(f"Период не должен превышать {core.MAX_REPORT_DAYS} дней")


def intervals_overlap(
    start: datetime,
    finish: datetime,
    window_start: datetime,
    window_finish: datetime,
) -> bool:
    return start < window_finish and finish > window_start


def has_night_site_mileage(
    report_day: date,
    points: Sequence[Point],
    site_polygon: list[tuple[float, float]],
) -> bool:
    """Return True only for site mileage within the report date's night windows.

    The flag belongs to the calendar date of the report and is set when there is
    positive mileage inside Akkuyu NPP during either interval:
    00:00 <= time < 05:00 or 22:00 <= time < 24:00.
    """
    track = core.sanitize_position_outliers(points)
    if len(track) < 2:
        return False

    site_states = core.smooth_boolean_states([
        core.point_in_polygon(point.lat, point.lon, site_polygon)
        for point in track
    ])
    day_start = datetime.combine(report_day, time.min)
    early_finish = datetime.combine(report_day, core.NIGHT_END)
    late_start = datetime.combine(report_day, core.NIGHT_START)
    day_finish = datetime.combine(report_day + timedelta(days=1), time.min)

    for index, (p1, p2) in enumerate(zip(track, track[1:])):
        gap = (p2.time - p1.time).total_seconds()
        if gap <= 0:
            continue

        distance = core.segment_distance(p1, p2)
        if not math.isfinite(distance) or distance <= 0:
            continue

        p1_inside = site_states[index]
        p2_inside = site_states[index + 1]
        if p1_inside == p2_inside:
            portions = [(p1_inside, 0.0, 1.0)]
        else:
            fraction = core.polygon_crossing_fraction(p1, p2, site_polygon)
            if not p1_inside and p2_inside:
                portions = [(False, 0.0, fraction), (True, fraction, 1.0)]
            else:
                portions = [(True, 0.0, fraction), (False, fraction, 1.0)]

        for is_inside, start_fraction, finish_fraction in portions:
            if not is_inside:
                continue
            fraction = max(0.0, finish_fraction - start_fraction)
            if distance * fraction <= 0:
                continue
            part_start = p1.time + (p2.time - p1.time) * start_fraction
            part_finish = p1.time + (p2.time - p1.time) * finish_fraction
            if intervals_overlap(part_start, part_finish, day_start, early_finish):
                return True
            if intervals_overlap(part_start, part_finish, late_start, day_finish):
                return True

    return False


def annotate_roster_usage(
    output_path: Path,
    rosters: list[DatedRoster],
    selected_by_day: dict[date, DatedRoster],
    registry: Any | None = None,
) -> None:
    workbook = load_workbook(output_path)
    try:
        parameters = workbook["Параметры"]
        for row in range(1, parameters.max_row + 1):
            label = parameters.cell(row, 1).value
            if label == "Разнарядка":
                parameters.cell(row, 2).value = "Несколько загруженных файлов"
            elif label == "Дата разнарядки":
                parameters.cell(row, 2).value = "выбирается отдельно для каждой даты отчёта"
            elif label == "Работа ночью":
                parameters.cell(row, 2).value = (
                    "есть положительный пробег внутри АЭС в дату отчёта "
                    "в интервалах 00:00–05:00 или 22:00–24:00"
                )

        parameters.append([])
        parameters.append(["Загруженные разнарядки", "Файл"])
        header_row = parameters.max_row
        for cell in parameters[header_row]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
        for roster in rosters:
            parameters.append([roster.day, roster.path.name])
            parameters.cell(parameters.max_row, 1).number_format = "dd.mm.yyyy"

        parameters.append([])
        parameters.append(["Дата отчёта", "Использованная разнарядка"])
        header_row = parameters.max_row
        for cell in parameters[header_row]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
        for report_day in sorted(selected_by_day):
            roster = selected_by_day[report_day]
            parameters.append([
                report_day,
                f"{roster.day:%d.%m.%Y} — {roster.path.name}",
            ])
            parameters.cell(parameters.max_row, 1).number_format = "dd.mm.yyyy"

        if registry is not None:
            parameters.append([])
            parameters.append(["Источник геозон", registry.source])
            if registry.version_refs:
                parameters.append(["Версии геозон", ", ".join(registry.version_refs)])

        diagnostics = workbook["Диагностика"]
        diagnostics.cell(1, 8).value = "Дата разнарядки"
        diagnostics.cell(1, 9).value = "Файл разнарядки"
        for column in (8, 9):
            cell = diagnostics.cell(1, column)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in range(2, diagnostics.max_row + 1):
            report_day = diagnostics.cell(row, 1).value
            if isinstance(report_day, datetime):
                report_day = report_day.date()
            if not isinstance(report_day, date):
                continue
            roster = selected_by_day.get(report_day)
            if roster is None:
                continue
            diagnostics.cell(row, 8).value = roster.day
            diagnostics.cell(row, 8).number_format = "dd.mm.yyyy"
            diagnostics.cell(row, 9).value = roster.path.name
        diagnostics.column_dimensions["H"].width = 18
        diagnostics.column_dimensions["I"].width = 42

        save_report_workbook(workbook, output_path)
    finally:
        workbook.close()


def generate_multi_roster_report(
    *,
    start_day: date,
    end_day: date,
    roster_paths: list[Path],
    output_path: Path,
    database_url: str,
    source_path: Path | None = None,
    route_kml: Path = core.DEFAULT_ROUTE_KML,
    geozones: Path = core.DEFAULT_GEOZONES,
) -> dict[str, Any]:
    validate_period(start_day, end_day)
    rosters = load_rosters(roster_paths)

    registry = core.load_registry(geozones)
    site_zone = core.find_site_boundary(registry)
    site_polygon = list(site_zone.points or [])
    route_polygon = registry.route_polygon or core.load_kml_polygon(route_kml)

    if source_path is not None:
        if not source_path.exists():
            raise ValueError(f"Выгрузка Arvento не найдена: {source_path}")
        tracks = core.iter_source_tracks(source_path, start_day, end_day)
        source_description = str(source_path)
    else:
        if not database_url:
            raise ValueError("DATABASE_URL не задан")
        tracks = core.iter_database_tracks(database_url, start_day, end_day)
        source_description = "PostgreSQL gps_points"

    rows: list[core.ReportRow] = []
    selected_by_day: dict[date, DatedRoster] = {}
    processed = 0
    for report_day, plate, points in tracks:
        roster = select_roster(rosters, report_day)
        selected_by_day[report_day] = roster
        analysis_points = suppress_speed_in_exclusions(points, registry)
        item = core.analyze_track(
            report_day,
            plate,
            analysis_points,
            roster.vehicles,
            site_polygon,
            route_polygon,
        )
        if item is not None:
            night_work = int(
                has_night_site_mileage(report_day, analysis_points, site_polygon)
            )
            item = replace(item, night_work=night_work)
            rows.append(item)
        processed += 1

    if not rows:
        raise ValueError("За выбранный период нет пригодных данных для отчёта")

    core.save_report(
        output_path,
        rows,
        rosters[0].path,
        rosters[0].day,
        Path("PostGIS: ROUTE") if registry.route_polygon else route_kml,
        site_zone,
        start_day,
        end_day,
        source_description,
    )
    annotate_roster_usage(output_path, rosters, selected_by_day, registry)
    return {
        "rows": len(rows),
        "processed_vehicle_days": processed,
        "missing_roster_rows": sum(not row.in_roster for row in rows),
        "rosters": len(rosters),
        "report_days": len(selected_by_day),
        "geofence_source": registry.source,
        "geofence_versions": len(registry.version_refs),
    }


def parse_day(value: str) -> date:
    try:
        return date.fromisoformat(value)
    except ValueError as exc:
        raise argparse.ArgumentTypeError("Дата должна быть в формате YYYY-MM-DD") from exc


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Сводный отчёт с несколькими разнарядками")
    period = parser.add_mutually_exclusive_group(required=True)
    period.add_argument("--date", type=parse_day)
    period.add_argument("--date-from", type=parse_day)
    parser.add_argument("--date-to", type=parse_day)
    parser.add_argument("--roster", type=Path, nargs="+", required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--source", type=Path)
    parser.add_argument("--database-url", default=os.environ.get("DATABASE_URL", ""))
    parser.add_argument("--route-kml", type=Path, default=core.DEFAULT_ROUTE_KML)
    parser.add_argument("--geozones", type=Path, default=core.DEFAULT_GEOZONES)
    return parser


def main() -> None:
    args = build_parser().parse_args()
    start_day = args.date or args.date_from
    end_day = args.date or args.date_to or start_day
    if start_day is None or end_day is None:
        raise ValueError("Не задан период")
    stats = generate_multi_roster_report(
        start_day=start_day,
        end_day=end_day,
        roster_paths=list(args.roster),
        output_path=args.output,
        database_url=args.database_url,
        source_path=args.source,
        route_kml=args.route_kml,
        geozones=args.geozones,
    )
    print(f"Готово: {args.output}")
    print(
        "Строк: {rows}; обработано автомобилей-дней: {processed_vehicle_days}; "
        "нет в выбранной разнарядке: {missing_roster_rows}; разнарядок: {rosters}".format(**stats)
    )


if __name__ == "__main__":
    main()
