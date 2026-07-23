from __future__ import annotations

from collections import defaultdict
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


SUMMARY_HEADERS = [
    "Госномер",
    "Водитель",
    "Грейд",
    "Должность",
    "Дирекция",
    "Дата разнарядки",
    "Первый въезд на площадку",
    "Последний выезд с площадки",
    "Начало периода",
    "Конец периода",
    "Состояние в начале",
    "Состояние в конце",
    "Въездов",
    "Выездов",
    "Пробег внутри, км",
    "Пробег снаружи, км",
    "Общий пробег, км",
    "Пробег внутри, %",
    "Пробег снаружи, %",
    "Время на территории",
    "Время в движении на территории",
    "Время стоянки на территории",
    "Количество стоянок",
    "Движение на территории, %",
    "Стоянка на территории, %",
]


def summary_row(item: dict[str, Any]) -> list[Any]:
    return [
        item["plate"],
        item.get("driver", ""),
        item.get("grade", ""),
        item.get("position", ""),
        item.get("directorate", ""),
        item.get("roster_date"),
        item.get("first_entry_time"),
        item.get("last_exit_time"),
        item["first_time"],
        item["last_time"],
        item["start_state"],
        item["end_state"],
        item["entries"],
        item["exits"],
        item["inside_km"],
        item["outside_km"],
        item["total_km"],
        item["inside_percent"],
        item["outside_percent"],
        item["inside_seconds"] / 86400.0,
        item["moving_seconds"] / 86400.0,
        item["stopped_seconds"] / 86400.0,
        item["stop_count"],
        item["moving_percent"],
        item["stopped_percent"],
    ]


def style_sheet(sheet) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.showGridLines = False
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[1].height = 48
    for column in range(1, sheet.max_column + 1):
        width = min(
            max(
                len(str(sheet.cell(row, column).value or ""))
                for row in range(1, min(sheet.max_row, 300) + 1)
            ) + 2,
            36,
        )
        sheet.column_dimensions[get_column_letter(column)].width = max(width, 11)


def format_summary_sheet(sheet, offset: int = 0) -> None:
    for row in sheet.iter_rows(min_row=2):
        row[offset + 5].number_format = "dd.mm.yyyy"
        for index in (6, 7, 8, 9):
            row[offset + index].number_format = "dd.mm.yyyy hh:mm:ss"
        for index in range(14, 17):
            row[offset + index].number_format = "0.000"
        for index in (17, 18, 23, 24):
            row[offset + index].number_format = "0.0%"
        for index in range(19, 22):
            row[offset + index].number_format = "[h]:mm:ss"
    style_sheet(sheet)


def aggregate_summaries(items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for item in items:
        grouped[item["plate"]].append(item)

    result = []
    for plate, rows in sorted(grouped.items()):
        total_km = sum(row["total_km"] for row in rows)
        inside_km = sum(row["inside_km"] for row in rows)
        inside_seconds = sum(row["inside_seconds"] for row in rows)
        moving_seconds = sum(row["moving_seconds"] for row in rows)
        stopped_seconds = sum(row["stopped_seconds"] for row in rows)
        first = min(rows, key=lambda row: row["first_time"])
        last = max(rows, key=lambda row: row["last_time"])
        latest_roster_row = max(
            rows,
            key=lambda row: (row.get("roster_date") or date.min, row["last_time"]),
        )
        entry_times = [row.get("first_entry_time") for row in rows if row.get("first_entry_time")]
        exit_times = [row.get("last_exit_time") for row in rows if row.get("last_exit_time")]

        result.append(
            {
                "plate": plate,
                "driver": latest_roster_row.get("driver", ""),
                "grade": latest_roster_row.get("grade", ""),
                "position": latest_roster_row.get("position", ""),
                "directorate": latest_roster_row.get("directorate", ""),
                "roster_date": latest_roster_row.get("roster_date"),
                "first_entry_time": min(entry_times) if entry_times else None,
                "last_exit_time": max(exit_times) if exit_times else None,
                "first_time": first["first_time"],
                "last_time": last["last_time"],
                "start_state": first["start_state"],
                "end_state": last["end_state"],
                "entries": sum(row["entries"] for row in rows),
                "exits": sum(row["exits"] for row in rows),
                "inside_km": inside_km,
                "outside_km": sum(row["outside_km"] for row in rows),
                "total_km": total_km,
                "inside_percent": inside_km / total_km if total_km else 0.0,
                "outside_percent": 1.0 - inside_km / total_km if total_km else 0.0,
                "inside_seconds": inside_seconds,
                "moving_seconds": moving_seconds,
                "stopped_seconds": stopped_seconds,
                "stop_count": sum(row["stop_count"] for row in rows),
                "moving_percent": moving_seconds / inside_seconds if inside_seconds else 0.0,
                "stopped_percent": stopped_seconds / inside_seconds if inside_seconds else 0.0,
            }
        )
    return result


def save_daily_book(path: Path, daily: dict[date, list[dict[str, Any]]]) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for day in sorted(daily):
        sheet = workbook.create_sheet(day.strftime("%d.%m.%Y"))
        sheet.append(SUMMARY_HEADERS)
        for item in daily[day]:
            sheet.append(summary_row(item))
        format_summary_sheet(sheet)
    workbook.save(path)


def save_summary_book(
    path: Path,
    daily: dict[date, list[dict[str, Any]]],
    stops: list[dict[str, Any]],
) -> None:
    workbook = Workbook()

    main = workbook.active
    main.title = "Сводка по автомобилям"
    main.append(SUMMARY_HEADERS)
    all_items = [item for day in sorted(daily) for item in daily[day]]
    for item in aggregate_summaries(all_items):
        main.append(summary_row(item))
    format_summary_sheet(main)

    by_day = workbook.create_sheet("Сводка по дням")
    by_day.append(["Дата"] + SUMMARY_HEADERS)
    for day in sorted(daily):
        for item in daily[day]:
            by_day.append([day] + summary_row(item))
    for row in by_day.iter_rows(min_row=2):
        row[0].number_format = "dd.mm.yyyy"
    format_summary_sheet(by_day, offset=1)

    detail = workbook.create_sheet("Стоянки")
    detail.append([
        "Дата", "Госномер", "Начало стоянки", "Конец стоянки", "Продолжительность",
        "Геозона", "Достоверность", "Широта до", "Долгота до", "Широта после", "Долгота после",
    ])
    for stop in sorted(stops, key=lambda item: (item["date"], item["plate"], item["start"])):
        detail.append([
            stop["date"], stop["plate"], stop["start"], stop["end"], stop["seconds"] / 86400.0,
            stop["zone"], stop["confidence"], stop["lat_before"], stop["lon_before"],
            stop["lat_after"], stop["lon_after"],
        ])
    for row in detail.iter_rows(min_row=2):
        row[0].number_format = "dd.mm.yyyy"
        row[2].number_format = "dd.mm.yyyy hh:mm:ss"
        row[3].number_format = "dd.mm.yyyy hh:mm:ss"
        row[4].number_format = "[h]:mm:ss"
        for index in range(7, 11):
            row[index].number_format = "0.000000"
    style_sheet(detail)

    zone_summary = workbook.create_sheet("Стоянки по геозонам")
    zone_summary.append(["Госномер", "Геозона", "Количество стоянок", "Общее время стоянок"])
    grouped: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for stop in stops:
        grouped[(stop["plate"], stop["zone"])].append(stop)
    for (plate, zone), rows in sorted(grouped.items()):
        zone_summary.append([plate, zone, len(rows), sum(row["seconds"] for row in rows) / 86400.0])
    for row in zone_summary.iter_rows(min_row=2):
        row[3].number_format = "[h]:mm:ss"
    style_sheet(zone_summary)

    workbook.save(path)
