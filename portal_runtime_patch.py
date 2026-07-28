from __future__ import annotations

"""Explicit runtime integration for the production report portal.

The patch is applied by :mod:`portal_entrypoint`; importing this module alone
has no application side effects. It provides:

* all three speed-violation categories in the browser preview;
* one-decimal formatting for floating-point indicators;
* current route/out-of-region wording in the form and summary.
"""

from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

import report_portal as implementation
import run_report_portal as current
from regional_speed_report import REGION_SHEET_NAME, ROUTE_SHEET_NAME, SITE_SHEET_NAME
from speed_violation_report import TURN_SHEET_NAME

SPEED_SHEET_NAMES = (
    TURN_SHEET_NAME,
    SITE_SHEET_NAME,
    ROUTE_SHEET_NAME,
    REGION_SHEET_NAME,
)

_original_json_cell = implementation.json_cell
_original_generate_report = current.generate_report_with_thresholds
_PATCHED = False


def json_cell_one_decimal(value: Any) -> Any:
    """Format floating-point web indicators with exactly one decimal place."""
    if isinstance(value, float):
        rounded = round(float(value), 1)
        return f"{rounded:.1f}"
    return _original_json_cell(value)


def violation_web_preview(path: Path) -> tuple[list[str], list[list[Any]], int]:
    """Read prohibited turns and all three speed categories from one workbook."""
    columns = [
        "Госномер",
        "Тип нарушения",
        "Дата",
        "Начало",
        "Окончание",
        "Максимальная скорость, км/ч",
        "Порог, км/ч",
        "Адрес",
        "Карта",
    ]
    records: list[tuple[str, datetime | None, list[Any]]] = []
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet_name in SPEED_SHEET_NAMES:
            if sheet_name not in workbook.sheetnames:
                continue
            sheet = workbook[sheet_name]
            headers = current._header_map(sheet)
            for row in sheet.iter_rows(min_row=2, values_only=True):
                plate = str(current._value(row, headers, "Госномер") or "").strip()
                if not plate:
                    continue

                if sheet_name == TURN_SHEET_NAME:
                    start = current._value(row, headers, "Начало прохода")
                    finish = current._value(row, headers, "Окончание прохода")
                    speeds = [
                        current._as_number(current._value(row, headers, name))
                        for name in ("Скорость в начале", "Скорость в конце")
                    ]
                    numeric_speeds = [value for value in speeds if value is not None]
                    max_speed = max(numeric_speeds) if numeric_speeds else None
                    address = " → ".join(
                        part
                        for part in (
                            str(current._value(row, headers, "Адрес начала") or "").strip(),
                            str(current._value(row, headers, "Адрес окончания") or "").strip(),
                        )
                        if part
                    )
                    map_url = current._map_url(
                        current._value(row, headers, "Координаты начала")
                        or current._value(row, headers, "Координаты окончания")
                    )
                    threshold = None
                else:
                    start = current._value(row, headers, "Начало нарушения")
                    finish = current._value(row, headers, "Окончание нарушения")
                    max_speed = current._as_number(
                        current._value(row, headers, "Максимальная скорость, км/ч")
                    )
                    threshold = current._as_number(
                        current._value(row, headers, "Порог фиксации, км/ч")
                    )
                    address = str(
                        current._value(row, headers, "Адрес максимума") or ""
                    ).strip()
                    map_url = current._map_url(
                        current._value(row, headers, "Координаты максимума")
                    )

                event_date = (
                    start.date()
                    if isinstance(start, datetime)
                    else current._value(row, headers, "Дата")
                )
                display = [
                    plate,
                    sheet_name,
                    json_cell_one_decimal(event_date),
                    json_cell_one_decimal(start),
                    json_cell_one_decimal(finish),
                    json_cell_one_decimal(max_speed),
                    json_cell_one_decimal(threshold),
                    address,
                    map_url,
                ]
                records.append(
                    (plate, start if isinstance(start, datetime) else None, display)
                )
    finally:
        workbook.close()

    records.sort(key=lambda item: (item[0], item[1] or datetime.min, item[2][1]))
    rows = [item[2] for item in records]
    return columns, rows, len(rows)


def generate_report_with_regional_summary(*args, **kwargs):
    """Keep the generated result but expose current threshold wording."""
    result = _original_generate_report(*args, **kwargs)
    report_type = args[0] if args else kwargs.get("report_type")
    if report_type == "violation":
        summary = result.get("summary", {})
        value = summary.pop("Порог вне площадки", None)
        if value is not None:
            summary["Порог маршрута и вне региона"] = value
    return result


def apply_runtime_patch() -> None:
    """Apply portal integration once and make the action directly testable."""
    global _PATCHED
    if _PATCHED:
        return

    implementation.json_cell = json_cell_one_decimal
    current.implementation.json_cell = json_cell_one_decimal
    current.violation_web_preview = violation_web_preview
    current.generate_report_with_thresholds = generate_report_with_regional_summary

    implementation.HTML = implementation.HTML.replace(
        "Порог вне площадки, км/ч",
        "Порог маршрута и вне региона, км/ч",
    )
    implementation.HTML = implementation.HTML.replace(
        "Порог вне площадки",
        "Порог маршрута и вне региона",
    )
    _PATCHED = True


__all__ = [
    "SPEED_SHEET_NAMES",
    "apply_runtime_patch",
    "json_cell_one_decimal",
    "violation_web_preview",
]
