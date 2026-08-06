#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Add daily refuelling totals from Fuel Monitor to the consolidated report."""
from __future__ import annotations

import os
from collections import defaultdict
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any

import psycopg
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from consolidated_time_logic import apply_consolidated_time_logic

apply_consolidated_time_logic()

from consolidated_mileage_logic import apply_consolidated_mileage_logic
from consolidated_multi_report import generate_multi_roster_report as generate_base_report
from consolidated_performance import apply_consolidated_performance
from mileage_review_policy import (
    annotate_mileage_review,
    apply_authoritative_mileage_policy,
)
from roster_registry import normalize_plate

apply_consolidated_performance()
apply_consolidated_mileage_logic()
apply_authoritative_mileage_policy()

FUEL_HEADER = "Заправка, л"
FUEL_SOURCE = "Fuel Monitor: public.fuel_events"


def psycopg_database_url(value: str) -> str:
    """Convert an SQLAlchemy PostgreSQL URL into a psycopg-compatible URL."""
    url = value.strip()
    for prefix in ("postgresql+psycopg://", "postgres+psycopg://"):
        if url.startswith(prefix):
            return "postgresql://" + url[len(prefix):]
    return url


def load_fuel_totals(
    database_url: str,
    start_day: date,
    end_day: date,
) -> dict[tuple[date, str], float]:
    """Return refuelled litres grouped by local calendar day and normalized plate."""
    if not database_url.strip():
        return {}

    start_at = datetime.combine(start_day, time.min)
    finish_at = datetime.combine(end_day + timedelta(days=1), time.min)
    timeout = int(os.environ.get("FUEL_DB_CONNECT_TIMEOUT_SECONDS", "10"))
    query = """
        SELECT
            event_dt::date AS event_day,
            plate,
            COALESCE(SUM(liters), 0) AS fuel_liters
        FROM public.fuel_events
        WHERE event_dt >= %s
          AND event_dt < %s
        GROUP BY event_dt::date, plate
        ORDER BY event_day, plate
    """

    totals: defaultdict[tuple[date, str], float] = defaultdict(float)
    try:
        with psycopg.connect(
            psycopg_database_url(database_url),
            connect_timeout=timeout,
        ) as connection:
            with connection.cursor() as cursor:
                cursor.execute(query, (start_at, finish_at))
                for event_day, plate, liters in cursor:
                    if isinstance(event_day, datetime):
                        event_day = event_day.date()
                    normalized = normalize_plate(plate)
                    if isinstance(event_day, date) and normalized:
                        totals[(event_day, normalized)] += float(liters or 0)
    except Exception as exc:
        raise RuntimeError(f"Не удалось прочитать заправки из Fuel Monitor: {exc}") from exc

    return dict(totals)


def parse_excel_day(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if value is None:
        return None
    text = str(value).strip()
    for parser in (
        lambda: date.fromisoformat(text),
        lambda: datetime.strptime(text, "%d.%m.%Y").date(),
    ):
        try:
            return parser()
        except ValueError:
            continue
    return None


def add_fuel_column(
    output_path: Path,
    totals: dict[tuple[date, str], float],
    *,
    configured: bool,
) -> tuple[float, int]:
    """Append the rightmost fuel column and return total litres and matched rows."""
    workbook = load_workbook(output_path)
    try:
        sheet = workbook["Сводный отчет"]
        headers = {
            str(cell.value or "").strip(): cell.column
            for cell in sheet[1]
        }
        day_column = headers.get("Дата")
        plate_column = next(
            (
                headers[name]
                for name in ("Госномер / Plaka", "Госномер", "Номерной знак")
                if name in headers
            ),
            None,
        )
        if day_column is None or plate_column is None:
            raise RuntimeError("В сводном отчёте не найдены столбцы даты или госномера")

        fuel_column = headers.get(FUEL_HEADER) or (sheet.max_column + 1)
        header = sheet.cell(1, fuel_column, FUEL_HEADER)
        header.font = Font(bold=True, color="FFFFFF")
        header.fill = PatternFill("solid", fgColor="1F4E78")
        header.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet.column_dimensions[get_column_letter(fuel_column)].width = 16

        total_liters = 0.0
        matched_rows = 0
        for row_index in range(2, sheet.max_row + 1):
            report_day = parse_excel_day(sheet.cell(row_index, day_column).value)
            plate = normalize_plate(sheet.cell(row_index, plate_column).value)
            cell = sheet.cell(row_index, fuel_column)
            if not configured or report_day is None or not plate:
                cell.value = None
            else:
                value = round(totals.get((report_day, plate), 0.0), 1)
                cell.value = value
                cell.number_format = "0.0"
                total_liters += value
                if value > 0:
                    matched_rows += 1
            cell.alignment = Alignment(vertical="center", wrap_text=True)

        sheet.auto_filter.ref = sheet.dimensions

        parameters = workbook["Параметры"]
        parameters.append([])
        parameters.append([
            "Источник заправок",
            FUEL_SOURCE if configured else "не настроен: FUEL_DATABASE_URL пуст",
        ])
        parameters.append([
            "Расчёт заправок",
            "сумма fuel_events.liters по календарной дате и нормализованному госномеру",
        ])
        for row in parameters.iter_rows(min_row=max(1, parameters.max_row - 2)):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        workbook.save(output_path)
        return round(total_liters, 1), matched_rows
    finally:
        workbook.close()


def generate_multi_roster_report(**kwargs: Any) -> dict[str, Any]:
    """Generate the report, enrich fuel and flag mileage-review candidates."""
    output_path = Path(kwargs["output_path"])
    start_day: date = kwargs["start_day"]
    end_day: date = kwargs["end_day"]
    database_url = str(kwargs["database_url"])
    stats: dict[str, Any] = dict(generate_base_report(**kwargs))

    fuel_database_url = os.environ.get("FUEL_DATABASE_URL", "").strip()
    totals = load_fuel_totals(fuel_database_url, start_day, end_day)
    fuel_liters, fuel_rows = add_fuel_column(
        output_path,
        totals,
        configured=bool(fuel_database_url),
    )
    review_stats = annotate_mileage_review(
        output_path,
        database_url,
        start_day,
        end_day,
        refresh=True,
    )
    stats["fuel_liters"] = fuel_liters
    stats["fuel_rows"] = fuel_rows
    stats["fuel_configured"] = int(bool(fuel_database_url))
    stats["mileage_review_candidates"] = review_stats["candidates"]
    stats["mileage_review_rows"] = review_stats["flagged_rows"]
    return stats


__all__ = ["generate_multi_roster_report", "load_fuel_totals", "add_fuel_column"]
