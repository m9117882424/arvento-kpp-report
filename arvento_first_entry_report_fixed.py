#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""Запуск отчёта по первому въезду с устойчивым определением стороны КПП.

Исправляет пропуски пересечений, когда промежуточная GPS-точка попадает в
пограничную зону менее MIN_SIDE_DISTANCE_M от линии ворот. Итоговый лист
«Первый въезд» не содержит отдельного столбца со ссылкой на карту и начинается
с двух строк заголовка, сформированных по дате отчёта, времени и разнарядке.
"""

from __future__ import annotations

import csv
import re
from collections import defaultdict
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, Literal

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from excel_formatting import save_report_workbook

import arvento_first_entry_report as base


_original_load_coordinate_points = base.load_coordinate_points
_original_create_report = base.create_report


DATE_PATTERNS = (
    re.compile(r"(?<!\d)(\d{1,2})[._-](\d{1,2})[._-](\d{4})(?!\d)"),
    re.compile(r"(?<!\d)(\d{4})[._-](\d{1,2})[._-](\d{1,2})(?!\d)"),
    re.compile(r"(?<!\d)(\d{1,2})[._-](\d{1,2})[._-](\d{2})(?!\d)"),
)


def detect_csv_encoding(path: Path) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp1251", "cp1254"):
        try:
            with path.open("r", encoding=encoding, newline="") as stream:
                stream.read(256 * 1024)
            return encoding
        except UnicodeDecodeError:
            continue
    raise ValueError("Не удалось определить кодировку CSV")


def detect_csv_delimiter(path: Path, encoding: str) -> str:
    with path.open("r", encoding=encoding, newline="") as stream:
        sample = stream.read(64 * 1024)
    try:
        return csv.Sniffer().sniff(sample, delimiters=";,\t|").delimiter
    except csv.Error:
        return ","


def load_coordinate_points(path: Path) -> list[base.Point]:
    """Загружает координатные точки из XLSX/XLSM или CSV."""
    if path.suffix.lower() != ".csv":
        return _original_load_coordinate_points(path)

    encoding = detect_csv_encoding(path)
    delimiter = detect_csv_delimiter(path, encoding)
    points: list[base.Point] = []
    columns: dict[str, int | None] | None = None
    epoch = datetime(1899, 12, 30)

    with path.open("r", encoding=encoding, newline="") as stream:
        reader = csv.reader(stream, delimiter=delimiter)
        for row in reader:
            if columns is None:
                headers = [base.clean(value) for value in row]
                candidate = {
                    key: base.find_column(headers, aliases)
                    for key, aliases in base.COORD_ALIASES.items()
                }
                if all(candidate.get(key) is not None for key in ("plate", "timestamp", "lat", "lon")):
                    columns = candidate
                continue

            plate = base.normalize_plate(base.get_value(row, columns["plate"]))
            timestamp = base.parse_datetime(base.get_value(row, columns["timestamp"]), epoch)
            lat = base.as_float(base.get_value(row, columns["lat"]))
            lon = base.as_float(base.get_value(row, columns["lon"]))
            if not plate or timestamp is None or lat is None or lon is None:
                continue

            points.append(
                base.Point(
                    plate=plate,
                    timestamp=timestamp,
                    lat=lat,
                    lon=lon,
                    region=base.clean(base.get_value(row, columns.get("region"))),
                )
            )

    if columns is None:
        raise ValueError(
            "В координатной CSV-выгрузке не найдены колонки: "
            "госномер, дата/время, широта, долгота"
        )
    if not points:
        raise ValueError("В координатной CSV-выгрузке нет пригодных GPS-точек")
    return points


def stable_side(distance_m: float) -> int:
    if distance_m >= base.MIN_SIDE_DISTANCE_M:
        return 1
    if distance_m <= -base.MIN_SIDE_DISTANCE_M:
        return -1
    return 0


def detect_coordinate_crossings(points: list[base.Point]) -> list[base.Crossing]:
    """Определяет пересечения через последнюю устойчивую сторону линии КПП.

    Последовательность вида +34 м -> -0,56 м -> -8,68 м теперь считается
    одним въездом: точка -0,56 м находится в пограничной зоне и не сбрасывает
    последнюю устойчивую внешнюю сторону.
    """
    by_plate: dict[str, list[base.Point]] = defaultdict(list)
    for point in points:
        by_plate[point.plate].append(point)

    crossings: list[base.Crossing] = []
    for plate, vehicle_points in by_plate.items():
        vehicle_points.sort(key=lambda item: item.timestamp)
        last_by_gate_direction: dict[tuple[str, str], datetime] = {}

        stable: dict[str, tuple[int, base.Point] | None] = {}
        for gate_name, gate in base.GATE_SEGMENTS.items():
            first_point = vehicle_points[0] if vehicle_points else None
            if first_point is None:
                stable[gate_name] = None
                continue
            first_side = stable_side(base.signed_gate_distance_m(first_point, gate))
            stable[gate_name] = (first_side, first_point) if first_side else None

        for current in vehicle_points[1:]:
            for gate_name, gate in base.GATE_SEGMENTS.items():
                current_side = stable_side(base.signed_gate_distance_m(current, gate))
                if current_side == 0:
                    continue

                previous = stable.get(gate_name)
                if previous is None:
                    stable[gate_name] = (current_side, current)
                    continue

                previous_side, previous_point = previous
                if current_side == previous_side:
                    stable[gate_name] = (current_side, current)
                    continue

                gap = (current.timestamp - previous_point.timestamp).total_seconds()
                if gap <= 0 or gap > base.MAX_GPS_GAP_SECONDS:
                    stable[gate_name] = (current_side, current)
                    continue

                direction: Literal["entry", "exit"] = (
                    "entry" if previous_side == 1 and current_side == -1 else "exit"
                )
                fraction = base.crossing_fraction(previous_point, current, gate)
                stable[gate_name] = (current_side, current)
                if fraction is None:
                    continue

                timestamp = previous_point.timestamp + (
                    current.timestamp - previous_point.timestamp
                ) * fraction
                key = (gate_name, direction)
                prior_event = last_by_gate_direction.get(key)
                if prior_event and (
                    timestamp - prior_event
                ).total_seconds() < base.CROSSING_COOLDOWN_SECONDS:
                    continue

                lat = previous_point.lat + (current.lat - previous_point.lat) * fraction
                lon = previous_point.lon + (current.lon - previous_point.lon) * fraction
                crossings.append(
                    base.Crossing(
                        plate=plate,
                        timestamp=timestamp,
                        gate=gate_name,
                        direction=direction,
                        lat=lat,
                        lon=lon,
                    )
                )
                last_by_gate_direction[key] = timestamp

    return sorted(crossings, key=lambda item: (item.timestamp, item.plate, item.gate))


def extract_date_from_text(value: Any) -> date | None:
    """Извлекает дату из имени файла или текста без привязки к языку."""
    text = str(value or "")
    for index, pattern in enumerate(DATE_PATTERNS):
        match = pattern.search(text)
        if not match:
            continue
        first, second, third = (int(part) for part in match.groups())
        if index == 1:
            year, month, day = first, second, third
        else:
            day, month, year = first, second, third
            if year < 100:
                year += 2000
        try:
            parsed = date(year, month, day)
        except ValueError:
            continue
        if 2020 <= parsed.year <= 2100:
            return parsed
    return None


def detect_roster_date(path: Path) -> date | None:
    """Определяет дату разнарядки сначала по имени, затем по верхним строкам."""
    filename_date = extract_date_from_text(path.stem)
    if filename_date is not None:
        return filename_date

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(min_row=1, max_row=12, max_col=20, values_only=True):
                for value in row:
                    if isinstance(value, datetime):
                        candidate = value.date()
                    elif isinstance(value, date):
                        candidate = value
                    else:
                        candidate = extract_date_from_text(value)
                    if candidate is not None and 2020 <= candidate.year <= 2100:
                        return candidate
    finally:
        workbook.close()
    return None


def detect_report_date(entries: list[base.Entry], coordinate_path: Path) -> date | None:
    if entries:
        return min(entry.timestamp.date() for entry in entries)
    return extract_date_from_text(coordinate_path.stem)


def display_title_time(value: time | None, *, leading_zero: bool) -> str:
    if value is None:
        return ""
    hour = f"{value.hour:02d}" if leading_zero else str(value.hour)
    return f"{hour}:{value.minute:02d}"


def build_report_titles(
    report_date: date | None,
    roster_date: date | None,
    filters: base.ReportFilters,
) -> tuple[str, str]:
    report_date_text = report_date.strftime("%d.%m.%Y") if report_date else "дата не определена"
    roster_date_text = roster_date.strftime("%d.%m.%Y") if roster_date else "дата не определена"

    if filters.time_from and filters.time_to:
        time_text = (
            f"(с {display_title_time(filters.time_from, leading_zero=False)} "
            f"до {display_title_time(filters.time_to, leading_zero=True)})"
        )
    elif filters.time_from:
        time_text = f"(с {display_title_time(filters.time_from, leading_zero=False)})"
    elif filters.time_to:
        time_text = f"(до {display_title_time(filters.time_to, leading_zero=True)})"
    else:
        time_text = "(за весь день)"

    title = (
        "Отчет по времени въезда служебных автомобилей в утреннее время "
        f"за {report_date_text} {time_text} без учёта повторных проездов через геозону."
    )
    subtitle = (
        "(принадлежность водителей проставлена по разнарядке "
        f"от {roster_date_text})"
    )
    return title, subtitle


def create_report_without_map_column(*args: Any, **kwargs: Any) -> None:
    """Build the report, remove the map column and add two dynamic title rows."""
    _original_create_report(*args, **kwargs)
    output_path = Path(args[0] if args else kwargs["output_path"])
    entries = list(args[1] if len(args) > 1 else kwargs.get("entries", []))
    coordinate_path = Path(args[3] if len(args) > 3 else kwargs["coordinate_path"])
    roster_path = Path(args[4] if len(args) > 4 else kwargs["roster_path"])
    filters = args[6] if len(args) > 6 else kwargs["filters"]

    report_date = detect_report_date(entries, coordinate_path)
    roster_date = detect_roster_date(roster_path)
    title, subtitle = build_report_titles(report_date, roster_date, filters)

    workbook = load_workbook(output_path)
    try:
        if "Первый въезд" not in workbook.sheetnames:
            raise RuntimeError("В отчёте отсутствует лист «Первый въезд»")
        sheet = workbook["Первый въезд"]

        # Column E in the base report contains only «Показать на карте» links.
        sheet.delete_cols(5, 1)
        sheet.insert_rows(1, amount=2)
        sheet.merge_cells("A1:J1")
        sheet.merge_cells("A2:J2")
        sheet["A1"] = title
        sheet["A2"] = subtitle
        sheet["A1"].font = Font(bold=True, size=14)
        sheet["A2"].font = Font(italic=True, size=11)
        sheet["A1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet["A2"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet.row_dimensions[1].height = 42
        sheet.row_dimensions[2].height = 32
        sheet.row_dimensions[3].height = 26

        max_row = max(sheet.max_row, 3)
        sheet.auto_filter.ref = f"A3:J{max_row}"
        sheet.freeze_panes = "A4"

        widths = [6, 15, 13, 13, 15, 24, 10, 38, 45, 55]
        for index, width in enumerate(widths, 1):
            sheet.column_dimensions[get_column_letter(index)].width = width
        sheet.column_dimensions["K"].hidden = True

        save_report_workbook(workbook, output_path)
    finally:
        workbook.close()


base.load_coordinate_points = load_coordinate_points
base.detect_coordinate_crossings = detect_coordinate_crossings
base.create_report = create_report_without_map_column


if __name__ == "__main__":
    try:
        base.main()
    except Exception as exc:
        print(f"ОШИБКА: {exc}")
        raise
