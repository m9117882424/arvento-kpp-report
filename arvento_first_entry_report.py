#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Отчёт по первому въезду автомобилей через KPP 4 / KPP 5.

Основной источник — координатная выгрузка Arvento. Въезд определяется по
фактическому пересечению линий КПП между двумя последовательными GPS-точками,
даже если геозона KPP в Arvento не привязана к автомобилю.

Опционально можно передать отчёт «Сигнал географического региона». Он
используется как резервный источник для автомобилей, по которым пересечение
по координатам определить не удалось.

Опция «учитывать предыдущие выезды» исключает въезд, если в этот же день
автомобиль до него уже пересёк линию одного из КПП в направлении выезда.
Это исключает повторный въезд машины, которая на начало дня находилась внутри.

Пример:
    python arvento_first_entry_report.py Report_tsmenerji3.xlsx roster.xlsx report.xlsx \
        --grade-from 7 --grade-to 14 --time-from 07:00 --time-to 09:00 \
        --consider-previous-exits

С дополнительным отчётом геозон:
    python arvento_first_entry_report.py Report_tsmenerji3.xlsx roster.xlsx report.xlsx \
        --geo-events "Report_Сигнал_географического_региона.xlsx" \
        --grade-from 7 --grade-to 14 --time-from 07:00 --time-to 09:00 \
        --consider-previous-exits
"""

from __future__ import annotations

import argparse
import math
import re
import warnings
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from itertools import chain
from pathlib import Path
from typing import Any, Iterable, Iterator, Literal, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel
from excel_formatting import save_report_workbook

warnings.filterwarnings(
    "ignore",
    message="Workbook contains no default style, apply openpyxl's default",
    category=UserWarning,
)

# Линии ворот. Порядок точек задаёт направление въезда: сторона + -> сторона -.
# Обратное пересечение (сторона - -> сторона +) считается выездом.
GATE_SEGMENTS: dict[str, tuple[tuple[float, float], tuple[float, float]]] = {
    "KPP 4 TEST": (
        (36.153566, 33.5728024),
        (36.153566, 33.5730806),
    ),
    "KPP 5 TEST": (
        (36.15886644084745, 33.549775396083724),
        (36.159167964778234, 33.55101996362877),
    ),
}

# В выгрузке 34KBF723 между точками у КПП был разрыв 8 мин 17 сек.
# 15 минут позволяют корректно интерполировать такое пересечение.
MAX_GPS_GAP_SECONDS = 15 * 60
GATE_EXTENSION_M = 35.0
MIN_SIDE_DISTANCE_M = 1.0
EVENT_MATCH_WINDOW_SECONDS = 15 * 60
CROSSING_COOLDOWN_SECONDS = 120

COORD_ALIASES = {
    "plate": ("Номерной знак", "Госномер", "Plaka", "License Plate"),
    "timestamp": ("Дата / время", "Дата/время", "Tarih / Saat", "Date / Time"),
    "lat": ("Latitudine", "Широта", "Latitude", "Enlem"),
    "lon": ("Долгота", "Longitude", "Boylam"),
    "region": ("Область", "Регион", "Геозона", "Region"),
}

EVENT_ALIASES = {
    "plate": ("Номерной знак", "Госномер", "Plaka", "License Plate"),
    "date": ("Дата / время (Дата)", "Дата", "Date"),
    "time": ("Дата / время (Час)", "Время", "Time"),
    "event_type": ("Тип аварии", "Тип события", "Event Type"),
    "region": ("Область", "Регион", "Геозона", "Region"),
}

ROSTER_ALIASES = {
    "plate": ("Гос рег знак", "PLAKA", "Госномер", "Номерной знак"),
    "model": ("Марка, модель", "Marka, model", "Модель"),
    "grade": ("Грейд", "SCALA", "Grade"),
    "driver": ("Пользователь", "KULLANICI", "Водитель"),
    "position": ("Должность", "GÖREVİ", "GOREVI"),
    "directorate": ("Дирекция", "Directorate"),
}


@dataclass(frozen=True)
class Point:
    plate: str
    timestamp: datetime
    lat: float
    lon: float
    region: str = ""


@dataclass(frozen=True)
class Crossing:
    plate: str
    timestamp: datetime
    gate: str
    direction: Literal["entry", "exit"]
    lat: float
    lon: float
    source: str = "пересечение по координатам"


@dataclass(frozen=True)
class Entry:
    plate: str
    timestamp: datetime
    gate: str
    lat: Optional[float]
    lon: Optional[float]
    source: str


@dataclass
class VehicleInfo:
    plate: str
    model: str = ""
    grade: str = ""
    driver: str = ""
    position: str = ""
    directorate: str = ""


@dataclass(frozen=True)
class ReportFilters:
    grade_from: Optional[float] = None
    grade_to: Optional[float] = None
    time_from: Optional[time] = None
    time_to: Optional[time] = None
    consider_previous_exits: bool = False


def clean(value: Any) -> str:
    return " ".join(str(value or "").replace("\n", " ").split()).strip()


def normalized(value: Any) -> str:
    return clean(value).lower().replace("ё", "е").replace("ı", "i")


def normalize_plate(value: Any) -> str:
    return "".join(ch for ch in clean(value).upper() if ch.isalnum())


def parse_grade_number(value: Any) -> Optional[float]:
    """Берёт только число: 7a, 7A и 7b считаются грейдом 7."""
    match = re.search(r"\d+(?:[.,]\d+)?", clean(value))
    return float(match.group().replace(",", ".")) if match else None


def parse_clock(value: Any) -> Optional[time]:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.time().replace(tzinfo=None)
    if isinstance(value, time):
        return value.replace(tzinfo=None)
    if isinstance(value, (int, float)):
        fraction = float(value) % 1.0
        total_seconds = int(round(fraction * 86400)) % 86400
        return (datetime.min + timedelta(seconds=total_seconds)).time()
    text = clean(value)
    if not text:
        return None
    for fmt in ("%H", "%H:%M", "%H:%M:%S"):
        try:
            return datetime.strptime(text, fmt).time()
        except ValueError:
            pass
    raise ValueError(f"Некорректное время: {value}")


def parse_datetime(value: Any, epoch: datetime) -> Optional[datetime]:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    if isinstance(value, date):
        return datetime.combine(value, time.min)
    if isinstance(value, (int, float)):
        converted = from_excel(value, epoch)
        if isinstance(converted, datetime):
            return converted.replace(tzinfo=None)
        if isinstance(converted, date):
            return datetime.combine(converted, time.min)
        return None
    text = clean(value)
    for fmt in (
        "%d.%m.%Y %H:%M:%S",
        "%d.%m.%Y %H:%M",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%m/%d/%Y %H:%M:%S",
    ):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            pass
    return None


def parse_date_and_time(date_value: Any, time_value: Any, epoch: datetime) -> Optional[datetime]:
    if isinstance(date_value, datetime):
        day = date_value.date()
    elif isinstance(date_value, date):
        day = date_value
    elif isinstance(date_value, (int, float)):
        converted = from_excel(date_value, epoch)
        day = converted.date() if isinstance(converted, datetime) else converted
    else:
        day = None
        text = clean(date_value)
        for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%m/%d/%Y"):
            try:
                day = datetime.strptime(text, fmt).date()
                break
            except ValueError:
                pass
    clock = parse_clock(time_value)
    return datetime.combine(day, clock or time.min) if isinstance(day, date) else None


def as_float(value: Any) -> Optional[float]:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(clean(value).replace(" ", "").replace(",", "."))
    except ValueError:
        return None


def find_column(headers: list[str], aliases: Iterable[str]) -> Optional[int]:
    values = [normalized(header) for header in headers]
    for alias in aliases:
        needle = normalized(alias)
        for index, header in enumerate(values):
            if header == needle or needle in header:
                return index
    return None


def find_header(
    rows: Iterator[tuple[Any, ...]],
    aliases: dict[str, Iterable[str]],
    required: tuple[str, ...],
    max_rows: int = 30,
) -> tuple[Optional[dict[str, Optional[int]]], Iterator[tuple[Any, ...]]]:
    buffered: list[tuple[Any, ...]] = []
    for _ in range(max_rows):
        try:
            row = next(rows)
        except StopIteration:
            break
        buffered.append(row)
        headers = [clean(value) for value in row]
        columns = {key: find_column(headers, names) for key, names in aliases.items()}
        if all(columns.get(key) is not None for key in required):
            return columns, chain(buffered[len(buffered):], rows)
    return None, iter(())


def get_value(row: tuple[Any, ...], index: Optional[int]) -> Any:
    return row[index] if index is not None and index < len(row) else None


def load_roster(path: Path) -> dict[str, VehicleInfo]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        result: dict[str, VehicleInfo] = {}
        for sheet in workbook.worksheets:
            iterator = iter(sheet.iter_rows(values_only=True))
            columns, data_rows = find_header(iterator, ROSTER_ALIASES, ("plate",))
            if columns is None:
                continue
            for row in data_rows:
                plate = normalize_plate(get_value(row, columns["plate"]))
                if not plate:
                    continue
                info = VehicleInfo(
                    plate=plate,
                    model=clean(get_value(row, columns.get("model"))),
                    grade=clean(get_value(row, columns.get("grade"))),
                    driver=clean(get_value(row, columns.get("driver"))),
                    position=clean(get_value(row, columns.get("position"))),
                    directorate=clean(get_value(row, columns.get("directorate"))),
                )
                old = result.get(plate)
                old_score = (
                    sum(bool(getattr(old, field)) for field in ("model", "grade", "driver", "position", "directorate"))
                    if old else -1
                )
                new_score = sum(
                    bool(getattr(info, field))
                    for field in ("model", "grade", "driver", "position", "directorate")
                )
                if new_score > old_score:
                    result[plate] = info
        if not result:
            raise ValueError("В разнарядке не найден столбец с госномером")
        return result
    finally:
        workbook.close()


def load_coordinate_points(path: Path) -> list[Point]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        points: list[Point] = []
        found = False
        for sheet in workbook.worksheets:
            iterator = iter(sheet.iter_rows(values_only=True))
            columns, data_rows = find_header(iterator, COORD_ALIASES, ("plate", "timestamp", "lat", "lon"))
            if columns is None:
                continue
            found = True
            for row in data_rows:
                plate = normalize_plate(get_value(row, columns["plate"]))
                timestamp = parse_datetime(get_value(row, columns["timestamp"]), workbook.epoch)
                lat = as_float(get_value(row, columns["lat"]))
                lon = as_float(get_value(row, columns["lon"]))
                if not plate or timestamp is None or lat is None or lon is None:
                    continue
                points.append(Point(
                    plate=plate,
                    timestamp=timestamp,
                    lat=lat,
                    lon=lon,
                    region=clean(get_value(row, columns.get("region"))),
                ))
        if not found:
            raise ValueError("В координатной выгрузке не найдены колонки: госномер, дата/время, широта, долгота")
        if not points:
            raise ValueError("В координатной выгрузке нет пригодных GPS-точек")
        return points
    finally:
        workbook.close()


def load_geo_entries(path: Path) -> list[Entry]:
    """Загружает только события входа в геозону как резервные кандидаты въезда.

    Событие выхода из небольшой геозоны КПП не равно выезду с территории,
    поэтому для условия предыдущего выезда оно не используется.
    """
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        events: list[Entry] = []
        for sheet in workbook.worksheets:
            iterator = iter(sheet.iter_rows(values_only=True))
            columns, data_rows = find_header(
                iterator,
                EVENT_ALIASES,
                ("plate", "date", "time", "event_type", "region"),
            )
            if columns is None:
                continue
            for row in data_rows:
                plate = normalize_plate(get_value(row, columns["plate"]))
                event_type = normalized(get_value(row, columns["event_type"]))
                region = clean(get_value(row, columns["region"]))
                gate = next((name for name in GATE_SEGMENTS if name.lower() in region.lower()), None)
                if not plate or gate is None or "вход" not in event_type:
                    continue
                timestamp = parse_date_and_time(
                    get_value(row, columns["date"]),
                    get_value(row, columns["time"]),
                    workbook.epoch,
                )
                if timestamp is None:
                    continue
                events.append(Entry(plate, timestamp, gate, None, None, "геозона Arvento"))
        return first_per_vehicle_day(events)
    finally:
        workbook.close()


def local_xy_m(lat: float, lon: float, origin_lat: float, origin_lon: float) -> tuple[float, float]:
    y = (lat - origin_lat) * 111_320.0
    x = (lon - origin_lon) * 111_320.0 * math.cos(math.radians(origin_lat))
    return x, y


def cross_2d(ax: float, ay: float, bx: float, by: float) -> float:
    return ax * by - ay * bx


def signed_gate_distance_m(
    point: Point,
    gate: tuple[tuple[float, float], tuple[float, float]],
) -> float:
    (a_lat, a_lon), (b_lat, b_lon) = gate
    origin_lat = (a_lat + b_lat) / 2.0
    origin_lon = (a_lon + b_lon) / 2.0
    ax, ay = local_xy_m(a_lat, a_lon, origin_lat, origin_lon)
    bx, by = local_xy_m(b_lat, b_lon, origin_lat, origin_lon)
    px, py = local_xy_m(point.lat, point.lon, origin_lat, origin_lon)
    vx, vy = bx - ax, by - ay
    length = math.hypot(vx, vy)
    if length == 0:
        return 0.0
    return cross_2d(vx, vy, px - ax, py - ay) / length


def crossing_fraction(
    p1: Point,
    p2: Point,
    gate: tuple[tuple[float, float], tuple[float, float]],
) -> Optional[float]:
    """Возвращает долю времени до пересечения конечного отрезка ворот."""
    (a_lat, a_lon), (b_lat, b_lon) = gate
    origin_lat = (a_lat + b_lat) / 2.0
    origin_lon = (a_lon + b_lon) / 2.0

    ax, ay = local_xy_m(a_lat, a_lon, origin_lat, origin_lon)
    bx, by = local_xy_m(b_lat, b_lon, origin_lat, origin_lon)
    x1, y1 = local_xy_m(p1.lat, p1.lon, origin_lat, origin_lon)
    x2, y2 = local_xy_m(p2.lat, p2.lon, origin_lat, origin_lon)

    sx, sy = bx - ax, by - ay
    gate_length = math.hypot(sx, sy)
    if gate_length == 0:
        return None

    ux, uy = sx / gate_length, sy / gate_length
    ax -= ux * GATE_EXTENSION_M
    ay -= uy * GATE_EXTENSION_M
    bx += ux * GATE_EXTENSION_M
    by += uy * GATE_EXTENSION_M
    sx, sy = bx - ax, by - ay

    rx, ry = x2 - x1, y2 - y1
    denominator = cross_2d(rx, ry, sx, sy)
    if abs(denominator) < 1e-9:
        return None

    qpx, qpy = ax - x1, ay - y1
    t = cross_2d(qpx, qpy, sx, sy) / denominator
    u = cross_2d(qpx, qpy, rx, ry) / denominator
    if -1e-9 <= t <= 1.0 + 1e-9 and -1e-9 <= u <= 1.0 + 1e-9:
        return max(0.0, min(1.0, t))
    return None


def detect_coordinate_crossings(points: list[Point]) -> list[Crossing]:
    """Определяет все въезды и выезды по направлению пересечения линий КПП."""
    by_plate: dict[str, list[Point]] = defaultdict(list)
    for point in points:
        by_plate[point.plate].append(point)

    crossings: list[Crossing] = []
    for plate, vehicle_points in by_plate.items():
        vehicle_points.sort(key=lambda item: item.timestamp)
        last_by_gate_direction: dict[tuple[str, str], datetime] = {}
        for p1, p2 in zip(vehicle_points, vehicle_points[1:]):
            gap = (p2.timestamp - p1.timestamp).total_seconds()
            if gap <= 0 or gap > MAX_GPS_GAP_SECONDS:
                continue
            for gate_name, gate in GATE_SEGMENTS.items():
                side1 = signed_gate_distance_m(p1, gate)
                side2 = signed_gate_distance_m(p2, gate)

                if side1 >= MIN_SIDE_DISTANCE_M and side2 <= -MIN_SIDE_DISTANCE_M:
                    direction: Literal["entry", "exit"] = "entry"
                elif side1 <= -MIN_SIDE_DISTANCE_M and side2 >= MIN_SIDE_DISTANCE_M:
                    direction = "exit"
                else:
                    continue

                fraction = crossing_fraction(p1, p2, gate)
                if fraction is None:
                    continue
                timestamp = p1.timestamp + (p2.timestamp - p1.timestamp) * fraction
                key = (gate_name, direction)
                previous = last_by_gate_direction.get(key)
                if previous and (timestamp - previous).total_seconds() < CROSSING_COOLDOWN_SECONDS:
                    continue

                lat = p1.lat + (p2.lat - p1.lat) * fraction
                lon = p1.lon + (p2.lon - p1.lon) * fraction
                crossings.append(Crossing(
                    plate=plate,
                    timestamp=timestamp,
                    gate=gate_name,
                    direction=direction,
                    lat=lat,
                    lon=lon,
                ))
                last_by_gate_direction[key] = timestamp

    return sorted(crossings, key=lambda item: (item.timestamp, item.plate, item.gate))


def coordinate_entries_from_crossings(crossings: Iterable[Crossing]) -> list[Entry]:
    return first_per_vehicle_day(
        Entry(
            plate=crossing.plate,
            timestamp=crossing.timestamp,
            gate=crossing.gate,
            lat=crossing.lat,
            lon=crossing.lon,
            source=crossing.source,
        )
        for crossing in crossings
        if crossing.direction == "entry"
    )


def first_per_vehicle_day(entries: Iterable[Entry]) -> list[Entry]:
    first: dict[tuple[str, date], Entry] = {}
    for entry in entries:
        key = (entry.plate, entry.timestamp.date())
        if key not in first or entry.timestamp < first[key].timestamp:
            first[key] = entry
    return sorted(first.values(), key=lambda item: (item.timestamp, item.plate))


def merge_entries(coordinate_entries: list[Entry], geo_entries: list[Entry]) -> list[Entry]:
    """Координаты имеют приоритет; геозона добавляет только пропущенные случаи."""
    merged: dict[tuple[str, date], Entry] = {
        (entry.plate, entry.timestamp.date()): entry for entry in coordinate_entries
    }
    for event in geo_entries:
        key = (event.plate, event.timestamp.date())
        coordinate = merged.get(key)
        if coordinate is None:
            merged[key] = event
            continue
        # Если оба источника относятся к одному въезду, оставляем расчёт по линии.
        if abs((coordinate.timestamp - event.timestamp).total_seconds()) <= EVENT_MATCH_WINDOW_SECONDS:
            continue
        if event.timestamp < coordinate.timestamp:
            merged[key] = event
    return sorted(merged.values(), key=lambda item: (item.timestamp, item.plate))


def exclude_entries_after_previous_exits(
    entries: list[Entry],
    crossings: Iterable[Crossing],
    enabled: bool,
) -> tuple[list[Entry], list[tuple[Entry, Crossing]]]:
    """Исключает въезд, если до него в этот же день уже был выезд.

    Выезд определяется только по направлению координатного пересечения линии КПП.
    Событие «выход из геозоны KPP» не используется, потому что оно означает
    покидание небольшой зоны ворот, а не обязательно выезд с территории.
    """
    if not enabled:
        return entries, []

    exits_by_vehicle_day: dict[tuple[str, date], list[Crossing]] = defaultdict(list)
    for crossing in crossings:
        if crossing.direction == "exit":
            exits_by_vehicle_day[(crossing.plate, crossing.timestamp.date())].append(crossing)

    result: list[Entry] = []
    excluded: list[tuple[Entry, Crossing]] = []
    for entry in entries:
        previous_exits = [
            crossing
            for crossing in exits_by_vehicle_day.get((entry.plate, entry.timestamp.date()), [])
            if crossing.timestamp < entry.timestamp
        ]
        if previous_exits:
            excluded.append((entry, max(previous_exits, key=lambda item: item.timestamp)))
        else:
            result.append(entry)
    return result, excluded


def apply_filters(
    entries: list[Entry],
    roster: dict[str, VehicleInfo],
    filters: ReportFilters,
) -> tuple[list[Entry], dict[str, int]]:
    stats = {
        "input": len(entries),
        "no_roster": 0,
        "no_grade": 0,
        "grade": 0,
        "time": 0,
    }
    result: list[Entry] = []
    for entry in entries:
        info = roster.get(entry.plate)
        if info is None:
            stats["no_roster"] += 1
        if filters.grade_from is not None or filters.grade_to is not None:
            grade = parse_grade_number(info.grade if info else "")
            if grade is None:
                stats["no_grade"] += 1
                continue
            if filters.grade_from is not None and grade < filters.grade_from:
                stats["grade"] += 1
                continue
            if filters.grade_to is not None and grade > filters.grade_to:
                stats["grade"] += 1
                continue
        clock = entry.timestamp.time()
        if filters.time_from is not None and clock < filters.time_from:
            stats["time"] += 1
            continue
        if filters.time_to is not None and clock > filters.time_to:
            stats["time"] += 1
            continue
        result.append(entry)
    stats["output"] = len(result)
    return result, stats


def display_filter_time(value: Optional[time]) -> str:
    return value.strftime("%H:%M:%S") if value else "без ограничения"


def display_filter_grade(value: Optional[float]) -> str:
    if value is None:
        return "без ограничения"
    return str(int(value)) if value.is_integer() else str(value)


def create_report(
    output_path: Path,
    entries: list[Entry],
    roster: dict[str, VehicleInfo],
    coordinate_path: Path,
    roster_path: Path,
    geo_events_path: Optional[Path],
    filters: ReportFilters,
    stats: dict[str, int],
    coordinate_entry_count: int,
    coordinate_exit_count: int,
    geo_entry_count: int,
    before_previous_exit_filter: int,
    excluded_previous_exits: list[tuple[Entry, Crossing]],
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Первый въезд"

    headers = [
        "№", "Номерной знак", "Дата въезда", "Время въезда", "",
        "Геозона", "Марка, модель", "Грейд", "Водитель", "Должность", "Дирекция",
    ]
    sheet.append(headers)

    header_fill = PatternFill("solid", fgColor="4472C4")
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    missing_roster: set[str] = set()
    for number, entry in enumerate(entries, 1):
        info = roster.get(entry.plate, VehicleInfo(entry.plate))
        if entry.plate not in roster:
            missing_roster.add(entry.plate)
        sheet.append([
            number,
            entry.plate,
            entry.timestamp.date(),
            entry.timestamp.time(),
            "Показать на карте" if entry.lat is not None and entry.lon is not None else "",
            entry.gate,
            info.model,
            info.grade,
            info.driver,
            info.position,
            info.directorate,
        ])
        row_number = sheet.max_row
        if entry.lat is not None and entry.lon is not None:
            map_cell = sheet.cell(row_number, 5)
            map_cell.hyperlink = f"https://www.google.com/maps?q={entry.lat:.7f},{entry.lon:.7f}"
            map_cell.style = "Hyperlink"
        sheet.cell(row_number, 3).number_format = "dd.mm.yyyy"
        sheet.cell(row_number, 4).number_format = "hh:mm:ss"

    max_row = max(sheet.max_row, 1)
    for row in sheet.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=11):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(
                horizontal="center" if cell.column <= 8 else "left",
                vertical="center",
                wrap_text=cell.column >= 9,
            )

    widths = [6, 15, 13, 13, 20, 15, 24, 10, 38, 45, 55]
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.row_dimensions[1].height = 26
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:K{max_row}"

    check = workbook.create_sheet("Проверка")
    check_rows = [
        ("Параметр", "Значение"),
        ("Координатная выгрузка", str(coordinate_path)),
        ("Разнарядка", str(roster_path)),
        ("Дополнительный отчёт геозон", str(geo_events_path) if geo_events_path else "не использован"),
        ("Грейд от", display_filter_grade(filters.grade_from)),
        ("Грейд до", display_filter_grade(filters.grade_to)),
        ("Время от", display_filter_time(filters.time_from)),
        ("Время до", display_filter_time(filters.time_to)),
        ("Учитывать предыдущие выезды", "да" if filters.consider_previous_exits else "нет"),
        ("Максимальный разрыв GPS", f"{MAX_GPS_GAP_SECONDS // 60} мин"),
        ("Въездов по координатам", coordinate_entry_count),
        ("Выездов по координатам", coordinate_exit_count),
        ("Въездов в отчёте геозон", geo_entry_count),
        ("До проверки предыдущих выездов", before_previous_exit_filter),
        ("Исключено из-за предыдущего выезда", len(excluded_previous_exits)),
        ("До фильтрации по грейду и времени", stats.get("input", 0)),
        ("После всех фильтров", stats.get("output", 0)),
        ("Исключено без числового грейда", stats.get("no_grade", 0)),
        ("Исключено по грейду", stats.get("grade", 0)),
        ("Исключено по времени", stats.get("time", 0)),
    ]
    for row in check_rows:
        check.append(row)
    check["A1"].font = Font(bold=True, color="FFFFFF")
    check["B1"].font = Font(bold=True, color="FFFFFF")
    check["A1"].fill = header_fill
    check["B1"].fill = header_fill

    check.append([])
    check.append(["Госномера в отчёте, отсутствующие в разнарядке"])
    check.cell(check.max_row, 1).font = Font(bold=True)
    if missing_roster:
        for plate in sorted(missing_roster):
            check.append([plate])
    else:
        check.append(["Нет"])

    check.append([])
    check.append(["Исключённые въезды из-за предыдущего выезда"])
    check.cell(check.max_row, 1).font = Font(bold=True)
    check.append(["Госномер", "Время выезда", "КПП выезда", "Время исключённого въезда", "КПП въезда"])
    if excluded_previous_exits:
        for entry, previous_exit in excluded_previous_exits:
            check.append([
                entry.plate,
                previous_exit.timestamp,
                previous_exit.gate,
                entry.timestamp,
                entry.gate,
            ])
            check.cell(check.max_row, 2).number_format = "dd.mm.yyyy hh:mm:ss"
            check.cell(check.max_row, 4).number_format = "dd.mm.yyyy hh:mm:ss"
    else:
        check.append(["Нет"])

    check.column_dimensions["A"].width = 42
    check.column_dimensions["B"].width = 95
    check.column_dimensions["C"].width = 20
    check.column_dimensions["D"].width = 26
    check.column_dimensions["E"].width = 20
    for row in check.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    save_report_workbook(workbook, output_path)


def choose_file(title: str, optional: bool = False) -> Optional[Path]:
    from tkinter import Tk, filedialog

    root = Tk()
    root.withdraw()
    try:
        selected = filedialog.askopenfilename(
            title=title,
            filetypes=[("Excel", "*.xlsx *.xlsm"), ("Все файлы", "*.*")],
        )
    finally:
        root.destroy()
    if not selected:
        if optional:
            return None
        raise SystemExit("Файл не выбран")
    return Path(selected).resolve()


def ask_filters() -> ReportFilters:
    from tkinter import Tk, messagebox, simpledialog

    root = Tk()
    root.withdraw()
    try:
        grade_from = simpledialog.askstring(
            "Фильтр", "Грейд от, например 7. Пусто — без фильтра:", parent=root
        )
        grade_to = simpledialog.askstring(
            "Фильтр", "Грейд до, например 14. Пусто — без фильтра:", parent=root
        )
        time_from = simpledialog.askstring(
            "Фильтр", "Время от, например 07:00. Пусто — без фильтра:", parent=root
        )
        time_to = simpledialog.askstring(
            "Фильтр", "Время до, например 09:00. Пусто — без фильтра:", parent=root
        )
        consider_previous_exits = messagebox.askyesno(
            "Дополнительное условие",
            "Учитывать предыдущие выезды?\n\n"
            "Да — если в этот день до въезда автомобиль уже выезжал с территории, "
            "такой повторный въезд не попадёт в отчёт.",
            parent=root,
        )
    finally:
        root.destroy()
    return ReportFilters(
        parse_grade_number(grade_from),
        parse_grade_number(grade_to),
        parse_clock(time_from),
        parse_clock(time_to),
        consider_previous_exits,
    )


def main() -> None:
    parser = argparse.ArgumentParser(description="Отчёт по первому въезду через KPP 4 / KPP 5")
    parser.add_argument("arvento", nargs="?", help="Координатная выгрузка Arvento")
    parser.add_argument("roster", nargs="?", help="Excel-разнарядка")
    parser.add_argument("output", nargs="?", help="Выходной XLSX")
    parser.add_argument("--geo-events", help="Опциональный отчёт «Сигнал географического региона»")
    parser.add_argument("--grade-from")
    parser.add_argument("--grade-to")
    parser.add_argument("--time-from")
    parser.add_argument("--time-to")
    parser.add_argument(
        "--consider-previous-exits",
        action=argparse.BooleanOptionalAction,
        default=None,
        help=(
            "Исключать въезд, если в этот же день до него был координатный выезд "
            "через KPP 4 или KPP 5"
        ),
    )
    parser.add_argument("--no-filter-dialog", action="store_true")
    args = parser.parse_args()

    arvento_path = (
        Path(args.arvento).resolve()
        if args.arvento
        else choose_file("Выберите координатную выгрузку Arvento")
    )
    roster_path = (
        Path(args.roster).resolve()
        if args.roster
        else choose_file("Выберите файл разнарядки")
    )
    assert arvento_path is not None and roster_path is not None
    output_path = (
        Path(args.output).resolve()
        if args.output
        else arvento_path.with_name(f"Первый въезд {arvento_path.stem}.xlsx")
    )
    geo_events_path = Path(args.geo_events).resolve() if args.geo_events else None

    cli_options = (
        args.no_filter_dialog
        or any(value is not None for value in (args.grade_from, args.grade_to, args.time_from, args.time_to))
        or args.consider_previous_exits is not None
    )
    if cli_options:
        filters = ReportFilters(
            parse_grade_number(args.grade_from),
            parse_grade_number(args.grade_to),
            parse_clock(args.time_from),
            parse_clock(args.time_to),
            bool(args.consider_previous_exits),
        )
    else:
        filters = ask_filters()

    if (
        filters.grade_from is not None
        and filters.grade_to is not None
        and filters.grade_from > filters.grade_to
    ):
        raise ValueError("Минимальный грейд больше максимального")
    if (
        filters.time_from is not None
        and filters.time_to is not None
        and filters.time_from > filters.time_to
    ):
        raise ValueError("Начальное время позже конечного")

    print(f"Координатная выгрузка: {arvento_path}")
    print(f"Разнарядка: {roster_path}")
    print(
        "Учитывать предыдущие выезды: "
        + ("да" if filters.consider_previous_exits else "нет")
    )
    if geo_events_path:
        print(f"Дополнительный отчёт геозон: {geo_events_path}")

    roster = load_roster(roster_path)
    points = load_coordinate_points(arvento_path)
    crossings = detect_coordinate_crossings(points)
    coordinate_entries = coordinate_entries_from_crossings(crossings)
    geo_entries = load_geo_entries(geo_events_path) if geo_events_path else []
    merged_entries = merge_entries(coordinate_entries, geo_entries)

    entries_after_exit_check, excluded_previous_exits = exclude_entries_after_previous_exits(
        merged_entries,
        crossings,
        filters.consider_previous_exits,
    )
    filtered_entries, stats = apply_filters(entries_after_exit_check, roster, filters)

    coordinate_entry_count = sum(crossing.direction == "entry" for crossing in crossings)
    coordinate_exit_count = sum(crossing.direction == "exit" for crossing in crossings)

    create_report(
        output_path=output_path,
        entries=filtered_entries,
        roster=roster,
        coordinate_path=arvento_path,
        roster_path=roster_path,
        geo_events_path=geo_events_path,
        filters=filters,
        stats=stats,
        coordinate_entry_count=coordinate_entry_count,
        coordinate_exit_count=coordinate_exit_count,
        geo_entry_count=len(geo_entries),
        before_previous_exit_filter=len(merged_entries),
        excluded_previous_exits=excluded_previous_exits,
    )

    print(f"Готово: {output_path}")
    print(f"Первых въездов до проверки предыдущих выездов: {len(merged_entries)}")
    print(f"Исключено из-за предыдущего выезда: {len(excluded_previous_exits)}")
    print(f"Строк после всех фильтров: {len(filtered_entries)}")


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"ОШИБКА: {exc}")
        raise
