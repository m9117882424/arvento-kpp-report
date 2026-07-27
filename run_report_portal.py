#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""Canonical ASGI entrypoint for the report portal.

The compatibility implementation remains in ``report_portal.py``. This module
adds current user-facing names, validated speed controls, grouped violation
preview, clickable map links, a client-side plate filter and database freshness
information.
"""

from __future__ import annotations

import base64
import tempfile
from datetime import date, datetime
from pathlib import Path
from typing import Any

import psycopg
from fastapi import File, Form, HTTPException, UploadFile
from fastapi.concurrency import run_in_threadpool
from openpyxl import load_workbook

import report_portal as implementation
from map_links import google_maps_url, parse_coordinate_pair
from speed_violation_report import (
    DEFAULT_OUTSIDE_SPEED_THRESHOLD_KMH,
    DEFAULT_SITE_SPEED_THRESHOLD_KMH,
    OUTSIDE_SHEET_NAME,
    SITE_SHEET_NAME,
    TURN_SHEET_NAME,
    validate_speed_thresholds,
)


_original_generate_report = implementation.generate_report


# User-facing report name and configurable speed thresholds.
implementation.HTML = implementation.HTML.replace(
    '<option value="violation">Запрещённый поворот</option>',
    '<option value="violation">Нарушения</option>',
)
implementation.HTML = implementation.HTML.replace(
    '        <div class="kpp-only"><label for="gradeFrom">Грейд от</label>',
    f'''        <div class="violation-only hidden">
          <label for="siteSpeedThreshold">Порог на площадке, км/ч</label>
          <input id="siteSpeedThreshold" name="site_speed_threshold" type="number"
                 min="5" max="200" step="0.1" value="{DEFAULT_SITE_SPEED_THRESHOLD_KMH:g}">
        </div>
        <div class="violation-only hidden">
          <label for="outsideSpeedThreshold">Порог вне площадки, км/ч</label>
          <input id="outsideSpeedThreshold" name="outside_speed_threshold" type="number"
                 min="20" max="250" step="0.1" value="{DEFAULT_OUTSIDE_SPEED_THRESHOLD_KMH:g}">
        </div>
        <div class="violation-only hidden" style="grid-column: span 2">
          <div class="note">Нарушение засчитывается только по плавной последовательности минимум из трёх GPS-точек продолжительностью не менее 10 секунд. Одиночные скачки скорости исключаются.</div>
        </div>
        <div class="kpp-only"><label for="gradeFrom">Грейд от</label>''',
)
implementation.HTML = implementation.HTML.replace(
    '    <div class="table-wrap"><table id="resultTable"></table></div>',
    '''    <div id="plateFilterBox" class="hidden" style="display:flex; gap:12px; align-items:end; margin-bottom:14px; max-width:520px">
      <div style="flex:1">
        <label for="plateFilter">Фильтр по госномеру</label>
        <select id="plateFilter"><option value="">Все госномера</option></select>
      </div>
      <div id="plateFilterCount" class="muted" style="padding-bottom:10px"></div>
    </div>
    <div class="table-wrap"><table id="resultTable"></table></div>''',
)

# Database freshness indicator in the upper-right corner.
implementation.HTML = implementation.HTML.replace(
    '<div class="muted">Данные берутся напрямую из PostgreSQL. Временные CSV и Excel после формирования удаляются.</div>',
    '''<div class="muted">Данные берутся напрямую из PostgreSQL. Временные CSV и Excel после формирования удаляются.</div>
    <div id="dbStatus" class="db-status loading" title="Состояние синхронизации GPS-данных">
      <div class="db-status-head"><span class="db-dot"></span><strong>Последняя запись в БД</strong></div>
      <div id="dbStatusTime" class="db-status-time">Проверка…</div>
      <div id="dbStatusMeta" class="db-status-meta"></div>
      <div id="dbCompleteness" class="db-completeness"></div>
    </div>''',
)
implementation.HTML = implementation.HTML.replace(
    "    tr:last-child td { border-bottom: 0; }",
    "    tr:last-child td { border-bottom: 0; }\n"
    "    tr.plate-start td { border-top: 2px solid #98a2b3; }\n"
    "    .card { position: relative; }\n"
    "    .db-status { position:absolute; top:18px; right:18px; width:310px; padding:11px 13px; border:1px solid #d0d5dd; border-radius:10px; background:#f8fafc; box-sizing:border-box; font-size:12px; }\n"
    "    .db-status-head { display:flex; gap:7px; align-items:center; margin-bottom:4px; }\n"
    "    .db-dot { width:9px; height:9px; border-radius:50%; background:#98a2b3; flex:0 0 auto; }\n"
    "    .db-status-time { font-size:13px; font-weight:700; }\n"
    "    .db-status-meta, .db-completeness { margin-top:3px; color:#667085; line-height:1.35; }\n"
    "    .db-status.fresh .db-dot { background:#12b76a; }\n"
    "    .db-status.warning .db-dot { background:#f79009; }\n"
    "    .db-status.stale .db-dot, .db-status.error .db-dot { background:#d92d20; }\n"
    "    .map-link { color:#1663d6; text-decoration:none; font-weight:650; }\n"
    "    .map-link:hover { text-decoration:underline; }\n"
    "    @media (max-width: 900px) { .db-status { position:static; width:100%; margin-top:14px; } }",
)
implementation.HTML = implementation.HTML.replace(
    "  const isEfficiency = type === 'efficiency';\n  const needsRoster = isKpp || isEfficiency;",
    "  const isEfficiency = type === 'efficiency';\n  const isViolation = type === 'violation';\n  const needsRoster = isKpp || isEfficiency;",
)
implementation.HTML = implementation.HTML.replace(
    "  document.querySelectorAll('.efficiency-only').forEach(x => x.classList.toggle('hidden', !isEfficiency));",
    "  document.querySelectorAll('.efficiency-only').forEach(x => x.classList.toggle('hidden', !isEfficiency));\n"
    "  document.querySelectorAll('.violation-only').forEach(x => x.classList.toggle('hidden', !isViolation));\n"
    "  document.getElementById('siteSpeedThreshold').required = isViolation;\n"
    "  document.getElementById('outsideSpeedThreshold').required = isViolation;",
)
implementation.HTML = implementation.HTML.replace(
    "const previewNote = document.getElementById('previewNote');\nlet excelBase64 = '';",
    "const previewNote = document.getElementById('previewNote');\n"
    "const plateFilterBox = document.getElementById('plateFilterBox');\n"
    "const plateFilter = document.getElementById('plateFilter');\n"
    "const plateFilterCount = document.getElementById('plateFilterCount');\n"
    "const dbStatus = document.getElementById('dbStatus');\n"
    "const dbStatusTime = document.getElementById('dbStatusTime');\n"
    "const dbStatusMeta = document.getElementById('dbStatusMeta');\n"
    "const dbCompleteness = document.getElementById('dbCompleteness');\n"
    "let databaseStatusData = null;\n"
    "let tableColumns = [];\nlet tableRows = [];\nlet excelBase64 = '';",
)
implementation.HTML = implementation.HTML.replace(
    "function renderTable(columns, rows) {\n"
    "  const head = `<thead><tr>${columns.map(value => `<th>${esc(value)}</th>`).join('')}</tr></thead>`;\n"
    "  const body = `<tbody>${rows.map(row => `<tr>${row.map(value => `<td>${esc(value)}</td>`).join('')}</tr>`).join('')}</tbody>`;\n"
    "  table.innerHTML = head + body;\n"
    "}",
    "function drawFilteredTable() {\n"
    "  const plateIndex = tableColumns.indexOf('Госномер');\n"
    "  const addressIndex = tableColumns.indexOf('Адрес');\n"
    "  const mapIndex = tableColumns.indexOf('Карта');\n"
    "  const visibleIndexes = tableColumns.map((_, index) => index).filter(index => index !== mapIndex);\n"
    "  const selected = plateFilter.value;\n"
    "  const rows = selected && plateIndex >= 0 ? tableRows.filter(row => String(row[plateIndex] ?? '') === selected) : tableRows;\n"
    "  const head = `<thead><tr>${visibleIndexes.map(index => `<th>${esc(tableColumns[index])}</th>`).join('')}</tr></thead>`;\n"
    "  let previousPlate = null;\n"
    "  const bodyRows = rows.map(row => {\n"
    "    const plate = plateIndex >= 0 ? String(row[plateIndex] ?? '') : '';\n"
    "    const className = plate && plate !== previousPlate ? ' class=\"plate-start\"' : '';\n"
    "    previousPlate = plate;\n"
    "    const cells = visibleIndexes.map(index => {\n"
    "      const value = row[index];\n"
    "      if (index === addressIndex && mapIndex >= 0 && row[mapIndex]) {\n"
    "        const label = value || 'Открыть на карте';\n"
    "        return `<td><a class=\"map-link\" href=\"${esc(row[mapIndex])}\" target=\"_blank\" rel=\"noopener noreferrer\">${esc(label)}</a></td>`;\n"
    "      }\n"
    "      return `<td>${esc(value)}</td>`;\n"
    "    }).join('');\n"
    "    return `<tr${className}>${cells}</tr>`;\n"
    "  }).join('');\n"
    "  table.innerHTML = head + `<tbody>${bodyRows}</tbody>`;\n"
    "  plateFilterCount.textContent = `Показано: ${rows.length} из ${tableRows.length}`;\n"
    "}\n"
    "function renderTable(columns, rows) {\n"
    "  tableColumns = columns;\n"
    "  tableRows = rows;\n"
    "  const plateIndex = columns.indexOf('Госномер');\n"
    "  if (plateIndex >= 0) {\n"
    "    const plates = [...new Set(rows.map(row => String(row[plateIndex] ?? '')).filter(Boolean))].sort();\n"
    "    plateFilter.innerHTML = '<option value=\"\">Все госномера</option>' + plates.map(value => `<option value=\"${esc(value)}\">${esc(value)}</option>`).join('');\n"
    "    plateFilterBox.classList.remove('hidden');\n"
    "  } else {\n"
    "    plateFilterBox.classList.add('hidden');\n"
    "  }\n"
    "  drawFilteredTable();\n"
    "}\n"
    "plateFilter.addEventListener('change', drawFilteredTable);",
)
implementation.HTML = implementation.HTML.replace(
    "loadDates();",
    "loadDates();\n\n"
    "function ageText(minutes) {\n"
    "  if (minutes == null) return '';\n"
    "  if (minutes < 1) return 'менее минуты назад';\n"
    "  if (minutes < 60) return `${Math.round(minutes)} мин назад`;\n"
    "  const hours = Math.floor(minutes / 60);\n"
    "  const rest = Math.round(minutes % 60);\n"
    "  return rest ? `${hours} ч ${rest} мин назад` : `${hours} ч назад`;\n"
    "}\n"
    "function updateDatabaseCompleteness() {\n"
    "  if (!databaseStatusData || !databaseStatusData.latest_date) return;\n"
    "  const targetDate = typeSelect.value === 'efficiency' && endDateInput.value ? endDateInput.value : dateInput.value;\n"
    "  const latestDate = databaseStatusData.latest_date;\n"
    "  const today = databaseStatusData.server_date;\n"
    "  let text = '';\n"
    "  if (!targetDate) {\n"
    "    text = '';\n"
    "  } else if (targetDate > latestDate) {\n"
    "    text = 'За выбранную дату записей ещё нет.';\n"
    "  } else if (targetDate < latestDate) {\n"
    "    text = 'Выбранный день закрыт: в БД уже есть данные более поздней даты.';\n"
    "  } else if (targetDate === today) {\n"
    "    text = 'Отчёт за сегодня предварительный до окончания суток.';\n"
    "  } else {\n"
    "    text = 'Это последняя доступная дата; полнота дня не подтверждена.';\n"
    "  }\n"
    "  if (databaseStatusData.status === 'warning' || databaseStatusData.status === 'stale') {\n"
    "    text += ' Синхронизация запаздывает — лучше подождать.';\n"
    "  }\n"
    "  dbCompleteness.textContent = text;\n"
    "}\n"
    "async function loadDatabaseStatus() {\n"
    "  try {\n"
    "    const response = await fetch('/api/database-status', {cache: 'no-store'});\n"
    "    const data = await response.json();\n"
    "    if (!response.ok) throw new Error(data.detail || 'Ошибка получения состояния БД');\n"
    "    databaseStatusData = data;\n"
    "    dbStatus.className = `db-status ${data.status}`;\n"
    "    dbStatusTime.textContent = data.latest_display || 'Записей нет';\n"
    "    const parts = [];\n"
    "    if (data.age_minutes != null) parts.push(ageText(data.age_minutes));\n"
    "    parts.push(`сегодня точек: ${Number(data.today_points || 0).toLocaleString('ru-RU')}`);\n"
    "    parts.push(`ТС: ${Number(data.today_vehicles || 0).toLocaleString('ru-RU')}`);\n"
    "    dbStatusMeta.textContent = `${data.label}. ${parts.join(' · ')}`;\n"
    "    updateDatabaseCompleteness();\n"
    "  } catch (error) {\n"
    "    dbStatus.className = 'db-status error';\n"
    "    dbStatusTime.textContent = 'Не удалось проверить БД';\n"
    "    dbStatusMeta.textContent = error.message;\n"
    "    dbCompleteness.textContent = '';\n"
    "  }\n"
    "}\n"
    "dateInput.addEventListener('change', updateDatabaseCompleteness);\n"
    "endDateInput.addEventListener('change', updateDatabaseCompleteness);\n"
    "typeSelect.addEventListener('change', updateDatabaseCompleteness);\n"
    "loadDatabaseStatus();\n"
    "setInterval(loadDatabaseStatus, 60000);",
)
implementation.HTML = implementation.HTML.replace(
    "  const data = new FormData(form);",
    "  if (typeSelect.value === 'violation') {\n"
    "    const site = Number(document.getElementById('siteSpeedThreshold').value);\n"
    "    const outside = Number(document.getElementById('outsideSpeedThreshold').value);\n"
    "    if (!Number.isFinite(site) || site < 5 || site > 200) {\n"
    "      statusBox.className = 'status error';\n"
    "      statusBox.textContent = 'Порог на площадке должен быть от 5 до 200 км/ч.';\n"
    "      generateBtn.disabled = false;\n"
    "      return;\n"
    "    }\n"
    "    if (!Number.isFinite(outside) || outside < 20 || outside > 250) {\n"
    "      statusBox.className = 'status error';\n"
    "      statusBox.textContent = 'Порог вне площадки должен быть от 20 до 250 км/ч.';\n"
    "      generateBtn.disabled = false;\n"
    "      return;\n"
    "    }\n"
    "    if (outside < site) {\n"
    "      statusBox.className = 'status error';\n"
    "      statusBox.textContent = 'Порог вне площадки не может быть ниже порога на площадке.';\n"
    "      generateBtn.disabled = false;\n"
    "      return;\n"
    "    }\n"
    "  }\n"
    "  const data = new FormData(form);",
)
implementation.HTML = implementation.HTML.replace(
    "fetch('/api/generate', {method: 'POST', body: data})",
    "fetch('/api/generate-v2', {method: 'POST', body: data})",
)


def parse_threshold(value: str, label: str) -> float:
    text = value.strip().replace(",", ".")
    if not text:
        raise ValueError(f"Не заполнено поле «{label}»")
    try:
        return float(text)
    except ValueError as exc:
        raise ValueError(f"Поле «{label}» должно содержать число") from exc


def _header_map(sheet) -> dict[str, int]:
    first = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), tuple())
    return {str(value or "").strip(): index for index, value in enumerate(first)}


def _value(row: tuple[Any, ...], headers: dict[str, int], name: str) -> Any:
    index = headers.get(name)
    if index is None or index >= len(row):
        return None
    return row[index]


def _as_number(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _map_url(value: Any) -> str:
    coordinates = parse_coordinate_pair(value)
    if coordinates is None:
        return ""
    return google_maps_url(*coordinates)


def violation_web_preview(path: Path) -> tuple[list[str], list[list[Any]], int]:
    """Build one detailed, plate-grouped table from all violation sheets."""
    columns = [
        "Госномер",
        "Тип нарушения",
        "Дата",
        "Начало",
        "Окончание",
        "Максимальная скорость, км/ч",
        "Порог, км/ч",
        "Адрес",
        "Карта",
    ]
    records: list[tuple[str, datetime | None, list[Any]]] = []
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet_name in (TURN_SHEET_NAME, SITE_SHEET_NAME, OUTSIDE_SHEET_NAME):
            if sheet_name not in workbook.sheetnames:
                continue
            sheet = workbook[sheet_name]
            headers = _header_map(sheet)
            for row in sheet.iter_rows(min_row=2, values_only=True):
                plate = str(_value(row, headers, "Госномер") or "").strip()
                if not plate:
                    continue

                if sheet_name == TURN_SHEET_NAME:
                    start = _value(row, headers, "Начало прохода")
                    finish = _value(row, headers, "Окончание прохода")
                    start_speed = _as_number(_value(row, headers, "Скорость в начале"))
                    finish_speed = _as_number(_value(row, headers, "Скорость в конце"))
                    speeds = [value for value in (start_speed, finish_speed) if value is not None]
                    max_speed = max(speeds) if speeds else None
                    address_parts = [
                        str(_value(row, headers, "Адрес начала") or "").strip(),
                        str(_value(row, headers, "Адрес окончания") or "").strip(),
                    ]
                    address = " → ".join(part for part in address_parts if part)
                    map_url = _map_url(
                        _value(row, headers, "Координаты начала")
                        or _value(row, headers, "Координаты окончания")
                    )
                    threshold = None
                else:
                    start = _value(row, headers, "Начало нарушения")
                    finish = _value(row, headers, "Окончание нарушения")
                    max_speed = _as_number(_value(row, headers, "Максимальная скорость, км/ч"))
                    threshold = _as_number(_value(row, headers, "Порог фиксации, км/ч"))
                    address = str(_value(row, headers, "Адрес максимума") or "").strip()
                    map_url = _map_url(_value(row, headers, "Координаты максимума"))

                event_date = start.date() if isinstance(start, datetime) else _value(row, headers, "Дата")
                display = [
                    plate,
                    sheet_name,
                    implementation.json_cell(event_date),
                    implementation.json_cell(start),
                    implementation.json_cell(finish),
                    implementation.json_cell(max_speed),
                    implementation.json_cell(threshold),
                    address,
                    map_url,
                ]
                records.append((plate, start if isinstance(start, datetime) else None, display))
    finally:
        workbook.close()

    records.sort(key=lambda item: (item[0], item[1] or datetime.min, item[2][1]))
    rows = [item[2] for item in records]
    return columns, rows, len(rows)


def generate_report_with_thresholds(
    report_type: str,
    report_day: date,
    report_end_day: date | None,
    roster_bytes: bytes | None,
    roster_filename: str,
    roster_suffix: str,
    grade_from: str,
    grade_to: str,
    time_from: str,
    time_to: str,
    consider_previous_exits: bool,
    site_speed_threshold: float,
    outside_speed_threshold: float,
) -> dict[str, Any]:
    if report_type != "violation":
        return _original_generate_report(
            report_type,
            report_day,
            report_end_day,
            roster_bytes,
            roster_filename,
            roster_suffix,
            grade_from,
            grade_to,
            time_from,
            time_to,
            consider_previous_exits,
        )

    implementation.validate_canonical_scripts()
    site_threshold, outside_threshold = validate_speed_thresholds(
        site_speed_threshold,
        outside_speed_threshold,
    )

    with tempfile.TemporaryDirectory(prefix="arvento_report_portal_") as temp_name:
        temp_dir = Path(temp_name)
        csv_path = temp_dir / f"gps_{report_day.isoformat()}.csv"
        gps_count = implementation.export_period_to_csv(report_day, report_day, csv_path)
        if gps_count == 0:
            raise ValueError("За выбранную дату GPS-точки отсутствуют")

        filename = f"Нарушения_{report_day.isoformat()}.xlsx"
        output_path = temp_dir / filename
        log = implementation.run_command([
            implementation.sys.executable,
            str(implementation.CANONICAL_REPORT_SCRIPTS[report_type]),
            str(csv_path),
            str(output_path),
            "--site-speed-threshold",
            f"{site_threshold:g}",
            "--outside-speed-threshold",
            f"{outside_threshold:g}",
        ])
        if not output_path.exists():
            raise RuntimeError("Построитель не создал Excel-файл")

        columns, rows, total_rows = violation_web_preview(output_path)
        plate_count = len({str(row[0]) for row in rows if row and row[0]})
        return {
            "filename": filename,
            "columns": columns,
            "rows": rows,
            "preview_truncated": False,
            "excel_base64": base64.b64encode(output_path.read_bytes()).decode("ascii"),
            "summary": {
                "Отчёт": "Нарушения",
                "Период": report_day.strftime("%d.%m.%Y"),
                "Порог на площадке": f"{site_threshold:g} км/ч",
                "Порог вне площадки": f"{outside_threshold:g} км/ч",
                "Госномеров": plate_count,
                "Нарушений": total_rows,
                "GPS-точек": gps_count,
            },
            "log": log,
        }


app = implementation.app


@app.get("/api/database-status")
def database_status() -> dict[str, Any]:
    query = """
        WITH bounds AS (
            SELECT
                date_trunc('day', now() AT TIME ZONE 'Europe/Istanbul')
                    AT TIME ZONE 'Europe/Istanbul' AS day_start,
                (date_trunc('day', now() AT TIME ZONE 'Europe/Istanbul') + interval '1 day')
                    AT TIME ZONE 'Europe/Istanbul' AS day_end
        )
        SELECT
            max(g.event_time) AS latest_event_time,
            now() AS database_time,
            count(*) FILTER (
                WHERE g.event_time >= bounds.day_start
                  AND g.event_time < bounds.day_end
            ) AS today_points,
            count(DISTINCT g.normalized_plate) FILTER (
                WHERE g.event_time >= bounds.day_start
                  AND g.event_time < bounds.day_end
            ) AS today_vehicles
        FROM gps_points g
        CROSS JOIN bounds
    """
    with psycopg.connect(implementation.db_url()) as connection:
        latest, database_time, today_points, today_vehicles = connection.execute(query).fetchone()

    if latest is None:
        return {
            "status": "empty",
            "label": "В БД пока нет GPS-записей",
            "latest_display": None,
            "latest_date": None,
            "server_date": database_time.astimezone(implementation.TZ).date().isoformat(),
            "age_minutes": None,
            "today_points": int(today_points or 0),
            "today_vehicles": int(today_vehicles or 0),
        }

    latest_local = latest.astimezone(implementation.TZ)
    database_local = database_time.astimezone(implementation.TZ)
    age_minutes = max(0.0, (database_time - latest).total_seconds() / 60.0)
    if age_minutes <= 60:
        status = "fresh"
        label = "Данные актуальны"
    elif age_minutes <= 180:
        status = "warning"
        label = "Есть задержка синхронизации"
    else:
        status = "stale"
        label = "Данные давно не обновлялись"

    return {
        "status": status,
        "label": label,
        "latest_display": latest_local.strftime("%d.%m.%Y %H:%M:%S"),
        "latest_iso": latest_local.isoformat(),
        "latest_date": latest_local.date().isoformat(),
        "server_date": database_local.date().isoformat(),
        "age_minutes": round(age_minutes, 1),
        "today_points": int(today_points or 0),
        "today_vehicles": int(today_vehicles or 0),
    }


@app.post("/api/generate-v2")
async def api_generate_v2(
    report_type: str = Form(...),
    report_date: str = Form(...),
    report_end_date: str = Form(default=""),
    roster: UploadFile | None = File(default=None),
    grade_from: str = Form(default=""),
    grade_to: str = Form(default=""),
    time_from: str = Form(default=""),
    time_to: str = Form(default=""),
    consider_previous_exits: bool = Form(default=False),
    site_speed_threshold: str = Form(default=str(DEFAULT_SITE_SPEED_THRESHOLD_KMH)),
    outside_speed_threshold: str = Form(default=str(DEFAULT_OUTSIDE_SPEED_THRESHOLD_KMH)),
) -> dict[str, Any]:
    try:
        day = implementation.parse_report_date(report_date)
        end_day = (
            implementation.parse_report_date(report_end_date)
            if report_end_date.strip()
            else None
        )

        site_threshold = parse_threshold(site_speed_threshold, "Порог на площадке")
        outside_threshold = parse_threshold(outside_speed_threshold, "Порог вне площадки")
        if report_type == "violation":
            site_threshold, outside_threshold = validate_speed_thresholds(
                site_threshold,
                outside_threshold,
            )

        roster_bytes: bytes | None = None
        roster_filename = "roster.xlsx"
        roster_suffix = ".xlsx"
        if roster is not None and roster.filename:
            roster_filename = Path(roster.filename).name
            roster_suffix = Path(roster_filename).suffix.lower()
            if roster_suffix not in {".xlsx", ".xlsm"}:
                raise ValueError("Разнарядка должна быть в формате XLSX или XLSM")
            roster_bytes = await roster.read(implementation.MAX_ROSTER_BYTES + 1)
            if len(roster_bytes) > implementation.MAX_ROSTER_BYTES:
                raise ValueError("Размер разнарядки превышает 25 МБ")

        return await run_in_threadpool(
            generate_report_with_thresholds,
            report_type,
            day,
            end_day,
            roster_bytes,
            roster_filename,
            roster_suffix,
            grade_from,
            grade_to,
            time_from,
            time_to,
            consider_previous_exits,
            site_threshold,
            outside_threshold,
        )
    except (ValueError, RuntimeError) as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


__all__ = ["app"]
