#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Authoritative mileage policy and mileage-review annotations.

VehicleDistanceReport remains the reported vehicle-day mileage even when the
coordinate track differs substantially. Coordinates are used only to distribute
the authoritative total between route/site segments and to identify rows that
require manual verification.
"""
from __future__ import annotations

import math
from copy import copy
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any, Iterator, Mapping, Sequence
from zoneinfo import ZoneInfo

import psycopg
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from excel_formatting import save_report_workbook

import consolidated_mileage_logic as mileage
import consolidated_report as core
from arvento_io import Point
from roster_registry import normalize_plate

TZ = ZoneInfo("Europe/Istanbul")
REVIEW_HEADER = "Проверка пробега"
REVIEW_VALUE = "Проверить"
REVIEW_SHEET = "Проверка пробега"
REVIEW_REASON_GAP = "Пробег VehicleDistanceReport существенно выше координатного"
REVIEW_REASON_NO_COORDINATES = "Нет пригодного координатного пробега"

_POLICY_PATCHED = False
_INCREMENTAL_PATCHED = False


@dataclass(frozen=True, slots=True)
class MileageReviewCandidate:
    report_day: date
    normalized_plate: str
    authoritative_km: float
    coordinate_km: float
    gap_km: float
    gap_percent: float | None
    reason: str


def _valid_nonnegative(value: object) -> float | None:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    if not math.isfinite(number) or number < 0:
        return None
    return number


def _valid_source_distance(value: object) -> float | None:
    number = _valid_nonnegative(value)
    if number is None or number > mileage.MAX_SEGMENT_DISTANCE_KM:
        return None
    return number


def _assign_distances(points: Sequence[Point], values: Sequence[float]) -> None:
    if len(points) != len(values):
        raise ValueError("Количество точек и подготовленных расстояний различается")
    for point, value in zip(points, values):
        point.source_distance = max(0.0, float(value))
        point.prepared_distance = True


def mileage_requires_review(authoritative_km: float, coordinate_km: float) -> bool:
    """Return True when the VehicleDistanceReport total requires manual review."""
    authoritative = _valid_nonnegative(authoritative_km)
    coordinate = _valid_nonnegative(coordinate_km)
    if authoritative is None:
        return False
    if coordinate is None or coordinate <= 0:
        return authoritative > mileage.HYBRID_ABSOLUTE_GAP_KM
    return (
        authoritative - coordinate > mileage.HYBRID_ABSOLUTE_GAP_KM
        and authoritative > coordinate * mileage.HYBRID_RATIO
    )


def authoritative_vehicle_day_distances(
    points: Sequence[Point],
    authoritative_distance_km: float | None = None,
) -> str:
    """Prepare a vehicle-day track without replacing authoritative mileage.

    Coordinates distribute VehicleDistanceReport mileage across segments. A
    large discrepancy is handled separately as a review flag; it never changes
    the reported total.
    """
    authoritative_km = _valid_nonnegative(authoritative_distance_km)

    if authoritative_km is not None:
        coordinate_values = mileage.coordinate_segment_distances(points)
        coordinate_km = sum(coordinate_values)
        if coordinate_km > 0:
            scale = authoritative_km / coordinate_km
            _assign_distances(points, [value * scale for value in coordinate_values])
            return "authoritative_scaled_coordinates"

        source_values = [
            _valid_source_distance(point.source_distance) or 0.0
            for point in points
        ]
        source_km = sum(source_values)
        if source_km > 0:
            scale = authoritative_km / source_km
            _assign_distances(points, [value * scale for value in source_values])
            return "authoritative_scaled_source"

        fallback_values = [0.0] * len(points)
        if len(fallback_values) >= 2:
            fallback_values[-1] = authoritative_km
        _assign_distances(points, fallback_values)
        return "authoritative_single_segment"

    has_arvento_distance = any(
        _valid_source_distance(point.source_distance) is not None
        for point in points
    )
    if not has_arvento_distance:
        return "coordinate_only"

    source_values = [
        _valid_source_distance(point.source_distance) or 0.0
        for point in points
    ]
    _assign_distances(points, source_values)
    return "general_report_distance"


def apply_authoritative_mileage_policy() -> None:
    """Patch the shared mileage normalizer once for the current process."""
    global _POLICY_PATCHED
    if _POLICY_PATCHED:
        return
    mileage.normalize_vehicle_day_distances = authoritative_vehicle_day_distances
    _POLICY_PATCHED = True


def apply_incremental_mileage_policy() -> None:
    """Apply the same authoritative policy to intraday incremental refreshes."""
    global _INCREMENTAL_PATCHED
    if _INCREMENTAL_PATCHED:
        return

    import consolidated_incremental_cache as incremental

    original_iter = incremental.iter_database_tracks

    def iter_tracks(
        database_url: str,
        report_day: date,
        normalized_plates: Sequence[str],
    ) -> Iterator[tuple[str, list[Point]]]:
        authoritative = mileage.load_authoritative_daily_distances(
            database_url,
            report_day,
            report_day,
        )
        for display_plate, points in original_iter(
            database_url,
            report_day,
            normalized_plates,
        ):
            track = core.sanitize_position_outliers(points)
            authoritative_vehicle_day_distances(
                track,
                authoritative.get((report_day, normalize_plate(display_plate))),
            )
            yield display_plate, track

    incremental.iter_database_tracks = iter_tracks
    _INCREMENTAL_PATCHED = True


def ensure_review_schema(connection: psycopg.Connection) -> None:
    with connection.cursor() as cursor:
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS mileage_review_candidates (
                report_day DATE NOT NULL,
                normalized_plate TEXT NOT NULL,
                authoritative_km DOUBLE PRECISION NOT NULL,
                coordinate_km DOUBLE PRECISION NOT NULL,
                gap_km DOUBLE PRECISION NOT NULL,
                gap_percent DOUBLE PRECISION,
                reason TEXT NOT NULL,
                calculated_at TIMESTAMPTZ NOT NULL DEFAULT now(),
                PRIMARY KEY (report_day, normalized_plate)
            )
            """
        )
        cursor.execute(
            """
            CREATE INDEX IF NOT EXISTS ix_mileage_review_candidates_day
            ON mileage_review_candidates (report_day, normalized_plate)
            """
        )


def _iter_coordinate_totals(
    database_url: str,
    start_day: date,
    end_day: date,
) -> Iterator[tuple[date, str, float]]:
    start_at = datetime.combine(start_day, time.min, tzinfo=TZ)
    finish_at = datetime.combine(end_day + timedelta(days=1), time.min, tzinfo=TZ)
    query = """
        SELECT
            (event_time AT TIME ZONE 'Europe/Istanbul')::date AS local_day,
            normalized_plate,
            COALESCE(NULLIF(plate, ''), normalized_plate) AS display_plate,
            event_time AT TIME ZONE 'Europe/Istanbul' AS local_time,
            latitude,
            longitude,
            speed_kmh,
            distance_km,
            COALESCE(address, '')
        FROM gps_points
        WHERE event_time >= %s
          AND event_time < %s
        ORDER BY local_day, normalized_plate, event_time
    """

    def prepared_total(points: list[Point]) -> float:
        track = core.sanitize_position_outliers(points)
        return float(sum(mileage.coordinate_segment_distances(track)))

    with psycopg.connect(database_url) as connection:
        with connection.cursor(name="mileage_review_coordinate_points") as cursor:
            cursor.itersize = 20_000
            cursor.execute(query, (start_at, finish_at))
            key: tuple[date, str] | None = None
            points: list[Point] = []
            for row in cursor:
                (
                    day_value,
                    normalized,
                    display_plate,
                    local_time,
                    latitude,
                    longitude,
                    speed,
                    distance,
                    address,
                ) = row
                row_key = (day_value, normalize_plate(normalized))
                if key is not None and row_key != key:
                    yield key[0], key[1], prepared_total(points)
                    points = []
                key = row_key
                points.append(
                    Point(
                        plate=str(display_plate or normalized),
                        time=local_time,
                        lat=float(latitude),
                        lon=float(longitude),
                        speed=float(speed) if speed is not None else None,
                        source_distance=float(distance) if distance is not None else None,
                        address=str(address or ""),
                    )
                )
            if key is not None:
                yield key[0], key[1], prepared_total(points)


def calculate_mileage_review_candidates(
    database_url: str,
    start_day: date,
    end_day: date,
) -> dict[tuple[date, str], MileageReviewCandidate]:
    authoritative = mileage.load_authoritative_daily_distances(
        database_url,
        start_day,
        end_day,
    )
    coordinate = {
        (report_day, normalized_plate): coordinate_km
        for report_day, normalized_plate, coordinate_km
        in _iter_coordinate_totals(database_url, start_day, end_day)
    }

    result: dict[tuple[date, str], MileageReviewCandidate] = {}
    for key, authoritative_km in authoritative.items():
        coordinate_km = float(coordinate.get(key, 0.0))
        if not mileage_requires_review(authoritative_km, coordinate_km):
            continue
        gap_km = float(authoritative_km - coordinate_km)
        gap_percent = (
            gap_km / coordinate_km * 100.0
            if coordinate_km > 0
            else None
        )
        reason = (
            REVIEW_REASON_GAP
            if coordinate_km > 0
            else REVIEW_REASON_NO_COORDINATES
        )
        result[key] = MileageReviewCandidate(
            report_day=key[0],
            normalized_plate=key[1],
            authoritative_km=float(authoritative_km),
            coordinate_km=coordinate_km,
            gap_km=gap_km,
            gap_percent=gap_percent,
            reason=reason,
        )
    return result


def refresh_mileage_review_candidates(
    database_url: str,
    start_day: date,
    end_day: date,
) -> dict[tuple[date, str], MileageReviewCandidate]:
    candidates = calculate_mileage_review_candidates(
        database_url,
        start_day,
        end_day,
    )
    with psycopg.connect(database_url) as connection:
        ensure_review_schema(connection)
        with connection.cursor() as cursor:
            cursor.execute(
                """
                DELETE FROM mileage_review_candidates
                WHERE report_day BETWEEN %s AND %s
                """,
                (start_day, end_day),
            )
            for item in candidates.values():
                cursor.execute(
                    """
                    INSERT INTO mileage_review_candidates(
                        report_day,
                        normalized_plate,
                        authoritative_km,
                        coordinate_km,
                        gap_km,
                        gap_percent,
                        reason,
                        calculated_at
                    ) VALUES (%s,%s,%s,%s,%s,%s,%s,now())
                    """,
                    (
                        item.report_day,
                        item.normalized_plate,
                        item.authoritative_km,
                        item.coordinate_km,
                        item.gap_km,
                        item.gap_percent,
                        item.reason,
                    ),
                )
        connection.commit()
    return candidates


def load_mileage_review_candidates(
    database_url: str,
    start_day: date,
    end_day: date,
) -> dict[tuple[date, str], MileageReviewCandidate]:
    with psycopg.connect(database_url) as connection:
        ensure_review_schema(connection)
        connection.commit()
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT
                    report_day,
                    normalized_plate,
                    authoritative_km,
                    coordinate_km,
                    gap_km,
                    gap_percent,
                    reason
                FROM mileage_review_candidates
                WHERE report_day BETWEEN %s AND %s
                ORDER BY report_day, normalized_plate
                """,
                (start_day, end_day),
            )
            return {
                (report_day, normalized_plate): MileageReviewCandidate(
                    report_day=report_day,
                    normalized_plate=normalized_plate,
                    authoritative_km=float(authoritative_km),
                    coordinate_km=float(coordinate_km),
                    gap_km=float(gap_km),
                    gap_percent=(float(gap_percent) if gap_percent is not None else None),
                    reason=str(reason or ""),
                )
                for (
                    report_day,
                    normalized_plate,
                    authoritative_km,
                    coordinate_km,
                    gap_km,
                    gap_percent,
                    reason,
                ) in cursor.fetchall()
            }


def _as_day(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if value is None:
        return None
    text = str(value).strip()
    for parser in (
        lambda: date.fromisoformat(text),
        lambda: datetime.strptime(text, "%d.%m.%Y").date(),
    ):
        try:
            return parser()
        except ValueError:
            continue
    return None


def _copy_header_style(source, target) -> None:
    target._style = copy(source._style)
    target.font = copy(source.font)
    target.fill = copy(source.fill)
    target.border = copy(source.border)
    target.alignment = copy(source.alignment)
    target.protection = copy(source.protection)
    target.number_format = source.number_format


def annotate_mileage_review_workbook(
    output_path: Path,
    candidates: Mapping[tuple[date, str], MileageReviewCandidate],
) -> dict[str, int]:
    workbook = load_workbook(output_path)
    try:
        if "Сводный отчет" not in workbook.sheetnames:
            raise RuntimeError("В книге отсутствует лист «Сводный отчет»")
        sheet = workbook["Сводный отчет"]
        headers = {
            str(cell.value or "").strip(): cell.column
            for cell in sheet[1]
        }
        day_column = headers.get("Дата")
        plate_column = next(
            (
                headers[name]
                for name in ("Госномер / Plaka", "Госномер", "Номерной знак")
                if name in headers
            ),
            None,
        )
        company_column = next(
            (
                headers[name]
                for name in ("Компания или фирма", "Компания", "Фирма")
                if name in headers
            ),
            None,
        )
        driver_column = next(
            (
                headers[name]
                for name in (
                    "Имя водителя",
                    "ПОЛЬЗОВАТЕЛЬ / KULLANICI",
                    "Пользователь",
                    "Водитель",
                )
                if name in headers
            ),
            None,
        )
        if day_column is None or plate_column is None:
            raise RuntimeError("В сводном отчёте не найдены столбцы даты или госномера")

        review_column = headers.get(REVIEW_HEADER) or (sheet.max_column + 1)
        header = sheet.cell(1, review_column, REVIEW_HEADER)
        template = sheet.cell(1, max(1, review_column - 1))
        _copy_header_style(template, header)
        header.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        sheet.column_dimensions[get_column_letter(review_column)].width = 18

        row_details: dict[tuple[date, str], tuple[str, str, str]] = {}
        flagged_rows = 0
        review_fill = PatternFill("solid", fgColor="FDE68A")
        review_font = Font(bold=True, color="9C0006")
        for row_index in range(2, sheet.max_row + 1):
            report_day = _as_day(sheet.cell(row_index, day_column).value)
            normalized = normalize_plate(sheet.cell(row_index, plate_column).value)
            if report_day is None or not normalized:
                continue
            key = (report_day, normalized)
            display_plate = str(sheet.cell(row_index, plate_column).value or normalized)
            company = (
                str(sheet.cell(row_index, company_column).value or "")
                if company_column is not None
                else ""
            )
            driver = (
                str(sheet.cell(row_index, driver_column).value or "")
                if driver_column is not None
                else ""
            )
            row_details[key] = (display_plate, company, driver)
            cell = sheet.cell(row_index, review_column)
            if key in candidates:
                cell.value = REVIEW_VALUE
                cell.fill = review_fill
                cell.font = review_font
                flagged_rows += 1
            else:
                cell.value = None
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        if REVIEW_SHEET in workbook.sheetnames:
            del workbook[REVIEW_SHEET]
        review = workbook.create_sheet(REVIEW_SHEET, 1)
        review_headers = [
            "Дата",
            "Госномер",
            "Компания",
            "Водитель",
            "Пробег VehicleDistanceReport, км",
            "Пробег по координатам, км",
            "Разница, км",
            "Расхождение, %",
            "Статус",
            "Причина",
        ]
        review.append(review_headers)
        for item in sorted(
            candidates.values(),
            key=lambda value: (value.report_day, value.normalized_plate),
        ):
            display_plate, company, driver = row_details.get(
                (item.report_day, item.normalized_plate),
                (item.normalized_plate, "", ""),
            )
            review.append([
                item.report_day,
                display_plate,
                company,
                driver,
                round(item.authoritative_km, 3),
                round(item.coordinate_km, 3),
                round(item.gap_km, 3),
                (round(item.gap_percent, 1) if item.gap_percent is not None else None),
                REVIEW_VALUE,
                item.reason,
            ])

        header_fill = PatternFill("solid", fgColor="1F4E78")
        for cell in review[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        review.freeze_panes = "A2"
        review.auto_filter.ref = review.dimensions
        review.sheet_view.showGridLines = False
        review.row_dimensions[1].height = 48
        widths = (12, 18, 26, 36, 28, 24, 16, 18, 16, 48)
        for index, width in enumerate(widths, 1):
            review.column_dimensions[get_column_letter(index)].width = width
        for row in review.iter_rows(min_row=2):
            row[0].number_format = "dd.mm.yyyy"
            for index in (4, 5, 6):
                row[index].number_format = "0.000"
            row[7].number_format = "0.0"
            row[8].fill = review_fill
            row[8].font = review_font
            for cell in row:
                cell.alignment = Alignment(vertical="center", wrap_text=True)

        sheet.auto_filter.ref = sheet.dimensions
        save_report_workbook(workbook, output_path)
        return {
            "candidates": len(candidates),
            "flagged_rows": flagged_rows,
            "sheets": len(workbook.sheetnames),
        }
    finally:
        workbook.close()


def annotate_mileage_review(
    output_path: Path,
    database_url: str,
    start_day: date,
    end_day: date,
    *,
    refresh: bool,
) -> dict[str, int]:
    candidates = (
        refresh_mileage_review_candidates(database_url, start_day, end_day)
        if refresh
        else load_mileage_review_candidates(database_url, start_day, end_day)
    )
    return annotate_mileage_review_workbook(output_path, candidates)


def apply_mileage_review_ui(implementation: Any) -> None:
    """Highlight rows marked for mileage review in the browser preview."""
    style_marker = ".mileage-review-row td"
    if style_marker not in implementation.HTML and "</style>" in implementation.HTML:
        implementation.HTML = implementation.HTML.replace(
            "</style>",
            """
    .mileage-review-row td { background:#fff7d6; }
    .mileage-review-row:hover td { background:#ffefad; }
    .mileage-review-cell { color:#9c0006; font-weight:700; white-space:nowrap; }
  </style>""",
            1,
        )

    map_anchor = "  const mapIndex = tableColumns.indexOf('Карта');"
    if "mileageReviewIndex" not in implementation.HTML and map_anchor in implementation.HTML:
        implementation.HTML = implementation.HTML.replace(
            map_anchor,
            map_anchor + "\n  const mileageReviewIndex = tableColumns.indexOf('Проверка пробега');",
            1,
        )

    class_anchor = "    const className = plate && plate !== previousPlate ? ' class=\"plate-start\"' : '';"
    if class_anchor in implementation.HTML:
        implementation.HTML = implementation.HTML.replace(
            class_anchor,
            """    const rowClasses = [];
    if (plate && plate !== previousPlate) rowClasses.push('plate-start');
    if (mileageReviewIndex >= 0 && String(row[mileageReviewIndex] ?? '') === 'Проверить') rowClasses.push('mileage-review-row');
    const className = rowClasses.length ? ` class="${rowClasses.join(' ')}"` : '';""",
            1,
        )

    cell_anchor = "      return `<td>${esc(value)}</td>`;"
    if "mileage-review-cell" not in implementation.HTML and cell_anchor in implementation.HTML:
        implementation.HTML = implementation.HTML.replace(
            cell_anchor,
            """      const reviewClass = index === mileageReviewIndex && String(value ?? '') === 'Проверить'
        ? ' class="mileage-review-cell"'
        : '';
      return `<td${reviewClass}>${esc(value)}</td>`;""",
            1,
        )


__all__ = [
    "MileageReviewCandidate",
    "REVIEW_HEADER",
    "REVIEW_SHEET",
    "REVIEW_VALUE",
    "annotate_mileage_review",
    "annotate_mileage_review_workbook",
    "apply_authoritative_mileage_policy",
    "apply_incremental_mileage_policy",
    "apply_mileage_review_ui",
    "authoritative_vehicle_day_distances",
    "calculate_mileage_review_candidates",
    "load_mileage_review_candidates",
    "mileage_requires_review",
    "refresh_mileage_review_candidates",
]
