#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""Полная проверка доступности методов Arvento Report Web Service.

Скрипт:
1. Получает актуальный список методов со страницы report.asmx.
2. Для каждого метода читает список параметров со страницы ?op=Method.
3. Подставляет учётные данные и безопасные тестовые значения.
4. Выполняет HTTP POST к методу.
5. Классифицирует ответ и сохраняет Excel, JSON и сырые ответы.

Пароль в отчёты и логи не записывается.
"""

from __future__ import annotations

import argparse
import getpass
import html
import json
import os
import re
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from dataclasses import asdict, dataclass
from datetime import datetime, timedelta
from html.parser import HTMLParser
from pathlib import Path
from typing import Iterable
from xml.etree import ElementTree

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BASE_URL = "https://ws.arvento.com/v1/report.asmx"
USER_AGENT = "ArventoApiFullCheck/1.0"

CORE_METHODS = {
    "BuildingListReportReturnObject",
    "GeneralReportReturnObject",
    "GeneralReport2ReturnObject",
    "GeneralReportWithDistanceReturnObject",
    "RegionAlarmReturnObject",
    "GetVehicleInfoReturnObject",
    "GetLicensePlateNodeMappingsReturnObject",
    "GetNodesReturnObject",
    "GetGroupsReturnObject",
    "GetVehicleStatusJSON",
    "GetVehicleStatusV4",
    "GetVehicleStatusWithCourseReturnObject",
    "VehicleDailyStatusReport",
    "ContactAlarmReturnObject",
    "IgnitionDurationReportReturnObject",
    "MotionAlarmReturnObject",
    "IdlingDurationReportReturnObject",
    "IdleVehiclesReportReturnObject",
    "GetDriverInfoReturnObject",
    "GetDriverNodeMappingsReturnObject",
    "GetDriverFromNode",
    "DriverInformationReturnObject",
    "GetVehicleAlarmStatusJson",
    "DriverBehaviorReport",
    "DriveSafeReport",
    "MaximumSpeedReport",
    "CanBusOBDGeneralReportReturnObject",
    "CanBusOBDFuelInfoReportReturnObject",
    "CanBusFuelInfoReturnObject",
    "CanBusOBDOdometerInfoReportReturnObject",
    "FuelConsumptionReportReturnObject",
    "CANBUSTachographLastStatus",
}

ACCESS_DENIED_MARKERS = (
    "access denied",
    "yetkiniz yok",
    "yetki yok",
    "unauthorized",
    "forbidden",
    "erişim reddedildi",
    "erisim reddedildi",
)

PARAM_ERROR_MARKERS = (
    "parameter",
    "parametre",
    "invalid",
    "geçersiz",
    "gecersiz",
    "required",
    "zorunlu",
    "format",
    "conversion failed",
    "input string",
)

EMPTY_MARKERS = (
    "<newdataset />",
    "<newdataset></newdataset>",
    "<documentelement />",
    "<documentelement></documentelement>",
    "[]",
    "null",
)


class OperationListParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.operations: list[str] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag.lower() != "a":
            return
        href = dict(attrs).get("href") or ""
        match = re.search(r"[?&]op=([^&#]+)", href, flags=re.I)
        if match:
            name = urllib.parse.unquote(match.group(1)).strip()
            if name and name not in self.operations:
                self.operations.append(name)


class ParameterParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.parameters: list[str] = []
        self.in_form = False

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        attrs_dict = {k.lower(): (v or "") for k, v in attrs}
        if tag.lower() == "form":
            self.in_form = True
            return
        if not self.in_form or tag.lower() not in {"input", "select", "textarea"}:
            return
        name = attrs_dict.get("name", "").strip()
        input_type = attrs_dict.get("type", "").lower()
        if not name or input_type in {"submit", "button", "reset"}:
            return
        if name.startswith("__"):
            return
        if name not in self.parameters:
            self.parameters.append(name)

    def handle_endtag(self, tag: str) -> None:
        if tag.lower() == "form":
            self.in_form = False


@dataclass
class CheckResult:
    method: str
    status: str
    http_status: int | None
    elapsed_seconds: float
    parameters: list[str]
    sent_parameters: list[str]
    response_length: int
    content_type: str
    summary: str
    error: str
    raw_file: str


def http_request(
    url: str,
    *,
    data: dict[str, str] | None = None,
    timeout: float = 60.0,
) -> tuple[int, dict[str, str], bytes]:
    encoded = None if data is None else urllib.parse.urlencode(data).encode("utf-8")
    request = urllib.request.Request(
        url,
        data=encoded,
        headers={
            "User-Agent": USER_AGENT,
            "Content-Type": "application/x-www-form-urlencoded",
            "Accept": "*/*",
        },
        method="GET" if data is None else "POST",
    )
    try:
        with urllib.request.urlopen(request, timeout=timeout) as response:
            return response.status, dict(response.headers.items()), response.read()
    except urllib.error.HTTPError as exc:
        return exc.code, dict(exc.headers.items()), exc.read()


def decode_body(body: bytes, content_type: str = "") -> str:
    charset_match = re.search(r"charset=([\w-]+)", content_type, flags=re.I)
    encodings = [charset_match.group(1)] if charset_match else []
    encodings.extend(["utf-8-sig", "utf-8", "windows-1254", "windows-1251", "latin-1"])
    for encoding in encodings:
        try:
            return body.decode(encoding)
        except (UnicodeDecodeError, LookupError):
            continue
    return body.decode("utf-8", errors="replace")


def discover_operations(timeout: float) -> list[str]:
    status, headers, body = http_request(BASE_URL, timeout=timeout)
    if status != 200:
        raise RuntimeError(f"Не удалось получить список методов: HTTP {status}")
    parser = OperationListParser()
    parser.feed(decode_body(body, headers.get("Content-Type", "")))
    if not parser.operations:
        raise RuntimeError("На странице сервиса не найден список методов")
    return sorted(parser.operations, key=str.casefold)


def discover_parameters(method: str, timeout: float) -> list[str]:
    url = f"{BASE_URL}?op={urllib.parse.quote(method)}"
    status, headers, body = http_request(url, timeout=timeout)
    if status != 200:
        return []
    parser = ParameterParser()
    parser.feed(decode_body(body, headers.get("Content-Type", "")))
    return parser.parameters


def normalize_name(name: str) -> str:
    return re.sub(r"[^a-z0-9]", "", name.lower())


def infer_value(
    parameter: str,
    *,
    username: str,
    pin1: str,
    pin2: str,
    start_date: str,
    end_date: str,
    node: str,
    plate: str,
    group: str,
    overrides: dict[str, str],
) -> str:
    if parameter in overrides:
        return str(overrides[parameter])

    normalized = normalize_name(parameter)
    override_by_normalized = {normalize_name(k): str(v) for k, v in overrides.items()}
    if normalized in override_by_normalized:
        return override_by_normalized[normalized]

    exact = {
        "username": username,
        "user": username,
        "kullaniciadi": username,
        "pin1": pin1,
        "password": pin1,
        "sifre": pin1,
        "pin2": pin2,
        "startdate": start_date,
        "begindate": start_date,
        "datefrom": start_date,
        "fromdate": start_date,
        "enddate": end_date,
        "finishdate": end_date,
        "dateto": end_date,
        "todate": end_date,
        "node": node,
        "nodeno": node,
        "device": node,
        "deviceno": node,
        "licenseplate": plate,
        "plate": plate,
        "plaka": plate,
        "group": group,
        "groupname": group,
        "locale": "tr-TR",
        "language": "tr-TR",
        "lang": "tr-TR",
        "latitude": "36.3173",
        "longitude": "33.8745",
        "lat": "36.3173",
        "lon": "33.8745",
        "radius": "100",
        "distance": "100",
        "workingdays": "1,2,3,4,5",
        "workingstarttime": "08:00",
        "workingendtime": "18:00",
    }
    if normalized in exact:
        return exact[normalized]

    if normalized.startswith("chk"):
        return "1"
    if "start" in normalized and "date" in normalized:
        return start_date
    if "end" in normalized and "date" in normalized:
        return end_date
    if "node" in normalized or "device" in normalized:
        return node
    if "plate" in normalized or "plaka" in normalized:
        return plate
    if "group" in normalized:
        return group
    if "date" in normalized:
        return start_date
    if "time" in normalized:
        return "00:00"
    if any(token in normalized for token in ("count", "limit", "page", "index")):
        return "1"
    if normalized.startswith("is") or normalized.startswith("use"):
        return "0"
    return ""


def unwrap_response(text: str) -> str:
    stripped = text.strip()
    if not stripped.startswith("<"):
        return stripped
    try:
        root = ElementTree.fromstring(stripped)
    except ElementTree.ParseError:
        return stripped
    if root.text and root.text.strip():
        return html.unescape(root.text.strip())
    return stripped


def classify_response(http_status: int, text: str) -> tuple[str, str]:
    cleaned = unwrap_response(text).strip()
    lowered = cleaned.lower()

    if http_status < 200 or http_status >= 300:
        return "HTTP_ERROR", cleaned[:500]
    if any(marker in lowered for marker in ACCESS_DENIED_MARKERS):
        return "ACCESS_DENIED", cleaned[:500]
    if any(marker in lowered for marker in PARAM_ERROR_MARKERS):
        return "PARAMETER_ERROR", cleaned[:500]
    if not cleaned or lowered in EMPTY_MARKERS:
        return "ALLOWED_EMPTY", "Пустой ответ"
    if "exception" in lowered or "server error" in lowered or "soap:fault" in lowered:
        return "SERVER_ERROR", cleaned[:500]
    return "ALLOWED", cleaned[:500]


def safe_filename(name: str) -> str:
    return re.sub(r"[^A-Za-z0-9_.-]+", "_", name)[:150]


def load_overrides(path: str | None) -> dict[str, str]:
    if not path:
        return {}
    data = json.loads(Path(path).read_text(encoding="utf-8"))
    if not isinstance(data, dict):
        raise ValueError("Файл overrides должен содержать JSON-объект")
    return {str(k): "" if v is None else str(v) for k, v in data.items()}


def style_sheet(sheet) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.showGridLines = False
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for column in range(1, sheet.max_column + 1):
        width = min(
            max(len(str(sheet.cell(row, column).value or "")) for row in range(1, min(sheet.max_row, 500) + 1)) + 2,
            70,
        )
        sheet.column_dimensions[get_column_letter(column)].width = max(width, 12)


def save_excel(path: Path, results: list[CheckResult], settings: dict[str, object]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Проверка API"
    sheet.append([
        "№", "Метод", "Статус", "HTTP", "Время, сек", "Параметры метода",
        "Переданные параметры", "Размер ответа", "Content-Type", "Краткий ответ",
        "Ошибка", "Файл ответа",
    ])
    for index, result in enumerate(results, start=1):
        sheet.append([
            index,
            result.method,
            result.status,
            result.http_status,
            result.elapsed_seconds,
            ", ".join(result.parameters),
            ", ".join(result.sent_parameters),
            result.response_length,
            result.content_type,
            result.summary,
            result.error,
            result.raw_file,
        ])
    for row in sheet.iter_rows(min_row=2):
        row[4].number_format = "0.000"
    style_sheet(sheet)

    summary = workbook.create_sheet("Сводка")
    summary.append(["Статус", "Количество"])
    counts: dict[str, int] = {}
    for result in results:
        counts[result.status] = counts.get(result.status, 0) + 1
    for status, count in sorted(counts.items()):
        summary.append([status, count])
    summary.append(["ВСЕГО", len(results)])
    style_sheet(summary)

    params = workbook.create_sheet("Параметры запуска")
    params.append(["Параметр", "Значение"])
    for key, value in settings.items():
        params.append([key, value])
    style_sheet(params)
    workbook.save(path)


def select_methods(operations: Iterable[str], scope: str, methods: list[str]) -> list[str]:
    available = list(operations)
    if methods:
        requested = {item.strip() for item in methods if item.strip()}
        return [name for name in available if name in requested]
    if scope == "core":
        return [name for name in available if name in CORE_METHODS]
    return available


def main() -> None:
    parser = argparse.ArgumentParser(description="Полная проверка методов Arvento API")
    parser.add_argument("--scope", choices=["all", "core"], default="all", help="all — все методы; core — основные")
    parser.add_argument("--method", action="append", default=[], help="Проверить только указанный метод; можно повторять")
    parser.add_argument("--username", default=os.getenv("ARVENTO_USERNAME", ""))
    parser.add_argument("--pin1", default=os.getenv("ARVENTO_PIN1", ""))
    parser.add_argument("--pin2", default=os.getenv("ARVENTO_PIN2", ""))
    parser.add_argument("--node", default=os.getenv("ARVENTO_NODE", ""), help="Номер устройства для методов, где он обязателен")
    parser.add_argument("--plate", default=os.getenv("ARVENTO_PLATE", ""), help="Госномер для методов, где он обязателен")
    parser.add_argument("--group", default=os.getenv("ARVENTO_GROUP", ""), help="Группа автомобилей")
    parser.add_argument("--start-date", help="Начало периода, например 2026-07-23 00:00:00")
    parser.add_argument("--end-date", help="Конец периода, например 2026-07-23 23:59:59")
    parser.add_argument("--delay", type=float, default=2.0, help="Пауза между запросами, секунд")
    parser.add_argument("--timeout", type=float, default=60.0, help="Тайм-аут одного запроса")
    parser.add_argument("--overrides", help="JSON-файл со значениями параметров")
    parser.add_argument("--output-dir", default="arvento_api_check", help="Каталог результатов")
    parser.add_argument("--list-only", action="store_true", help="Только показать методы и параметры, без POST-запросов")
    args = parser.parse_args()

    username = args.username.strip() or input("Arvento Username: ").strip()
    pin1 = args.pin1 or getpass.getpass("Arvento PIN1: ")
    pin2 = args.pin2 if args.pin2 != "" else getpass.getpass("Arvento PIN2 (Enter, если пусто): ")

    yesterday = datetime.now() - timedelta(days=1)
    start_date = args.start_date or yesterday.strftime("%Y-%m-%d 00:00:00")
    end_date = args.end_date or yesterday.strftime("%Y-%m-%d 23:59:59")
    overrides = load_overrides(args.overrides)

    output_dir = Path(args.output_dir).expanduser().resolve()
    raw_dir = output_dir / "raw"
    raw_dir.mkdir(parents=True, exist_ok=True)

    print("Получение списка методов Arvento...")
    operations = discover_operations(args.timeout)
    selected = select_methods(operations, args.scope, args.method)
    if not selected:
        raise SystemExit("Не найдено методов для проверки")

    print(f"Обнаружено методов: {len(operations)}")
    print(f"Будет проверено: {len(selected)}")

    results: list[CheckResult] = []
    for index, method in enumerate(selected, start=1):
        print(f"[{index}/{len(selected)}] {method}")
        parameters = discover_parameters(method, args.timeout)

        if args.list_only:
            results.append(CheckResult(
                method=method,
                status="LIST_ONLY",
                http_status=None,
                elapsed_seconds=0.0,
                parameters=parameters,
                sent_parameters=[],
                response_length=0,
                content_type="",
                summary="",
                error="",
                raw_file="",
            ))
            continue

        payload = {
            parameter: infer_value(
                parameter,
                username=username,
                pin1=pin1,
                pin2=pin2,
                start_date=start_date,
                end_date=end_date,
                node=args.node,
                plate=args.plate,
                group=args.group,
                overrides=overrides,
            )
            for parameter in parameters
        }

        # На некоторых страницах HTML-форма не распознаётся. Минимальный набор
        # учётных данных всё равно передаётся.
        if not parameters:
            payload = {"Username": username, "PIN1": pin1, "PIN2": pin2}

        started = time.perf_counter()
        try:
            status_code, headers, body = http_request(
                f"{BASE_URL}/{urllib.parse.quote(method)}",
                data=payload,
                timeout=args.timeout,
            )
            elapsed = time.perf_counter() - started
            content_type = headers.get("Content-Type", "")
            text = decode_body(body, content_type)
            status, summary = classify_response(status_code, text)
            raw_path = raw_dir / f"{index:03d}_{safe_filename(method)}.txt"
            raw_path.write_text(text, encoding="utf-8", errors="replace")
            result = CheckResult(
                method=method,
                status=status,
                http_status=status_code,
                elapsed_seconds=round(elapsed, 3),
                parameters=parameters,
                sent_parameters=[name for name in payload if normalize_name(name) not in {"pin1", "pin2", "password", "sifre"}],
                response_length=len(body),
                content_type=content_type,
                summary=summary,
                error="",
                raw_file=str(raw_path.relative_to(output_dir)),
            )
        except Exception as exc:  # noqa: BLE001
            elapsed = time.perf_counter() - started
            result = CheckResult(
                method=method,
                status="REQUEST_ERROR",
                http_status=None,
                elapsed_seconds=round(elapsed, 3),
                parameters=parameters,
                sent_parameters=[name for name in payload if normalize_name(name) not in {"pin1", "pin2", "password", "sifre"}],
                response_length=0,
                content_type="",
                summary="",
                error=f"{type(exc).__name__}: {exc}",
                raw_file="",
            )

        results.append(result)
        print(f"    {result.status} | HTTP {result.http_status} | {result.elapsed_seconds:.3f} сек")
        if args.delay > 0 and index < len(selected):
            time.sleep(args.delay)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    json_path = output_dir / f"arvento_api_check_{timestamp}.json"
    xlsx_path = output_dir / f"arvento_api_check_{timestamp}.xlsx"

    json_path.write_text(
        json.dumps([asdict(item) for item in results], ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    save_excel(
        xlsx_path,
        results,
        {
            "Дата проверки": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "URL": BASE_URL,
            "Username": username,
            "PIN1/PIN2": "не сохраняются",
            "Режим": args.scope,
            "Начало периода": start_date,
            "Конец периода": end_date,
            "Node": args.node,
            "Plate": args.plate,
            "Group": args.group,
            "Пауза между запросами, сек": args.delay,
            "Количество методов": len(results),
        },
    )

    counts: dict[str, int] = {}
    for result in results:
        counts[result.status] = counts.get(result.status, 0) + 1

    print("\nГотово")
    print(f"Excel: {xlsx_path}")
    print(f"JSON:  {json_path}")
    print(f"Сырые ответы: {raw_dir}")
    for status, count in sorted(counts.items()):
        print(f"{status}: {count}")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\nПроверка остановлена пользователем", file=sys.stderr)
        raise SystemExit(130)
