#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Offline checks for authoritative mileage and review annotations."""
from __future__ import annotations

import tempfile
from datetime import date, datetime, timedelta
from pathlib import Path
from types import SimpleNamespace

from openpyxl import Workbook, load_workbook

from arvento_io import Point
from mileage_review_policy import (
    MileageReviewCandidate,
    REVIEW_HEADER,
    REVIEW_SHEET,
    REVIEW_VALUE,
    annotate_mileage_review_workbook,
    apply_authoritative_mileage_policy,
    apply_mileage_review_ui,
    authoritative_vehicle_day_distances,
    mileage_requires_review,
)

REPORT_DAY = date(2026, 8, 4)
START = datetime(2026, 8, 4, 8, 0)


def point(seconds: int, lon: float) -> Point:
    return Point(
        plate="34PKY310",
        time=START + timedelta(seconds=seconds),
        lat=36.0,
        lon=lon,
        speed=30.0,
    )


def check_authoritative_total_is_never_replaced() -> None:
    points = [
        point(0, 33.0000),
        point(60, 33.0010),
        point(120, 33.0020),
    ]
    authoritative_km = 4300.2
    mode = authoritative_vehicle_day_distances(points, authoritative_km)
    assert mode == "authoritative_scaled_coordinates"
    actual = sum(item.source_distance or 0.0 for item in points)
    assert abs(actual - authoritative_km) < 1e-6, actual
    coordinate_reference = 0.2
    assert mileage_requires_review(authoritative_km, coordinate_reference)

    apply_authoritative_mileage_policy()


def check_workbook_annotation() -> None:
    candidate = MileageReviewCandidate(
        report_day=REPORT_DAY,
        normalized_plate="34PKY310",
        authoritative_km=4300.2,
        coordinate_km=104.2,
        gap_km=4196.0,
        gap_percent=4026.9,
        reason="Тестовое расхождение",
    )

    with tempfile.TemporaryDirectory(prefix="verify_mileage_review_") as temp_name:
        output_path = Path(temp_name) / "report.xlsx"
        workbook = Workbook()
        report = workbook.active
        report.title = "Сводный отчет"
        report.append([
            "Дата",
            "Компания или фирма",
            "Госномер / Plaka",
            "ПОЛЬЗОВАТЕЛЬ / KULLANICI",
            "Пробег общий, км",
        ])
        report.append([
            REPORT_DAY,
            "TSM ENERJI",
            "34 PKY 310",
            "Тестовый водитель",
            4300.2,
        ])
        workbook.create_sheet("Диагностика")
        workbook.create_sheet("Параметры")
        workbook.save(output_path)
        workbook.close()

        stats = annotate_mileage_review_workbook(
            output_path,
            {(REPORT_DAY, "34PKY310"): candidate},
        )
        assert stats["candidates"] == 1
        assert stats["flagged_rows"] == 1

        result = load_workbook(output_path, read_only=True, data_only=True)
        try:
            assert REVIEW_SHEET in result.sheetnames, result.sheetnames
            report = result["Сводный отчет"]
            headers = [str(cell.value or "") for cell in report[1]]
            review_column = headers.index(REVIEW_HEADER) + 1
            assert report.cell(2, review_column).value == REVIEW_VALUE

            review = result[REVIEW_SHEET]
            assert review.max_row == 2
            assert review.cell(2, 2).value == "34 PKY 310"
            assert review.cell(2, 5).value == 4300.2
            assert review.cell(2, 6).value == 104.2
            assert review.cell(2, 9).value == REVIEW_VALUE
        finally:
            result.close()


def check_portal_marker_patch() -> None:
    implementation = SimpleNamespace(
        HTML="""
<style></style>
<script>
  const mapIndex = tableColumns.indexOf('Карта');
    const className = plate && plate !== previousPlate ? ' class="plate-start"' : '';
      return `<td>${esc(value)}</td>`;
</script>
"""
    )
    apply_mileage_review_ui(implementation)
    assert "mileageReviewIndex" in implementation.HTML
    assert "mileage-review-row" in implementation.HTML
    assert "Проверить" in implementation.HTML


if __name__ == "__main__":
    check_authoritative_total_is_never_replaced()
    check_workbook_annotation()
    check_portal_marker_patch()
    print("OK: VDR mileage is preserved and review candidates are marked")
