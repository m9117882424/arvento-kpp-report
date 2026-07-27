#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""Google Maps links for violation coordinates in Excel and the web portal."""

from __future__ import annotations

from pathlib import Path
from typing import Any
from urllib.parse import quote

from openpyxl import load_workbook

GOOGLE_MAPS_SEARCH_URL = "https://www.google.com/maps/search/?api=1&query="


def google_maps_url(lat: float, lon: float) -> str:
    """Return a stable Google Maps search URL for a latitude/longitude point."""
    latitude = float(lat)
    longitude = float(lon)
    if not -90 <= latitude <= 90:
        raise ValueError(f"Некорректная широта: {latitude}")
    if not -180 <= longitude <= 180:
        raise ValueError(f"Некорректная долгота: {longitude}")
    query = f"{latitude:.7f},{longitude:.7f}"
    return GOOGLE_MAPS_SEARCH_URL + quote(query, safe=",")


def parse_coordinate_pair(value: Any) -> tuple[float, float] | None:
    """Parse the report coordinate format ``lat, lon``."""
    if value in (None, ""):
        return None
    text = str(value).strip().replace(";", ",")
    parts = [part.strip() for part in text.split(",") if part.strip()]
    if len(parts) < 2:
        return None
    try:
        lat = float(parts[0].replace(" ", ""))
        lon = float(parts[1].replace(" ", ""))
    except ValueError:
        return None
    if not (-90 <= lat <= 90 and -180 <= lon <= 180):
        return None
    return lat, lon


def _header_columns(sheet) -> dict[str, int]:
    return {
        str(cell.value or "").strip(): cell.column
        for cell in sheet[1]
        if str(cell.value or "").strip()
    }


def _set_address_link(sheet, row: int, address_header: str, coordinates_header: str) -> bool:
    headers = _header_columns(sheet)
    address_column = headers.get(address_header)
    coordinates_column = headers.get(coordinates_header)
    if address_column is None or coordinates_column is None:
        return False

    coordinates = parse_coordinate_pair(sheet.cell(row, coordinates_column).value)
    if coordinates is None:
        return False

    cell = sheet.cell(row, address_column)
    if not str(cell.value or "").strip():
        cell.value = "Открыть на карте"
    cell.hyperlink = google_maps_url(*coordinates)
    cell.style = "Hyperlink"
    return True


def add_violation_map_links(workbook_path: Path) -> int:
    """Make violation address cells clickable in the generated Excel workbook."""
    workbook = load_workbook(workbook_path)
    links = 0
    try:
        for sheet_name in ("Скорость на площадке", "Скорость вне площадки"):
            if sheet_name not in workbook.sheetnames:
                continue
            sheet = workbook[sheet_name]
            for row in range(2, sheet.max_row + 1):
                links += int(
                    _set_address_link(
                        sheet,
                        row,
                        "Адрес максимума",
                        "Координаты максимума",
                    )
                )

        if "Запрещённый поворот" in workbook.sheetnames:
            sheet = workbook["Запрещённый поворот"]
            for row in range(2, sheet.max_row + 1):
                links += int(
                    _set_address_link(
                        sheet,
                        row,
                        "Адрес начала",
                        "Координаты начала",
                    )
                )
                links += int(
                    _set_address_link(
                        sheet,
                        row,
                        "Адрес окончания",
                        "Координаты окончания",
                    )
                )

        if "Параметры" in workbook.sheetnames:
            workbook["Параметры"].append([
                "Ссылки на карту",
                f"Google Maps, добавлено ссылок: {links}",
            ])

        workbook.save(workbook_path)
    finally:
        workbook.close()
    return links
