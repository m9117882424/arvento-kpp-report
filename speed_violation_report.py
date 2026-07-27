#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""Validated speed-violation detection and Excel sheets.

The site/outside state uses the same KPP crossing logic as the passenger-vehicle
utilisation report. A speed event is accepted only when it contains a sustained,
physically plausible sequence of GPS speeds. Isolated spikes are rejected.
"""

from __future__ import annotations

import math
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Sequence

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from arvento_analysis import EVENT_COOLDOWN_SECONDS, crossing_fraction, near_gate, side
from arvento_io import Point
from geozone_registry import Registry

SITE_SPEED_LIMIT_KMH = 30.0
OUTSIDE_SPEED_LIMIT_KMH = 95.0
DEFAULT_SITE_SPEED_THRESHOLD_KMH = 33.0
DEFAULT_OUTSIDE_SPEED_THRESHOLD_KMH = 104.5
MIN_SITE_THRESHOLD_KMH = 5.0
MAX_SITE_THRESHOLD_KMH = 200.0
MIN_OUTSIDE_THRESHOLD_KMH = 20.0
MAX_OUTSIDE_THRESHOLD_KMH = 250.0
MAX_VALID_GPS_SPEED_KMH = 250.0

# Anti-spike validation. A valid event must be sustained and smooth.
MIN_SPEED_EVENT_POINTS = 3
MIN_SPEED_EVENT_DURATION_SECONDS = 10
MAX_SPEED_EVENT_GAP_SECONDS = 5 * 60
MAX_ACCELERATION_MPS2 = 5.0
MAX_LOCAL_SPEED_DEVIATION_KMH = 25.0

TURN_SHEET_NAME = "Запрещённый поворот"
SUMMARY_SHEET_NAME = "Сводка по госномерам"
SITE_SHEET_NAME = "Скорость на площадке"
OUTSIDE_SHEET_NAME = "Скорость вне площадки"


@dataclass(frozen=True)
class SpeedViolation:
    plate: str
    start_point: Any
    finish_point: Any
    max_point: Any
    point_count: int
    speed_limit_kmh: float
    threshold_kmh: float
    on_site: bool

    @property
    def start(self) -> datetime:
        return self.start_point.timestamp

    @property
    def finish(self) -> datetime:
        return self.finish_point.timestamp

    @property
    def duration_seconds(self) -> int:
        return max(0, int((self.finish - self.start).total_seconds()))

    @property
    def max_speed_kmh(self) -> float:
        return float(self.max_point.speed)

    @property
    def tolerance_percent(self) -> float:
        if self.speed_limit_kmh <= 0:
            return 0.0
        return max(0.0, self.threshold_kmh / self.speed_limit_kmh - 1.0)


def validate_speed_thresholds(
    site_threshold_kmh: float,
    outside_threshold_kmh: float,
) -> tuple[float, float]:
    """Validate user-configured thresholds and return normalized floats."""
    site = float(site_threshold_kmh)
    outside = float(outside_threshold_kmh)
    if not math.isfinite(site) or not math.isfinite(outside):
        raise ValueError("Пороги скорости должны быть конечными числами")
    if not MIN_SITE_THRESHOLD_KMH <= site <= MAX_SITE_THRESHOLD_KMH:
        raise ValueError(
            f"Порог на площадке должен быть от {MIN_SITE_THRESHOLD_KMH:g} "
            f"до {MAX_SITE_THRESHOLD_KMH:g} км/ч"
        )
    if not MIN_OUTSIDE_THRESHOLD_KMH <= outside <= MAX_OUTSIDE_THRESHOLD_KMH:
        raise ValueError(
            f"Порог вне площадки должен быть от {MIN_OUTSIDE_THRESHOLD_KMH:g} "
            f"до {MAX_OUTSIDE_THRESHOLD_KMH:g} км/ч"
        )
    if outside < site:
        raise ValueError("Порог вне площадки не может быть ниже порога на площадке")
    return site, outside


def _as_gate_point(point: Any) -> Point:
    return Point(
        plate=point.plate,
        time=point.timestamp,
        lat=float(point.lat),
        lon=float(point.lon),
        speed=float(point.speed) if point.speed is not None else None,
        address=point.address or "",
    )


def _crossing_events(track: Sequence[Any], registry: Registry) -> dict[int, str]:
    """Return KPP crossing kind keyed by the index of the point after crossing."""
    if len(track) < 2:
        return {}

    first = _as_gate_point(track[0])
    stable_side = {gate.name: side(first, gate) for gate in registry.gates}
    stable_time = {gate.name: first.time for gate in registry.gates}
    last_event_time: datetime | None = None
    events: dict[int, str] = {}

    for index in range(1, len(track)):
        p1 = _as_gate_point(track[index - 1])
        p2 = _as_gate_point(track[index])
        candidates: list[tuple[float, str, datetime]] = []

        for gate in registry.gates:
            side2 = side(p2, gate)
            previous = stable_side[gate.name]
            if side2 != 0 and previous != 0 and side2 != previous:
                elapsed = (p2.time - stable_time[gate.name]).total_seconds()
                if elapsed <= 15 * 60 and (near_gate(p1, gate) or near_gate(p2, gate)):
                    fraction = crossing_fraction(p1, p2, gate)
                    kind = "Въезд" if previous == 1 and side2 == -1 else "Выезд"
                    event_time = p1.time + (p2.time - p1.time) * fraction
                    candidates.append((fraction, kind, event_time))
            if side2 != 0:
                stable_side[gate.name] = side2
                stable_time[gate.name] = p2.time

        if not candidates:
            continue

        _, kind, event_time = min(candidates, key=lambda item: item[0])
        if last_event_time is not None:
            if (event_time - last_event_time).total_seconds() < EVENT_COOLDOWN_SECONDS:
                continue
        events[index] = kind
        last_event_time = event_time

    return events


def classify_site_state(track: Sequence[Any], registry: Registry) -> list[tuple[Any, bool]]:
    """Classify each point as on-site or outside using KPP entry/exit events."""
    if not track:
        return []

    events = _crossing_events(track, registry)
    first_event = events[min(events)] if events else None
    inside = first_event == "Выезд"
    classified: list[tuple[Any, bool]] = [(track[0], inside)]

    for index in range(1, len(track)):
        kind = events.get(index)
        if kind == "Въезд":
            inside = True
        elif kind == "Выезд":
            inside = False
        classified.append((track[index], inside))

    return classified


def _valid_speed(value: Any) -> float | None:
    if value is None:
        return None
    try:
        speed = float(value)
    except (TypeError, ValueError):
        return None
    if not math.isfinite(speed) or speed < 0 or speed > MAX_VALID_GPS_SPEED_KMH:
        return None
    return speed


def _event_is_smooth(points: Sequence[Any]) -> bool:
    """Reject isolated spikes and physically implausible speed sequences."""
    if len(points) < MIN_SPEED_EVENT_POINTS:
        return False
    duration = (points[-1].timestamp - points[0].timestamp).total_seconds()
    if duration < MIN_SPEED_EVENT_DURATION_SECONDS:
        return False

    speeds = [_valid_speed(point.speed) for point in points]
    if any(speed is None for speed in speeds):
        return False
    numeric = [float(speed) for speed in speeds if speed is not None]

    for index in range(1, len(points)):
        seconds = (points[index].timestamp - points[index - 1].timestamp).total_seconds()
        if seconds <= 0 or seconds > MAX_SPEED_EVENT_GAP_SECONDS:
            return False
        acceleration = abs(numeric[index] - numeric[index - 1]) / 3.6 / seconds
        if acceleration > MAX_ACCELERATION_MPS2:
            return False

    # A single point that sharply rises above or falls below both neighbours is
    # treated as a GPS outlier. Smooth monotonic acceleration/deceleration passes.
    for index in range(1, len(points) - 1):
        before = numeric[index - 1]
        current = numeric[index]
        after = numeric[index + 1]
        dt_before = (points[index].timestamp - points[index - 1].timestamp).total_seconds()
        dt_after = (points[index + 1].timestamp - points[index].timestamp).total_seconds()
        total = dt_before + dt_after
        expected = before + (after - before) * (dt_before / total)
        is_local_extreme = current > max(before, after) or current < min(before, after)
        if is_local_extreme and abs(current - expected) > MAX_LOCAL_SPEED_DEVIATION_KMH:
            return False

    return True


def detect_speed_violations(
    track: Sequence[Any],
    registry: Registry,
    site_threshold_kmh: float = DEFAULT_SITE_SPEED_THRESHOLD_KMH,
    outside_threshold_kmh: float = DEFAULT_OUTSIDE_SPEED_THRESHOLD_KMH,
) -> tuple[list[SpeedViolation], list[SpeedViolation]]:
    """Detect grouped, sustained and smooth speed violations."""
    site_threshold, outside_threshold = validate_speed_thresholds(
        site_threshold_kmh,
        outside_threshold_kmh,
    )
    site_violations: list[SpeedViolation] = []
    outside_violations: list[SpeedViolation] = []
    active: dict[str, Any] | None = None

    def close_active() -> None:
        nonlocal active
        if active is None:
            return
        points = active["points"]
        if _event_is_smooth(points):
            max_point = max(points, key=lambda point: float(point.speed))
            violation = SpeedViolation(
                plate=points[0].plate,
                start_point=points[0],
                finish_point=points[-1],
                max_point=max_point,
                point_count=len(points),
                speed_limit_kmh=active["limit"],
                threshold_kmh=active["threshold"],
                on_site=active["on_site"],
            )
            (site_violations if violation.on_site else outside_violations).append(violation)
        active = None

    for point, on_site in classify_site_state(track, registry):
        limit = SITE_SPEED_LIMIT_KMH if on_site else OUTSIDE_SPEED_LIMIT_KMH
        threshold = site_threshold if on_site else outside_threshold
        speed = _valid_speed(point.speed)
        exceeds = speed is not None and speed > threshold

        if active is not None:
            previous = active["points"][-1]
            gap = (point.timestamp - previous.timestamp).total_seconds()
            if (
                active["on_site"] != on_site
                or gap <= 0
                or gap > MAX_SPEED_EVENT_GAP_SECONDS
                or not exceeds
            ):
                close_active()

        if not exceeds:
            continue

        if active is None:
            active = {
                "points": [point],
                "limit": limit,
                "threshold": threshold,
                "on_site": on_site,
            }
        else:
            active["points"].append(point)

    close_active()
    return site_violations, outside_violations


def _style_sheet(sheet) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.showGridLines = False
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="C00000")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[1].height = 45
    for column in range(1, sheet.max_column + 1):
        width = min(
            max(
                len(str(sheet.cell(row, column).value or ""))
                for row in range(1, min(sheet.max_row, 300) + 1)
            ) + 2,
            34,
        )
        sheet.column_dimensions[get_column_letter(column)].width = max(width, 12)


def _group_rows_by_plate(sheet, plate_column: int = 2) -> None:
    """Create visible Excel outline groups for contiguous plate rows."""
    if sheet.max_row < 2:
        return
    sheet.sheet_properties.outlinePr.summaryBelow = False
    start = 2
    current = str(sheet.cell(start, plate_column).value or "")
    for row in range(3, sheet.max_row + 2):
        value = str(sheet.cell(row, plate_column).value or "") if row <= sheet.max_row else None
        if value == current:
            continue
        end = row - 1
        if current:
            sheet.cell(start, plate_column).font = Font(bold=True)
            sheet.cell(start, plate_column).fill = PatternFill("solid", fgColor="D9EAF7")
            if end > start:
                sheet.row_dimensions.group(start + 1, end, outline_level=1, hidden=False)
        start = row
        current = value or ""


def _write_speed_sheet(
    workbook,
    title: str,
    violations: Sequence[SpeedViolation],
    index: int,
) -> None:
    if title in workbook.sheetnames:
        del workbook[title]
    sheet = workbook.create_sheet(title, index)
    sheet.append([
        "№",
        "Госномер",
        "Дата",
        "Начало нарушения",
        "Окончание нарушения",
        "Продолжительность",
        "Базовое ограничение, км/ч",
        "Допуск к ограничению, %",
        "Порог фиксации, км/ч",
        "Максимальная скорость, км/ч",
        "Превышение порога, км/ч",
        "Точек нарушения",
        "Координаты максимума",
        "Адрес максимума",
    ])

    ordered = sorted(violations, key=lambda item: (item.plate, item.start))
    for number, item in enumerate(ordered, start=1):
        max_point = item.max_point
        sheet.append([
            number,
            item.plate,
            item.start.date(),
            item.start,
            item.finish,
            item.duration_seconds / 86400.0,
            item.speed_limit_kmh,
            item.tolerance_percent,
            item.threshold_kmh,
            item.max_speed_kmh,
            item.max_speed_kmh - item.threshold_kmh,
            item.point_count,
            f"{max_point.lat:.7f}, {max_point.lon:.7f}",
            max_point.address,
        ])

    for row in sheet.iter_rows(min_row=2):
        row[2].number_format = "dd.mm.yyyy"
        row[3].number_format = "dd.mm.yyyy hh:mm:ss"
        row[4].number_format = "dd.mm.yyyy hh:mm:ss"
        row[5].number_format = "[h]:mm:ss"
        row[6].number_format = "0.0"
        row[7].number_format = "0.0%"
        row[8].number_format = "0.0"
        row[9].number_format = "0.0"
        row[10].number_format = "0.0"
    _style_sheet(sheet)
    _group_rows_by_plate(sheet)


def _write_plate_summary(
    workbook,
    site_violations: Sequence[SpeedViolation],
    outside_violations: Sequence[SpeedViolation],
) -> None:
    if SUMMARY_SHEET_NAME in workbook.sheetnames:
        del workbook[SUMMARY_SHEET_NAME]

    turn_counts: dict[str, int] = defaultdict(int)
    if TURN_SHEET_NAME in workbook.sheetnames:
        turn_sheet = workbook[TURN_SHEET_NAME]
        for row in turn_sheet.iter_rows(min_row=2, values_only=True):
            plate = str(row[1] or "").strip() if len(row) > 1 else ""
            if plate:
                turn_counts[plate] += 1

    stats: dict[str, dict[str, Any]] = defaultdict(
        lambda: {
            "site_count": 0,
            "site_max": None,
            "outside_count": 0,
            "outside_max": None,
        }
    )
    for item in site_violations:
        entry = stats[item.plate]
        entry["site_count"] += 1
        entry["site_max"] = max(entry["site_max"] or 0.0, item.max_speed_kmh)
    for item in outside_violations:
        entry = stats[item.plate]
        entry["outside_count"] += 1
        entry["outside_max"] = max(entry["outside_max"] or 0.0, item.max_speed_kmh)

    plates = sorted(set(turn_counts) | set(stats))
    sheet = workbook.create_sheet(SUMMARY_SHEET_NAME, 0)
    sheet.append([
        "Госномер",
        "Запрещённых поворотов",
        "Нарушений скорости на площадке",
        "Макс. скорость на площадке, км/ч",
        "Нарушений скорости вне площадки",
        "Макс. скорость вне площадки, км/ч",
        "Всего нарушений",
    ])
    for plate in plates:
        entry = stats[plate]
        turn_count = turn_counts.get(plate, 0)
        total = turn_count + entry["site_count"] + entry["outside_count"]
        sheet.append([
            plate,
            turn_count,
            entry["site_count"],
            entry["site_max"],
            entry["outside_count"],
            entry["outside_max"],
            total,
        ])
    for row in sheet.iter_rows(min_row=2):
        row[3].number_format = "0.0"
        row[5].number_format = "0.0"
    _style_sheet(sheet)


def append_speed_sheets(
    workbook_path: Path,
    site_violations: Sequence[SpeedViolation],
    outside_violations: Sequence[SpeedViolation],
    site_threshold_kmh: float = DEFAULT_SITE_SPEED_THRESHOLD_KMH,
    outside_threshold_kmh: float = DEFAULT_OUTSIDE_SPEED_THRESHOLD_KMH,
) -> None:
    """Append speed sheets, grouped details and a per-plate summary."""
    site_threshold, outside_threshold = validate_speed_thresholds(
        site_threshold_kmh,
        outside_threshold_kmh,
    )
    workbook = load_workbook(workbook_path)

    # The report itself is called "Нарушения"; keep the specific turn type as a
    # dedicated sheet name.
    if "Нарушения" in workbook.sheetnames and TURN_SHEET_NAME not in workbook.sheetnames:
        workbook["Нарушения"].title = TURN_SHEET_NAME

    _write_speed_sheet(workbook, SITE_SHEET_NAME, site_violations, 1)
    _write_speed_sheet(workbook, OUTSIDE_SHEET_NAME, outside_violations, 2)

    if TURN_SHEET_NAME in workbook.sheetnames:
        _group_rows_by_plate(workbook[TURN_SHEET_NAME])

    _write_plate_summary(workbook, site_violations, outside_violations)

    if "Параметры" in workbook.sheetnames:
        settings = workbook["Параметры"]
        settings.append(["Базовое ограничение на площадке, км/ч", SITE_SPEED_LIMIT_KMH])
        settings.append(["Порог фиксации на площадке, км/ч", site_threshold])
        settings.append(["Базовое ограничение вне площадки, км/ч", OUTSIDE_SPEED_LIMIT_KMH])
        settings.append(["Порог фиксации вне площадки, км/ч", outside_threshold])
        settings.append(["Разделение площадка/вне площадки", "по въездам и выездам через КПП 4 и КПП 5"])
        settings.append(["Минимум GPS-точек для валидного нарушения", MIN_SPEED_EVENT_POINTS])
        settings.append(["Минимальная длительность нарушения, секунд", MIN_SPEED_EVENT_DURATION_SECONDS])
        settings.append(["Максимально допустимая GPS-скорость, км/ч", MAX_VALID_GPS_SPEED_KMH])
        settings.append(["Максимальное ускорение/замедление, м/с²", MAX_ACCELERATION_MPS2])
        settings.append(["Максимальное локальное отклонение скорости, км/ч", MAX_LOCAL_SPEED_DEVIATION_KMH])
        settings.append(["Максимальный разрыв одного нарушения", MAX_SPEED_EVENT_GAP_SECONDS / 86400.0])
        settings.cell(settings.max_row, 2).number_format = "[h]:mm:ss"
        settings.append(["Нарушений скорости на площадке", len(site_violations)])
        settings.append(["Нарушений скорости вне площадки", len(outside_violations)])

    workbook.save(workbook_path)
