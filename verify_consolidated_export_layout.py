#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Offline verification for the downloadable consolidated workbook layout."""
from __future__ import annotations

import tempfile
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from consolidated_export_layout import (
    EXPORT_HEADERS,
    REPORT_SHEET,
    RosterExportDetails,
    finalize_consolidated_workbook,
)


def main() -> int:
    with tempfile.TemporaryDirectory(prefix="verify_consolidated_export_") as temp_name:
        output_path = Path(temp_name) / "report.xlsx"
        workbook = Workbook()
        report = workbook.active
        report.title = REPORT_SHEET
        report.append([
            "Дата",
            "Госномер / Plaka",
            "ПОЛЬЗОВАТЕЛЬ / KULLANICI",
            "Грейд / SCALA",
            "Заправка, л",
        ])
        report.append([date(2026, 8, 4), "33 ABC 123", "Водитель из отчёта", "7a", 10.0])
        report.append([date(2026, 8, 4), "34 XYZ 987", "Резервный водитель", "8b", 5.0])
        for cell in report[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        workbook.create_sheet("Диагностика")
        workbook.create_sheet("Параметры")
        workbook.save(output_path)
        workbook.close()

        stats = finalize_consolidated_workbook(
            output_path,
            {
                (date(2026, 8, 4), "33ABC123"): RosterExportDetails(
                    driver_name="Иван Иванов",
                    grade="9a",
                    directorate="Дирекция механизации",
                    responsible="Пётр Петров",
                )
            },
        )

        result = load_workbook(output_path, read_only=True, data_only=True)
        try:
            assert result.sheetnames == [REPORT_SHEET], result.sheetnames
            sheet = result[REPORT_SHEET]
            headers = [str(cell.value or "") for cell in sheet[1]]
            assert tuple(headers[-4:]) == EXPORT_HEADERS, headers
            assert [sheet.cell(2, column).value for column in range(sheet.max_column - 3, sheet.max_column + 1)] == [
                "Иван Иванов",
                "9a",
                "Дирекция механизации",
                "Пётр Петров",
            ]
            # Missing central details must preserve the existing driver and grade
            # while leaving the new organizational fields empty.
            assert [sheet.cell(3, column).value for column in range(sheet.max_column - 3, sheet.max_column + 1)] == [
                "Резервный водитель",
                "8b",
                None,
                None,
            ]
            assert stats["sheets"] == 1
            assert stats["rows"] == 2
            assert stats["enriched_rows"] == 1
        finally:
            result.close()

    print("OK: consolidated download keeps one sheet and appends roster columns")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
