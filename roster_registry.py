from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook

ROSTER_ALIASES = {
    "plate": ("Гос рег знак", "PLAKA", "Госномер", "Номерной знак"),
    "grade": ("Грейд", "SCALA", "Grade"),
    "driver": ("Пользователь", "KULLANICI", "Водитель"),
    "position": ("Должность", "GÖREVİ", "GOREVI"),
    "directorate": ("Дирекция", "Directorate"),
}

DATE_PATTERNS = (
    re.compile(r"(?<!\d)(\d{2})[._-](\d{2})[._-](\d{4})(?!\d)"),
    re.compile(r"(?<!\d)(\d{4})[._-](\d{2})[._-](\d{2})(?!\d)"),
)


@dataclass(frozen=True, slots=True)
class VehicleInfo:
    plate: str
    driver: str = ""
    grade: str = ""
    position: str = ""
    directorate: str = ""


@dataclass(frozen=True, slots=True)
class Roster:
    day: date
    path: Path
    vehicles: dict[str, VehicleInfo]


def clean(value: Any) -> str:
    return " ".join(str(value or "").replace("\n", " ").split()).strip()


def normalized(value: Any) -> str:
    return clean(value).lower().replace("ё", "е").replace("ı", "i")


def normalize_plate(value: Any) -> str:
    return "".join(ch for ch in clean(value).upper() if ch.isalnum())


def parse_date_value(value: Any) -> Optional[date]:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = clean(value)
    if not text:
        return None
    for pattern_index, pattern in enumerate(DATE_PATTERNS):
        match = pattern.search(text)
        if not match:
            continue
        try:
            if pattern_index == 0:
                day, month, year = map(int, match.groups())
            else:
                year, month, day = map(int, match.groups())
            return date(year, month, day)
        except ValueError:
            continue
    for fmt in ("%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def date_from_filename(path: Path) -> Optional[date]:
    return parse_date_value(path.stem)


def find_column(headers: list[Any], aliases: tuple[str, ...]) -> Optional[int]:
    values = [normalized(value) for value in headers]
    for alias in aliases:
        needle = normalized(alias)
        for index, header in enumerate(values):
            if header == needle or needle in header:
                return index
    return None


def find_roster_date_in_workbook(workbook) -> Optional[date]:
    for sheet in workbook.worksheets[:3]:
        for row in sheet.iter_rows(min_row=1, max_row=5, values_only=True):
            for value in row:
                parsed = parse_date_value(value)
                if parsed:
                    return parsed
    return None


def load_roster_file(path: Path) -> Optional[Roster]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        roster_day = date_from_filename(path) or find_roster_date_in_workbook(workbook)
        if roster_day is None:
            return None

        result: dict[str, VehicleInfo] = {}
        found_plate_column = False
        for sheet in workbook.worksheets:
            rows = iter(sheet.iter_rows(values_only=True))
            columns: Optional[dict[str, Optional[int]]] = None
            for _ in range(50):
                try:
                    row = next(rows)
                except StopIteration:
                    break
                candidate = {
                    key: find_column(list(row), aliases)
                    for key, aliases in ROSTER_ALIASES.items()
                }
                if candidate["plate"] is not None:
                    columns = candidate
                    found_plate_column = True
                    break
            if columns is None:
                continue

            max_index = max(index for index in columns.values() if index is not None)
            for row in rows:
                if len(row) <= max_index:
                    continue
                plate_index = columns["plate"]
                plate = normalize_plate(row[plate_index] if plate_index is not None else None)
                if not plate:
                    continue

                def value(key: str) -> str:
                    index = columns.get(key)
                    return clean(row[index]) if index is not None and index < len(row) else ""

                info = VehicleInfo(
                    plate=plate,
                    driver=value("driver"),
                    grade=value("grade"),
                    position=value("position"),
                    directorate=value("directorate"),
                )
                old = result.get(plate)
                old_score = sum(bool(getattr(old, field)) for field in ("driver", "grade", "position", "directorate")) if old else -1
                new_score = sum(bool(getattr(info, field)) for field in ("driver", "grade", "position", "directorate"))
                if new_score > old_score:
                    result[plate] = info

        if not found_plate_column or not result:
            return None
        return Roster(roster_day, path, result)
    finally:
        workbook.close()


def discover_rosters(folder: Path, source: Optional[Path] = None) -> list[Roster]:
    rosters: list[Roster] = []
    for path in sorted(folder.glob("*.xls*")):
        if source and path.resolve() == source.resolve():
            continue
        lowered = path.name.lower()
        if any(token in lowered for token in ("_по_дням", "_итоговая_сводка", "первый въезд")):
            continue
        try:
            roster = load_roster_file(path)
        except Exception as exc:
            print(f"  пропуск файла разнарядки {path.name}: {exc}")
            continue
        if roster:
            print(f"  разнарядка {roster.day.strftime('%d.%m.%Y')}: {path.name}; машин {len(roster.vehicles)}")
            rosters.append(roster)
    return sorted(rosters, key=lambda item: (item.day, item.path.name))


def select_roster(rosters: list[Roster], report_day: date) -> Optional[Roster]:
    exact = [item for item in rosters if item.day == report_day]
    if exact:
        return exact[-1]
    previous = [item for item in rosters if item.day <= report_day]
    if previous:
        return previous[-1]
    return rosters[-1] if rosters else None


def enrich_daily_with_rosters(daily: dict[date, list[dict[str, Any]]], rosters: list[Roster]) -> None:
    for report_day, rows in daily.items():
        roster = select_roster(rosters, report_day)
        for item in rows:
            info = roster.vehicles.get(normalize_plate(item.get("plate"))) if roster else None
            item["driver"] = info.driver if info else ""
            item["grade"] = info.grade if info else ""
            item["position"] = info.position if info else ""
            item["directorate"] = info.directorate if info else ""
            item["roster_date"] = roster.day if roster else None
            item["roster_file"] = roster.path.name if roster else ""
