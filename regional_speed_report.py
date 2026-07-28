#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Three-zone speed violation classification and Excel output.

Priority:
1. Akkuyu site polygon;
2. Tasucu–Akkuyu route polygon, outside the site;
3. outside every enabled geozone and outside the route.

Points inside ``purpose=speed_exclusion`` zones are ignored completely for
speed maxima and violations, while remaining available to mileage reports.
"""
from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Sequence

from openpyxl import load_workbook

from consolidated_report import DEFAULT_ROUTE_KML, load_kml_polygon
from geozone_registry import Registry, find_site_boundary, point_in_polygon, point_in_zone
from speed_violation_report import (
    DEFAULT_OUTSIDE_SPEED_THRESHOLD_KMH,
    DEFAULT_SITE_SPEED_THRESHOLD_KMH,
    MAX_SPEED_EVENT_GAP_SECONDS,
    OUTSIDE_SPEED_LIMIT_KMH,
    SITE_SPEED_LIMIT_KMH,
    TURN_SHEET_NAME,
    _event_is_smooth,
    _group_rows_by_plate,
    _style_sheet,
    _valid_speed,
    _write_speed_sheet,
    validate_speed_thresholds,
)

SITE_SHEET_NAME = "Скорость на площадке"
ROUTE_SHEET_NAME = "Скорость Ташуджу - Аккую"
REGION_SHEET_NAME = "Скорость вне региона"
SUMMARY_SHEET_NAME = "Сводка по госномерам"


@dataclass(frozen=True)
class RegionalSpeedViolation:
    plate: str
    start_point: Any
    finish_point: Any
    max_point: Any
    point_count: int
    speed_limit_kmh: float
    threshold_kmh: float
    category: str

    @property
    def start(self) -> datetime:
        return self.start_point.timestamp

    @property
    def finish(self) -> datetime:
        return self.finish_point.timestamp

    @property
    def duration_seconds(self) -> int:
        return max(0, int((self.finish - self.start).total_seconds()))

    @property
    def max_speed_kmh(self) -> float:
        return float(self.max_point.speed)

    @property
    def tolerance_percent(self) -> float:
        if self.speed_limit_kmh <= 0:
            return 0.0
        return max(0.0, self.threshold_kmh / self.speed_limit_kmh - 1.0)


def classify_speed_category(
    point: Any,
    registry: Registry,
    route_polygon: list[tuple[float, float]],
) -> str | None:
    """Classify one GPS point using site → route → outside-region priority."""
    lat = float(point.lat)
    lon = float(point.lon)
    exclusions = [zone for zone in registry.zones if zone.purpose == "speed_exclusion"]
    if any(point_in_zone(lat, lon, zone) for zone in exclusions):
        return None

    boundary = find_site_boundary(registry)
    if point_in_zone(lat, lon, boundary):
        return "site"
    if point_in_polygon(lat, lon, route_polygon):
        return "route"

    other_zones = [
        zone
        for zone in registry.zones
        if zone.purpose not in {"site_boundary", "speed_exclusion"}
    ]
    if any(point_in_zone(lat, lon, zone) for zone in other_zones):
        return None
    return "region"


def classify_speed_categories(
    track: Sequence[Any],
    registry: Registry,
    route_polygon: list[tuple[float, float]],
) -> list[str | None]:
    """Classify a track and suppress one-point non-tunnel geofence jitter."""
    raw = [classify_speed_category(point, registry, route_polygon) for point in track]
    stable = list(raw)
    for index in range(1, len(raw) - 1):
        if raw[index] is None:
            # A tunnel/excluded point must always split a speed event.
            continue
        if raw[index - 1] == raw[index + 1] != raw[index]:
            stable[index] = raw[index - 1]
    return stable


def detect_regional_speed_violations(
    track: Sequence[Any],
    registry: Registry,
    site_threshold_kmh: float = DEFAULT_SITE_SPEED_THRESHOLD_KMH,
    outside_threshold_kmh: float = DEFAULT_OUTSIDE_SPEED_THRESHOLD_KMH,
    route_kml: Path = DEFAULT_ROUTE_KML,
) -> tuple[
    list[RegionalSpeedViolation],
    list[RegionalSpeedViolation],
    list[RegionalSpeedViolation],
]:
    """Detect sustained violations in site, route and outside-region categories."""
    site_threshold, outside_threshold = validate_speed_thresholds(
        site_threshold_kmh,
        outside_threshold_kmh,
    )
    route_polygon = load_kml_polygon(route_kml)
    categories = classify_speed_categories(track, registry, route_polygon)
    result: dict[str, list[RegionalSpeedViolation]] = {
        "site": [],
        "route": [],
        "region": [],
    }
    active: dict[str, Any] | None = None

    def close_active() -> None:
        nonlocal active
        if active is None:
            return
        points = active["points"]
        if _event_is_smooth(points):
            result[active["category"]].append(
                RegionalSpeedViolation(
                    plate=points[0].plate,
                    start_point=points[0],
                    finish_point=points[-1],
                    max_point=max(points, key=lambda item: float(item.speed)),
                    point_count=len(points),
                    speed_limit_kmh=active["limit"],
                    threshold_kmh=active["threshold"],
                    category=active["category"],
                )
            )
        active = None

    for point, category in zip(track, categories):
        speed = _valid_speed(point.speed)
        if category == "site":
            limit, threshold = SITE_SPEED_LIMIT_KMH, site_threshold
        elif category in {"route", "region"}:
            limit, threshold = OUTSIDE_SPEED_LIMIT_KMH, outside_threshold
        else:
            limit = threshold = None
        exceeds = speed is not None and threshold is not None and speed > threshold

        if active is not None:
            previous = active["points"][-1]
            gap = (point.timestamp - previous.timestamp).total_seconds()
            if (
                active["category"] != category
                or gap <= 0
                or gap > MAX_SPEED_EVENT_GAP_SECONDS
                or not exceeds
            ):
                close_active()

        if not exceeds or category is None:
            continue
        if active is None:
            active = {
                "points": [point],
                "category": category,
                "limit": limit,
                "threshold": threshold,
            }
        else:
            active["points"].append(point)

    close_active()
    return result["site"], result["route"], result["region"]


def _round_speed_sheet(sheet) -> None:
    """Store speed values at one-decimal precision, not only as an Excel format."""
    headers = {
        str(cell.value or "").strip(): cell.column
        for cell in sheet[1]
        if str(cell.value or "").strip()
    }
    one_decimal_headers = (
        "Базовое ограничение, км/ч",
        "Порог фиксации, км/ч",
        "Максимальная скорость, км/ч",
        "Превышение порога, км/ч",
    )
    for header in one_decimal_headers:
        column = headers.get(header)
        if column is None:
            continue
        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row, column)
            if isinstance(cell.value, (int, float)):
                cell.value = round(float(cell.value), 1)
                cell.number_format = "0.0"
    tolerance_column = headers.get("Допуск к ограничению, %")
    if tolerance_column is not None:
        for row in range(2, sheet.max_row + 1):
            sheet.cell(row, tolerance_column).number_format = "0.0%"


def _write_summary(workbook, site, route, region) -> None:
    if SUMMARY_SHEET_NAME in workbook.sheetnames:
        del workbook[SUMMARY_SHEET_NAME]
    turn_counts: dict[str, int] = defaultdict(int)
    if TURN_SHEET_NAME in workbook.sheetnames:
        for row in workbook[TURN_SHEET_NAME].iter_rows(min_row=2, values_only=True):
            plate = str(row[1] or "").strip() if len(row) > 1 else ""
            if plate:
                turn_counts[plate] += 1

    stats = defaultdict(
        lambda: {
            "site_count": 0,
            "site_max": None,
            "route_count": 0,
            "route_max": None,
            "region_count": 0,
            "region_max": None,
        }
    )
    for key, items in (("site", site), ("route", route), ("region", region)):
        for item in items:
            entry = stats[item.plate]
            entry[f"{key}_count"] += 1
            current = entry[f"{key}_max"] or 0.0
            entry[f"{key}_max"] = round(max(current, item.max_speed_kmh), 1)

    sheet = workbook.create_sheet(SUMMARY_SHEET_NAME, 0)
    sheet.append(
        [
            "Госномер",
            "Запрещённых поворотов",
            "Нарушений скорости на площадке",
            "Макс. скорость на площадке, км/ч",
            "Нарушений скорости Ташуджу - Аккую",
            "Макс. скорость Ташуджу - Аккую, км/ч",
            "Нарушений скорости вне региона",
            "Макс. скорость вне региона, км/ч",
            "Всего нарушений",
        ]
    )
    for plate in sorted(set(turn_counts) | set(stats)):
        entry = stats[plate]
        total = turn_counts.get(plate, 0) + sum(
            entry[f"{key}_count"] for key in ("site", "route", "region")
        )
        sheet.append(
            [
                plate,
                turn_counts.get(plate, 0),
                entry["site_count"],
                entry["site_max"],
                entry["route_count"],
                entry["route_max"],
                entry["region_count"],
                entry["region_max"],
                total,
            ]
        )
    for row in sheet.iter_rows(min_row=2):
        for index in (3, 5, 7):
            row[index].number_format = "0.0"
    _style_sheet(sheet)


def append_regional_speed_sheets(
    workbook_path: Path,
    site_violations: Sequence[RegionalSpeedViolation],
    route_violations: Sequence[RegionalSpeedViolation],
    region_violations: Sequence[RegionalSpeedViolation],
    site_threshold_kmh: float = DEFAULT_SITE_SPEED_THRESHOLD_KMH,
    outside_threshold_kmh: float = DEFAULT_OUTSIDE_SPEED_THRESHOLD_KMH,
) -> None:
    site_threshold, outside_threshold = validate_speed_thresholds(
        site_threshold_kmh,
        outside_threshold_kmh,
    )
    workbook = load_workbook(workbook_path)
    try:
        if "Нарушения" in workbook.sheetnames and TURN_SHEET_NAME not in workbook.sheetnames:
            workbook["Нарушения"].title = TURN_SHEET_NAME
        _write_speed_sheet(workbook, SITE_SHEET_NAME, site_violations, 1)
        _write_speed_sheet(workbook, ROUTE_SHEET_NAME, route_violations, 2)
        _write_speed_sheet(workbook, REGION_SHEET_NAME, region_violations, 3)
        for sheet_name in (SITE_SHEET_NAME, ROUTE_SHEET_NAME, REGION_SHEET_NAME):
            _round_speed_sheet(workbook[sheet_name])
        if TURN_SHEET_NAME in workbook.sheetnames:
            _group_rows_by_plate(workbook[TURN_SHEET_NAME])
        _write_summary(workbook, site_violations, route_violations, region_violations)
        if "Параметры" in workbook.sheetnames:
            sheet = workbook["Параметры"]
            sheet.append(["Порог фиксации на площадке, км/ч", round(site_threshold, 1)])
            sheet.append(
                [
                    "Порог фиксации Ташуджу - Аккую и вне региона, км/ч",
                    round(outside_threshold, 1),
                ]
            )
            sheet.append(["Тоннельные геозоны", "скорость полностью исключена"])
            sheet.append(["Нарушений скорости на площадке", len(site_violations)])
            sheet.append(["Нарушений скорости Ташуджу - Аккую", len(route_violations)])
            sheet.append(["Нарушений скорости вне региона", len(region_violations)])
        workbook.save(workbook_path)
    finally:
        workbook.close()
