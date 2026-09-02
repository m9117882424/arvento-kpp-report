#!/usr/bin/env python3
"""Runtime smoke tests executed during the server image build.

The checks cover report logic, deployed ASGI entrypoint, three speed regions,
tunnel exclusions, browser preview and one-decimal Excel/web output.
"""
from __future__ import annotations

import math
import os
import tempfile
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook
from excel_formatting import save_report_workbook

import consolidated_report
from arvento_first_entry_report import ReportFilters
from arvento_first_entry_report_fixed import build_report_titles, extract_date_from_text
from arvento_io import Point
from arvento_reports import round_metric, round_ratio
from consolidated_multi_report import DatedRoster, has_night_site_mileage, select_roster
from consolidated_report import (
    HEADERS as CONSOLIDATED_HEADERS,
    allowed_entry_exit_time,
    load_kml_polygon,
    validated_speed_indices,
)
from geozone_registry import (
    Geozone,
    find_site_boundary,
    load_registry,
    point_in_polygon,
    point_in_zone,
    suppress_speed_in_exclusions,
)
from map_links import google_maps_url, parse_coordinate_pair
from regional_speed_report import (
    REGION_SHEET_NAME,
    ROUTE_SHEET_NAME,
    SITE_SHEET_NAME,
    classify_speed_category,
    detect_regional_speed_violations,
)
from site_boundary_speed import classify_site_state_by_polygon
from speed_violation_report import TURN_SHEET_NAME, _event_is_smooth, validate_speed_thresholds

STRICT_ENV = "ARVENTO_RUNTIME_CHECK_STRICT"
OPTIONAL_PORTAL_DEPENDENCIES = {
    "fastapi",
    "openpyxl",
    "psycopg",
    "pydantic",
    "starlette",
    "uvicorn",
}


@dataclass
class FakePoint:
    timestamp: datetime
    speed: float | None
    lat: float = 36.145
    lon: float = 33.55
    plate: str = "TEST"
    address: str = ""


def points(*speeds: float, seconds: int = 10) -> list[FakePoint]:
    start = datetime(2026, 1, 1, 8, 0, 0)
    return [
        FakePoint(start + timedelta(seconds=index * seconds), speed)
        for index, speed in enumerate(speeds)
    ]


def track_at(lat: float, lon: float, *speeds: float) -> list[FakePoint]:
    start = datetime(2026, 1, 1, 8, 0, 0)
    return [
        FakePoint(
            timestamp=start + timedelta(seconds=index * 10),
            speed=speed,
            lat=lat,
            lon=lon,
        )
        for index, speed in enumerate(speeds)
    ]


def consolidated_points(*speeds: float, seconds: int = 10) -> list[Point]:
    start = datetime(2026, 1, 1, 8, 0, 0)
    return [
        Point(
            plate="TEST",
            time=start + timedelta(seconds=index * seconds),
            lat=36.145 + index * 0.00001,
            lon=33.55 + index * 0.00001,
            speed=speed,
        )
        for index, speed in enumerate(speeds)
    ]


def movement_track(start: datetime, seconds: int = 60) -> list[Point]:
    return [
        Point(
            plate="TEST",
            time=start,
            lat=36.145000,
            lon=33.550000,
            speed=15,
        ),
        Point(
            plate="TEST",
            time=start + timedelta(seconds=seconds),
            lat=36.145500,
            lon=33.550500,
            speed=18,
        ),
    ]


def strict_mode() -> bool:
    return os.environ.get(STRICT_ENV, "").strip().lower() in {"1", "true", "yes", "on"}


def polygon_area(points_value: list[tuple[float, float]]) -> float:
    return abs(
        sum(
            points_value[index][1] * points_value[(index + 1) % len(points_value)][0]
            - points_value[(index + 1) % len(points_value)][1] * points_value[index][0]
            for index in range(len(points_value))
        )
    ) / 2.0


def find_polygon_interior(polygon: list[tuple[float, float]]) -> tuple[float, float]:
    min_lat = min(lat for lat, _ in polygon)
    max_lat = max(lat for lat, _ in polygon)
    min_lon = min(lon for _, lon in polygon)
    max_lon = max(lon for _, lon in polygon)
    for lat_step in range(1, 40):
        lat = min_lat + (max_lat - min_lat) * lat_step / 40.0
        for lon_step in range(1, 40):
            lon = min_lon + (max_lon - min_lon) * lon_step / 40.0
            if point_in_polygon(lat, lon, polygon):
                return lat, lon
    raise AssertionError("Не найдена внутренняя точка полигона")


def find_zone_interior(zone: Geozone) -> tuple[float, float]:
    if zone.zone_type == "circle" and zone.center:
        return zone.center
    polygon = list(zone.points or [])
    return find_polygon_interior(polygon)


def find_category_point(category: str, registry, route_polygon) -> tuple[float, float]:
    if category == "site":
        return find_zone_interior(find_site_boundary(registry))
    if category == "region":
        candidate = FakePoint(datetime(2026, 1, 1, 8), 0, 36.5, 34.5)
        assert classify_speed_category(candidate, registry, route_polygon) == "region"
        return candidate.lat, candidate.lon

    min_lat = min(lat for lat, _ in route_polygon)
    max_lat = max(lat for lat, _ in route_polygon)
    min_lon = min(lon for _, lon in route_polygon)
    max_lon = max(lon for _, lon in route_polygon)
    for lat_step in range(1, 60):
        lat = min_lat + (max_lat - min_lat) * lat_step / 60.0
        for lon_step in range(1, 60):
            lon = min_lon + (max_lon - min_lon) * lon_step / 60.0
            candidate = FakePoint(datetime(2026, 1, 1, 8), 0, lat, lon)
            if classify_speed_category(candidate, registry, route_polygon) == category:
                return lat, lon
    raise AssertionError(f"Не найдена тестовая точка категории {category}")


def check_regional_speed_runtime(registry, route_polygon) -> None:
    site_lat, site_lon = find_category_point("site", registry, route_polygon)
    route_lat, route_lon = find_category_point("route", registry, route_polygon)
    region_lat, region_lon = find_category_point("region", registry, route_polygon)

    assert classify_speed_category(
        FakePoint(datetime(2026, 1, 1, 8), 0, site_lat, site_lon),
        registry,
        route_polygon,
    ) == "site"
    assert classify_speed_category(
        FakePoint(datetime(2026, 1, 1, 8), 0, route_lat, route_lon),
        registry,
        route_polygon,
    ) == "route"
    assert classify_speed_category(
        FakePoint(datetime(2026, 1, 1, 8), 0, region_lat, region_lon),
        registry,
        route_polygon,
    ) == "region"

    site, route, region = detect_regional_speed_violations(
        track_at(site_lat, site_lon, 55, 57, 59),
        registry,
    )
    assert (len(site), len(route), len(region)) == (1, 0, 0)

    site, route, region = detect_regional_speed_violations(
        track_at(route_lat, route_lon, 110, 112, 114),
        registry,
    )
    assert (len(site), len(route), len(region)) == (0, 1, 0)

    site, route, region = detect_regional_speed_violations(
        track_at(region_lat, region_lon, 110, 112, 114),
        registry,
    )
    assert (len(site), len(route), len(region)) == (0, 0, 1)

    exclusions = [zone for zone in registry.zones if zone.purpose == "speed_exclusion"]
    assert {zone.name for zone in exclusions} == {
        "Тоннели около Ташуджу",
        "Тоннели около Аккую",
    }
    for zone in exclusions:
        lat, lon = find_zone_interior(zone)
        candidate = FakePoint(datetime(2026, 1, 1, 8), 200, lat, lon)
        assert classify_speed_category(candidate, registry, route_polygon) is None
        detected = detect_regional_speed_violations(
            track_at(lat, lon, 190, 195, 200),
            registry,
        )
        assert all(not items for items in detected), f"Скорость в «{zone.name}» попала в нарушения"

        consolidated_track = [
            Point(
                plate="TEST",
                time=datetime(2026, 1, 1, 8) + timedelta(seconds=index * 10),
                lat=lat,
                lon=lon,
                speed=speed,
            )
            for index, speed in enumerate((190, 195, 200))
        ]
        row = consolidated_report.analyze_track(
            date(2026, 1, 1),
            "TEST",
            suppress_speed_in_exclusions(consolidated_track, registry),
            {},
            list(find_site_boundary(registry).points or []),
            route_polygon,
        )
        assert row is not None
        assert row.max_speed is None and row.route_max_speed is None and row.site_max_speed is None


def check_excel_rounding() -> None:
    with tempfile.TemporaryDirectory(prefix="arvento_rounding_test_") as temp_name:
        path = Path(temp_name) / "rounding.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Показатель", "Широта", "Процент"])
        sheet.append([33.700001, 36.1234567, 0.32654])
        sheet["C2"].number_format = "0.000%"
        save_report_workbook(workbook, path)
        workbook.close()

        check = load_workbook(path, data_only=True)
        try:
            sheet = check.active
            assert sheet["A2"].value == 33.7
            assert sheet["A2"].number_format == "0.0"
            assert math.isclose(float(sheet["B2"].value), 36.1234567, abs_tol=1e-9)
            assert sheet["C2"].number_format == "0.0%"
        finally:
            check.close()


def add_speed_preview_sheet(workbook, title: str, speed: float) -> None:
    sheet = workbook.create_sheet(title)
    sheet.append(
        [
            "№",
            "Госномер",
            "Дата",
            "Начало нарушения",
            "Окончание нарушения",
            "Продолжительность",
            "Базовое ограничение, км/ч",
            "Допуск к ограничению, %",
            "Порог фиксации, км/ч",
            "Максимальная скорость, км/ч",
            "Превышение порога, км/ч",
            "Точек нарушения",
            "Координаты максимума",
            "Адрес максимума",
        ]
    )
    start = datetime(2026, 1, 1, 8, 0, 0)
    sheet.append(
        [
            1,
            "TEST",
            start.date(),
            start,
            start + timedelta(seconds=20),
            20 / 86400,
            95.0,
            0.1,
            104.5,
            speed,
            speed - 104.5,
            3,
            "36.2000000, 33.7000000",
            "Тест",
        ]
    )


def check_portal_runtime() -> bool:
    try:
        import portal_entrypoint
        import portal_runtime_patch
        import run_report_portal
    except ModuleNotFoundError as exc:
        missing = (exc.name or "").split(".", 1)[0]
        if strict_mode() or missing not in OPTIONAL_PORTAL_DEPENDENCIES:
            raise
        print(
            "ПРЕДУПРЕЖДЕНИЕ: проверка импорта веб-портала пропущена на хосте — "
            f"не установлен пакет {missing!r}. Полная проверка будет выполнена "
            "внутри Docker при сборке report-portal."
        )
        return False

    implementation = portal_entrypoint.portal.implementation
    html = implementation.HTML
    for token in (
        'value="consolidated">Сводный отчёт',
        'id="consolidatedRosters"',
        'name="rosters"',
        "multiple",
        "type === 'consolidated'",
        "/api/generate-v3",
        'value="violation">Нарушения',
        'id="siteSpeedThreshold"',
        'id="outsideSpeedThreshold"',
        "Порог маршрута и вне региона, км/ч",
        'id="plateFilter" type="text"',
        'id="plateSuggestions"',
        'class="sortable',
        'class="sort-indicator"',
        ".sort-indicator::before",
        ".sort-indicator::after",
        "cursor:pointer !important",
        "data-sort-index",
        "plateFilter.addEventListener('input'",
        'id="dbStatus"',
        "/api/database-status",
        'target="_blank"',
    ):
        assert token in html, f"В интерфейсе отсутствует обязательный элемент: {token}"

    duration = timedelta(hours=3, minutes=4, seconds=56, microseconds=176000)
    assert implementation.json_cell(duration) == "3:04:56"
    assert implementation.json_cell(33.700001) == "33.7"
    assert portal_runtime_patch.json_cell_one_decimal(11.0) == "11.0"
    assert run_report_portal.violation_web_preview is portal_runtime_patch.violation_web_preview
    assert (
        run_report_portal.generate_report_with_thresholds
        is portal_runtime_patch.generate_report_with_regional_summary
    )

    paths = {route.path for route in portal_entrypoint.app.routes}
    assert "/api/database-status" in paths
    assert "/api/generate-v2" in paths
    assert "/api/generate-v3" in paths

    with tempfile.TemporaryDirectory(prefix="arvento_portal_preview_") as temp_name:
        path = Path(temp_name) / "preview.xlsx"
        workbook = Workbook()
        workbook.active.title = TURN_SHEET_NAME
        workbook[TURN_SHEET_NAME].append(["№", "Госномер"])
        add_speed_preview_sheet(workbook, SITE_SHEET_NAME, 33.700001)
        add_speed_preview_sheet(workbook, ROUTE_SHEET_NAME, 111.700001)
        add_speed_preview_sheet(workbook, REGION_SHEET_NAME, 121.700001)
        workbook.save(path)
        workbook.close()

        _, rows, total = portal_runtime_patch.violation_web_preview(path)
        assert total == 3
        types = {row[1] for row in rows}
        assert types == {SITE_SHEET_NAME, ROUTE_SHEET_NAME, REGION_SHEET_NAME}
        values = {row[1]: row[5] for row in rows}
        assert values[SITE_SHEET_NAME] == "33.7"
        assert values[ROUTE_SHEET_NAME] == "111.7"
        assert values[REGION_SHEET_NAME] == "121.7"

    return True


def main() -> None:
    roster_22 = DatedRoster(Path("22.07.2026.xlsx"), date(2026, 7, 22), {})
    roster_24 = DatedRoster(Path("24.07.2026.xlsx"), date(2026, 7, 24), {})
    rosters = [roster_22, roster_24]
    assert select_roster(rosters, date(2026, 7, 24)) is roster_24
    assert select_roster(rosters, date(2026, 7, 23)) is roster_22
    try:
        select_roster(rosters, date(2026, 7, 20))
    except ValueError as exc:
        assert "нет разнарядки с этой или более ранней датой" in str(exc)
    else:
        raise AssertionError("Будущая разнарядка не должна применяться назад")

    route_polygon = load_kml_polygon(Path(__file__).resolve().parent / "route_akkuyu_tasucu.kml")
    assert len(route_polygon) >= 3
    assert polygon_area(route_polygon) > 1e-8
    assert len(CONSOLIDATED_HEADERS) == 25
    assert validated_speed_indices(consolidated_points(40, 45, 50)) == {0, 1, 2}
    assert not validated_speed_indices(consolidated_points(40, 120, 42))
    assert allowed_entry_exit_time(datetime(2026, 1, 1, 5, 0)) == time(5, 0)
    assert allowed_entry_exit_time(datetime(2026, 1, 1, 4, 59)) is None
    assert allowed_entry_exit_time(datetime(2026, 1, 1, 23, 1)) is None

    registry = load_registry(Path(__file__).resolve().parent / "geozones.json")
    boundary = find_site_boundary(registry)
    site_polygon = list(boundary.points or [])
    report_day = date(2026, 1, 1)
    assert has_night_site_mileage(
        report_day,
        movement_track(datetime(2026, 1, 1, 0, 30)),
        site_polygon,
    )
    assert has_night_site_mileage(
        report_day,
        movement_track(datetime(2026, 1, 1, 22, 30)),
        site_polygon,
    )
    assert has_night_site_mileage(
        report_day,
        movement_track(datetime(2026, 1, 1, 4, 59), seconds=60),
        site_polygon,
    )
    assert not has_night_site_mileage(
        report_day,
        movement_track(datetime(2026, 1, 1, 5, 0)),
        site_polygon,
    )
    assert not has_night_site_mileage(
        report_day,
        movement_track(datetime(2026, 1, 1, 12, 0)),
        site_polygon,
    )

    assert extract_date_from_text("23.07.2026 SON GUNCEL.xlsx") == date(2026, 7, 23)
    assert extract_date_from_text("gps_2026-07-24.csv") == date(2026, 7, 24)
    title, subtitle = build_report_titles(
        date(2026, 7, 24),
        date(2026, 7, 23),
        ReportFilters(time_from=time(7, 0), time_to=time(9, 0)),
    )
    assert title == (
        "Отчет по времени въезда служебных автомобилей в утреннее время за "
        "24.07.2026 (с 7:00 до 09:00) без учёта повторных проездов через геозону."
    )
    assert subtitle == (
        "(принадлежность водителей проставлена по разнарядке от 23.07.2026)"
    )

    assert round_metric(12.349) == 12.3
    assert round_metric(12.351) == 12.4
    assert round_ratio(0.32654) == 0.327

    assert _event_is_smooth(points(40, 45, 50))
    assert not _event_is_smooth(points(40, 120, 42))
    assert not _event_is_smooth(points(40, 45))
    assert not _event_is_smooth(points(40, 45, 50, seconds=1))

    assert validate_speed_thresholds(33, 104.5) == (33.0, 104.5)
    try:
        validate_speed_thresholds(120, 100)
    except ValueError:
        pass
    else:
        raise AssertionError("Некорректное соотношение порогов не отклонено")

    assert boundary.name == "Площадка АЭС АККУЮ"
    assert len(boundary.points or []) >= 3
    assert point_in_zone(36.145, 33.55, boundary)
    assert not point_in_zone(36.17, 33.55, boundary)

    jitter_track = [
        FakePoint(datetime(2026, 1, 1, 8, 0, 0), 40, 36.145, 33.55),
        FakePoint(datetime(2026, 1, 1, 8, 0, 10), 42, 36.17, 33.55),
        FakePoint(datetime(2026, 1, 1, 8, 0, 20), 44, 36.145, 33.55),
    ]
    states = [inside for _, inside in classify_site_state_by_polygon(jitter_track, registry)]
    assert states == [True, True, True]

    assert parse_coordinate_pair("36.145, 33.55") == (36.145, 33.55)
    assert google_maps_url(36.145, 33.55).startswith("https://www.google.com/maps/search/")

    check_regional_speed_runtime(registry, route_polygon)
    check_excel_rounding()
    portal_checked = check_portal_runtime()
    if portal_checked:
        print(
            "OK: runtime-проверки геозон, тоннелей, трёх категорий скорости, "
            "округления, сводного отчёта и production-портала пройдены."
        )
    else:
        print(
            "OK: runtime-проверки расчётной логики пройдены; production-портал "
            "будет полностью проверен внутри Docker."
        )


if __name__ == "__main__":
    main()
