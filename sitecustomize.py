from __future__ import annotations

"""Runtime-wide report rules.

- speeds inside ``purpose=speed_exclusion`` geozones are ignored;
- the violations portal exposes site, Tasucu–Akkuyu route and out-of-region tabs;
- measurable numeric indicators are displayed with one decimal in every XLSX.
"""

from copy import copy
from datetime import datetime
from pathlib import Path
from types import SimpleNamespace
from typing import Any, Sequence


def _install_speed_exclusions() -> None:
    try:
        import consolidated_report
        import speed_violation_report
        from geozone_registry import load_registry, point_in_zone
    except Exception:
        return

    registry_path = Path(__file__).resolve().parent / "geozones.json"
    try:
        registry = load_registry(registry_path)
    except Exception:
        return
    exclusion_zones = [
        zone for zone in registry.zones if zone.purpose == "speed_exclusion"
    ]
    if not exclusion_zones:
        return

    def excluded(point: Any) -> bool:
        try:
            lat = float(point.lat)
            lon = float(point.lon)
        except (AttributeError, TypeError, ValueError):
            return False
        return any(point_in_zone(lat, lon, zone) for zone in exclusion_zones)

    def without_speed(point: Any) -> Any:
        if not excluded(point):
            return point
        try:
            cloned = copy(point)
            cloned.speed = None
            return cloned
        except Exception:
            values = dict(getattr(point, "__dict__", {}))
            values.update(
                plate=getattr(point, "plate", ""),
                timestamp=getattr(point, "timestamp", getattr(point, "time", None)),
                time=getattr(point, "time", getattr(point, "timestamp", None)),
                lat=getattr(point, "lat", None),
                lon=getattr(point, "lon", None),
                speed=None,
                address=getattr(point, "address", ""),
            )
            return SimpleNamespace(**values)

    original_detect = speed_violation_report.detect_speed_violations
    if not getattr(original_detect, "_speed_exclusion_installed", False):
        def detect_speed_violations(
            track: Sequence[Any],
            registry: Any,
            site_threshold_kmh: float = speed_violation_report.DEFAULT_SITE_SPEED_THRESHOLD_KMH,
            outside_threshold_kmh: float = speed_violation_report.DEFAULT_OUTSIDE_SPEED_THRESHOLD_KMH,
        ):
            return original_detect(
                [without_speed(point) for point in track],
                registry,
                site_threshold_kmh=site_threshold_kmh,
                outside_threshold_kmh=outside_threshold_kmh,
            )
        detect_speed_violations._speed_exclusion_installed = True
        speed_violation_report.detect_speed_violations = detect_speed_violations

    original_analyze = consolidated_report.analyze_track
    if not getattr(original_analyze, "_speed_exclusion_installed", False):
        def analyze_track(day, display_plate, points, roster, site_polygon, route_polygon):
            return original_analyze(
                day,
                display_plate,
                [without_speed(point) for point in points],
                roster,
                site_polygon,
                route_polygon,
            )
        analyze_track._speed_exclusion_installed = True
        consolidated_report.analyze_track = analyze_track


def _install_excel_rounding() -> None:
    try:
        from openpyxl.styles.numbers import is_date_format
        from openpyxl.workbook.workbook import Workbook
    except Exception:
        return

    original_save = Workbook.save
    if getattr(original_save, "_one_decimal_installed", False):
        return

    coordinate_tokens = (
        "координат", "latitude", "longitude", "широт", "долгот", "lat", "lon"
    )

    def apply_rounding(workbook: Workbook) -> None:
        for sheet in workbook.worksheets:
            headers = {
                column: str(sheet.cell(1, column).value or "").strip().casefold()
                for column in range(1, sheet.max_column + 1)
            }
            for row in sheet.iter_rows(min_row=2):
                for cell in row:
                    value = cell.value
                    if not isinstance(value, float):
                        continue
                    header = headers.get(cell.column, "")
                    if any(token in header for token in coordinate_tokens):
                        continue
                    number_format = str(cell.number_format or "General")
                    if is_date_format(number_format):
                        continue
                    if "%" in number_format:
                        cell.number_format = "0.0%"
                    else:
                        cell.value = round(value, 1)
                        cell.number_format = "0.0"

    def save_with_rounding(self: Workbook, filename) -> None:
        apply_rounding(self)
        original_save(self, filename)

    save_with_rounding._one_decimal_installed = True
    Workbook.save = save_with_rounding


def _install_portal_region_view() -> None:
    try:
        import run_report_portal as portal
        from regional_speed_report import (
            REGION_SHEET_NAME,
            ROUTE_SHEET_NAME,
            SITE_SHEET_NAME,
        )
        from speed_violation_report import TURN_SHEET_NAME
        from openpyxl import load_workbook
    except Exception:
        return

    portal.SITE_SHEET_NAME = SITE_SHEET_NAME
    portal.OUTSIDE_SHEET_NAME = ROUTE_SHEET_NAME
    portal.REGION_SHEET_NAME = REGION_SHEET_NAME
    portal.implementation.HTML = portal.implementation.HTML.replace(
        "Порог вне площадки, км/ч", "Порог маршрута и вне региона, км/ч"
    )
    portal.implementation.HTML = portal.implementation.HTML.replace(
        "Порог вне площадки", "Порог маршрута и вне региона"
    )

    def violation_web_preview(path: Path):
        columns = [
            "Госномер", "Тип нарушения", "Дата", "Начало", "Окончание",
            "Максимальная скорость, км/ч", "Порог, км/ч", "Адрес", "Карта",
        ]
        records: list[tuple[str, datetime | None, list[Any]]] = []
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            for sheet_name in (
                TURN_SHEET_NAME,
                SITE_SHEET_NAME,
                ROUTE_SHEET_NAME,
                REGION_SHEET_NAME,
            ):
                if sheet_name not in workbook.sheetnames:
                    continue
                sheet = workbook[sheet_name]
                headers = portal._header_map(sheet)
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    plate = str(portal._value(row, headers, "Госномер") or "").strip()
                    if not plate:
                        continue
                    if sheet_name == TURN_SHEET_NAME:
                        start = portal._value(row, headers, "Начало прохода")
                        finish = portal._value(row, headers, "Окончание прохода")
                        speeds = [
                            portal._as_number(portal._value(row, headers, name))
                            for name in ("Скорость в начале", "Скорость в конце")
                        ]
                        speeds = [value for value in speeds if value is not None]
                        max_speed = max(speeds) if speeds else None
                        address = " → ".join(
                            value for value in (
                                str(portal._value(row, headers, "Адрес начала") or "").strip(),
                                str(portal._value(row, headers, "Адрес окончания") or "").strip(),
                            ) if value
                        )
                        map_url = portal._map_url(
                            portal._value(row, headers, "Координаты начала")
                            or portal._value(row, headers, "Координаты окончания")
                        )
                        threshold = None
                    else:
                        start = portal._value(row, headers, "Начало нарушения")
                        finish = portal._value(row, headers, "Окончание нарушения")
                        max_speed = portal._as_number(
                            portal._value(row, headers, "Максимальная скорость, км/ч")
                        )
                        threshold = portal._as_number(
                            portal._value(row, headers, "Порог фиксации, км/ч")
                        )
                        address = str(
                            portal._value(row, headers, "Адрес максимума") or ""
                        ).strip()
                        map_url = portal._map_url(
                            portal._value(row, headers, "Координаты максимума")
                        )
                    event_date = (
                        start.date() if isinstance(start, datetime)
                        else portal._value(row, headers, "Дата")
                    )
                    display = [
                        plate, sheet_name,
                        portal.implementation.json_cell(event_date),
                        portal.implementation.json_cell(start),
                        portal.implementation.json_cell(finish),
                        portal.implementation.json_cell(max_speed),
                        portal.implementation.json_cell(threshold),
                        address, map_url,
                    ]
                    records.append((
                        plate,
                        start if isinstance(start, datetime) else None,
                        display,
                    ))
        finally:
            workbook.close()
        records.sort(key=lambda item: (item[0], item[1] or datetime.min, item[2][1]))
        rows = [item[2] for item in records]
        return columns, rows, len(rows)

    portal.violation_web_preview = violation_web_preview
    original_generate = portal.generate_report_with_thresholds
    if not getattr(original_generate, "_regional_summary_installed", False):
        def generate_report_with_thresholds(*args, **kwargs):
            result = original_generate(*args, **kwargs)
            report_type = args[0] if args else kwargs.get("report_type")
            if report_type == "violation":
                summary = result.get("summary", {})
                value = summary.pop("Порог вне площадки", None)
                if value is not None:
                    summary["Порог маршрута и вне региона"] = value
            return result
        generate_report_with_thresholds._regional_summary_installed = True
        portal.generate_report_with_thresholds = generate_report_with_thresholds


_install_speed_exclusions()
_install_excel_rounding()
_install_portal_region_view()
