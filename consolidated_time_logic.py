#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Confirmed arrival/departure rules, worked hours, and date-only rendering."""
from __future__ import annotations

import re
from dataclasses import replace
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, Sequence

from openpyxl import load_workbook
from excel_formatting import save_report_workbook

import consolidated_report as core
from arvento_io import Point
from business_rules import (
    ENTRY_EXIT_TIME_FROM as WINDOW_START,
    ENTRY_EXIT_TIME_TO as WINDOW_END,
    SITE_EXIT_DISTANCE_THRESHOLD_KM,
)

PREFERRED_COMPANY_HEADER = "Эксплуатирующая фирма / ÇALIŞTIĞI FİRMA"

_ORIGINAL_ANALYZE_TRACK = core.analyze_track
_ORIGINAL_SAVE_REPORT = core.save_report
_TIME_LOGIC_PATCHED = False
_PREVIEW_PATCHED_IDS: set[int] = set()


def _apply_preferred_company_alias() -> None:
    """Prefer the exact operating-company column from the roster workbook."""
    aliases = tuple(getattr(core, "COMPANY_ALIASES", ()))
    preferred_normalized = core.normalized_text(PREFERRED_COMPANY_HEADER)
    ordered = (PREFERRED_COMPANY_HEADER,) + tuple(
        alias
        for alias in aliases
        if core.normalized_text(alias) != preferred_normalized
    )
    core.COMPANY_ALIASES = ordered
    core.ROSTER_ALIASES["company"] = ordered


_apply_preferred_company_alias()


def _day_moment(report_day: date, clock: time) -> datetime:
    return datetime.combine(report_day, clock)


def _site_state_at(
    track: Sequence[Point],
    states: Sequence[bool],
    moment: datetime,
    polygon: list[tuple[float, float]],
) -> bool:
    """Return the best known site state at a requested clock time."""
    if moment <= track[0].time:
        return bool(states[0])
    if moment >= track[-1].time:
        return bool(states[-1])

    for index, (p1, p2) in enumerate(zip(track, track[1:])):
        if moment > p2.time:
            continue
        state1 = bool(states[index])
        state2 = bool(states[index + 1])
        if state1 == state2 or p2.time <= p1.time:
            return state1
        fraction = core.polygon_crossing_fraction(p1, p2, polygon)
        crossing = p1.time + (p2.time - p1.time) * fraction
        return state1 if moment < crossing else state2
    return bool(states[-1])


def _crossings(
    track: Sequence[Point],
    states: Sequence[bool],
    polygon: list[tuple[float, float]],
) -> tuple[list[datetime], list[datetime]]:
    entries: list[datetime] = []
    exits: list[datetime] = []
    for index, (p1, p2) in enumerate(zip(track, track[1:])):
        if p2.time <= p1.time:
            continue
        state1 = bool(states[index])
        state2 = bool(states[index + 1])
        if state1 == state2:
            continue
        fraction = core.polygon_crossing_fraction(p1, p2, polygon)
        crossing = p1.time + (p2.time - p1.time) * fraction
        if not state1 and state2:
            entries.append(crossing)
        else:
            exits.append(crossing)
    return entries, exits


def _inside_seconds_between(
    track: Sequence[Point],
    states: Sequence[bool],
    polygon: list[tuple[float, float]],
    interval_start: datetime,
    interval_end: datetime,
) -> float:
    """Sum site-presence time strictly inside a confirmed entry/exit interval."""
    if interval_end <= interval_start:
        return 0.0

    total_seconds = 0.0
    for index, (p1, p2) in enumerate(zip(track, track[1:])):
        if p2.time <= p1.time:
            continue
        state1 = bool(states[index])
        state2 = bool(states[index + 1])
        if state1 == state2:
            portions = [(state1, 0.0, 1.0)]
        else:
            fraction = core.polygon_crossing_fraction(p1, p2, polygon)
            portions = [
                (state1, 0.0, fraction),
                (state2, fraction, 1.0),
            ]

        for is_inside, start_fraction, finish_fraction in portions:
            if not is_inside or finish_fraction <= start_fraction:
                continue
            part_start = p1.time + (p2.time - p1.time) * start_fraction
            part_finish = p1.time + (p2.time - p1.time) * finish_fraction
            clipped_start = max(part_start, interval_start)
            clipped_finish = min(part_finish, interval_end)
            if clipped_finish > clipped_start:
                total_seconds += (clipped_finish - clipped_start).total_seconds()

    return total_seconds


def calculate_operational_metrics(
    report_day: date,
    points: Sequence[Point],
    site_polygon: list[tuple[float, float]],
) -> tuple[time | None, time | None, float | None]:
    """Return confirmed entry, confirmed exit, and worked hours."""
    track = core.sanitize_position_outliers(points)
    if len(track) < 2:
        return None, None, None

    states = core.smooth_boolean_states([
        core.point_in_polygon(point.lat, point.lon, site_polygon)
        for point in track
    ])
    window_start = _day_moment(report_day, WINDOW_START)
    window_end = _day_moment(report_day, WINDOW_END)
    inside_at_start = _site_state_at(track, states, window_start, site_polygon)
    inside_at_end = _site_state_at(track, states, window_end, site_polygon)

    entries, exits = _crossings(track, states, site_polygon)
    entries = [value for value in entries if window_start <= value < window_end]
    exits = [value for value in exits if window_start < value <= window_end]

    confirmed_entry = entries[0] if not inside_at_start and entries else None
    confirmed_exit = exits[-1] if not inside_at_end and exits else None

    worked_hours: float | None = None
    if (
        confirmed_entry is not None
        and confirmed_exit is not None
        and confirmed_exit > confirmed_entry
    ):
        worked_hours = _inside_seconds_between(
            track,
            states,
            site_polygon,
            confirmed_entry,
            confirmed_exit,
        ) / 3600.0

    return (
        confirmed_entry.time().replace(tzinfo=None) if confirmed_entry else None,
        confirmed_exit.time().replace(tzinfo=None) if confirmed_exit else None,
        worked_hours,
    )


def calculate_arrival_departure(
    report_day: date,
    points: Sequence[Point],
    site_polygon: list[tuple[float, float]],
) -> tuple[time | None, time | None]:
    """Return only confirmed boundary-crossing entry and exit events."""
    arrived, departed, _worked_hours = calculate_operational_metrics(
        report_day,
        points,
        site_polygon,
    )
    return arrived, departed


def calculate_worked_hours(
    report_day: date,
    points: Sequence[Point],
    site_polygon: list[tuple[float, float]],
    *,
    inside_km: float,
) -> float | None:
    """Return worked hours only with mileage and confirmed entry plus exit."""
    if inside_km <= 0:
        return None
    _arrived, _departed, worked_hours = calculate_operational_metrics(
        report_day,
        points,
        site_polygon,
    )
    return worked_hours


def _outside_distance_is_insignificant(total_km: Any, inside_km: Any) -> bool:
    """Return True when total and inside mileage differ by no more than 10 km."""
    try:
        difference = abs(float(total_km or 0) - float(inside_km or 0))
    except (TypeError, ValueError):
        return False
    return difference <= SITE_EXIT_DISTANCE_THRESHOLD_KM


def analyze_track_with_operational_times(
    day: date,
    display_plate: str,
    points: Sequence[Point],
    roster: dict[str, core.RosterVehicle],
    site_polygon: list[tuple[float, float]],
    route_polygon: list[tuple[float, float]],
) -> core.ReportRow | None:
    row = _ORIGINAL_ANALYZE_TRACK(
        day,
        display_plate,
        points,
        roster,
        site_polygon,
        route_polygon,
    )
    if row is None:
        return None

    if _outside_distance_is_insignificant(row.total_km, row.inside_km):
        return replace(
            row,
            entry_time=None,
            exit_time=None,
            worked_hours=None,
        )

    arrived, departed, worked_hours = calculate_operational_metrics(
        day,
        points,
        site_polygon,
    )
    return replace(
        row,
        entry_time=arrived,
        exit_time=departed,
        worked_hours=(worked_hours if row.inside_km > 0 else None),
    )


def _has_confirmed_time(value: Any) -> bool:
    if value in (None, "", 0, 0.0):
        return False
    if isinstance(value, str) and value.strip() in {"", "0", "00:00", "00:00:00"}:
        return False
    return True


def _float_cell(value: Any) -> float | None:
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def save_report_with_date_only(*args: Any, **kwargs: Any) -> None:
    _ORIGINAL_SAVE_REPORT(*args, **kwargs)
    output_path = Path(args[0] if args else kwargs["output_path"])
    workbook = load_workbook(output_path)
    try:
        sheet = workbook["Сводный отчет"]
        headers = {str(cell.value or "").strip(): cell.column for cell in sheet[1]}
        date_column = headers.get("Дата")
        if date_column is not None:
            for row_index in range(2, sheet.max_row + 1):
                cell = sheet.cell(row_index, date_column)
                if isinstance(cell.value, datetime):
                    cell.value = cell.value.date()
                cell.number_format = "dd.mm.yyyy"

        total_km_column = headers.get("Пробег общий, км")
        inside_km_column = headers.get("Пробег внутри АЭС АККУЮ, км")
        entry_column = headers.get("Прибыл / Giriş")
        exit_column = headers.get("Убыл / Çıkış")
        worked_hours_column = headers.get("Всего отработано часов")

        for row_index in range(2, sheet.max_row + 1):
            total_km = (
                _float_cell(sheet.cell(row_index, total_km_column).value)
                if total_km_column is not None else None
            )
            inside_km = (
                _float_cell(sheet.cell(row_index, inside_km_column).value)
                if inside_km_column is not None else None
            )
            insignificant_outside_distance = (
                total_km is not None
                and inside_km is not None
                and _outside_distance_is_insignificant(total_km, inside_km)
            )

            if insignificant_outside_distance:
                if entry_column is not None:
                    sheet.cell(row_index, entry_column).value = None
                if exit_column is not None:
                    sheet.cell(row_index, exit_column).value = None
                if worked_hours_column is not None:
                    sheet.cell(row_index, worked_hours_column).value = None
                continue

            if worked_hours_column is None:
                continue
            has_site_mileage = inside_km is not None and inside_km > 0
            has_entry = entry_column is not None and _has_confirmed_time(
                sheet.cell(row_index, entry_column).value
            )
            has_exit = exit_column is not None and _has_confirmed_time(
                sheet.cell(row_index, exit_column).value
            )
            if not (has_site_mileage and has_entry and has_exit):
                sheet.cell(row_index, worked_hours_column).value = None

        parameters = workbook["Параметры"]
        for row_index in range(1, parameters.max_row + 1):
            label = parameters.cell(row_index, 1).value
            if label == "Допустимое время Прибыл/Убыл":
                parameters.cell(row_index, 2).value = "операционное окно 05:00–23:00"
        parameters.append([
            "Логика Прибыл",
            "только подтверждённое пересечение границы снаружи внутрь; "
            "если в 05:00 автомобиль уже внутри площадки, значение не заполняется",
        ])
        parameters.append([
            "Логика Убыл",
            "только подтверждённое пересечение границы изнутри наружу; "
            "если в 23:00 автомобиль остаётся внутри площадки, значение не заполняется",
        ])
        parameters.append([
            "Фильтр Прибыл/Убыл по пробегу",
            "если разница между общим пробегом и пробегом внутри площадки не превышает 10 км, "
            "поля Прибыл, Убыл и Всего отработано часов остаются пустыми",
        ])
        parameters.append([
            "Источник компании",
            f"столбец разнарядки «{PREFERRED_COMPANY_HEADER}» имеет приоритет",
        ])
        parameters.append([
            "Логика Всего отработано часов",
            "суммарное время внутри площадки между подтверждёнными въездом и выездом; "
            "без пробега внутри либо без одного из двух событий поле остаётся пустым",
        ])
        save_report_workbook(workbook, output_path)
    finally:
        workbook.close()


def apply_consolidated_time_logic() -> None:
    global _TIME_LOGIC_PATCHED
    _apply_preferred_company_alias()
    if _TIME_LOGIC_PATCHED:
        return
    core.analyze_track = analyze_track_with_operational_times
    core.save_report = save_report_with_date_only
    _TIME_LOGIC_PATCHED = True


def apply_consolidated_date_preview(implementation: Any) -> None:
    """Render the exact ``Дата`` column without a midnight time in the web table."""
    identity = id(implementation)
    if identity in _PREVIEW_PATCHED_IDS:
        return
    original = implementation.workbook_preview

    def workbook_preview_date_only(path: Path):
        columns, rows, total = original(path)
        date_indexes = [index for index, name in enumerate(columns) if str(name).strip() == "Дата"]
        for row in rows:
            for index in date_indexes:
                if index >= len(row):
                    continue
                value = row[index]
                if isinstance(value, datetime):
                    row[index] = value.strftime("%d.%m.%Y")
                elif isinstance(value, date):
                    row[index] = value.strftime("%d.%m.%Y")
                elif isinstance(value, str):
                    match = re.fullmatch(r"(\d{2}\.\d{2}\.\d{4})(?:\s+00:00(?::00)?)?", value.strip())
                    if match:
                        row[index] = match.group(1)
        return columns, rows, total

    implementation.workbook_preview = workbook_preview_date_only
    _PREVIEW_PATCHED_IDS.add(identity)


__all__ = [
    "PREFERRED_COMPANY_HEADER",
    "SITE_EXIT_DISTANCE_THRESHOLD_KM",
    "apply_consolidated_date_preview",
    "apply_consolidated_time_logic",
    "calculate_arrival_departure",
    "calculate_operational_metrics",
    "calculate_worked_hours",
]
