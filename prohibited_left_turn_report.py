#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""Excel-отчёт о запрещённом повороте налево.

Нарушение фиксируется, когда автомобиль проходит линейную геозону шириной
30 метров справа налево: от первой (восточной) точки полилинии к последней
(западной) не более чем за 5 минут.

Поддерживаемые исходники: XLSX, XLSM, CSV — координатные выгрузки Arvento.
"""

from __future__ import annotations

import argparse
import math
import shutil
import sqlite3
import tempfile
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from tkinter import Tk, filedialog
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from sqlite_store import import_source_to_sqlite


# Полилиния задана справа налево (с востока на запад).
CORRIDOR_POINTS: tuple[tuple[float, float], ...] = (
    (36.3172538740116, 33.874474367432896),
    (36.31719343922596, 33.874075574912666),
    (36.31714948662511, 33.8735029497554),
    (36.31720168033584, 33.87281614041499),
)

DEFAULT_WIDTH_M = 30.0
DEFAULT_MAX_SEQUENCE_SECONDS = 5 * 60
DEFAULT_COOLDOWN_SECONDS = 5 * 60
START_PROGRESS_MAX = 0.25
FINISH_PROGRESS_MIN = 0.75
MAX_BACKTRACK_M = 40.0


@dataclass(frozen=True)
class TrackPoint:
    plate: str
    timestamp: datetime
    lat: float
    lon: float
    speed: float | None
    address: str


@dataclass(frozen=True)
class CorridorPosition:
    distance_m: float
    along_m: float
    progress: float


@dataclass(frozen=True)
class Violation:
    plate: str
    start_point: TrackPoint
    finish_point: TrackPoint
    start_position: CorridorPosition
    finish_position: CorridorPosition
    min_distance_m: float
    max_progress: float
    point_count: int

    @property
    def start(self) -> datetime:
        return self.start_point.timestamp

    @property
    def finish(self) -> datetime:
        return self.finish_point.timestamp

    @property
    def duration_seconds(self) -> int:
        return max(0, int((self.finish - self.start).total_seconds()))


def choose_source() -> Path:
    root = Tk()
    root.withdraw()
    selected = filedialog.askopenfilename(
        title="Выберите координатную выгрузку Arvento",
        filetypes=[("Excel / CSV", "*.xlsx *.xlsm *.csv"), ("Все файлы", "*.*")],
    )
    root.destroy()
    if not selected:
        raise SystemExit("Файл не выбран")
    return Path(selected).resolve()


def local_xy_m(lat: float, lon: float, origin_lat: float, origin_lon: float) -> tuple[float, float]:
    y = (lat - origin_lat) * 111_320.0
    x = (lon - origin_lon) * 111_320.0 * math.cos(math.radians(origin_lat))
    return x, y


def build_polyline_xy() -> tuple[list[tuple[float, float]], list[float], float, float, float]:
    origin_lat = sum(p[0] for p in CORRIDOR_POINTS) / len(CORRIDOR_POINTS)
    origin_lon = sum(p[1] for p in CORRIDOR_POINTS) / len(CORRIDOR_POINTS)
    xy = [local_xy_m(lat, lon, origin_lat, origin_lon) for lat, lon in CORRIDOR_POINTS]
    cumulative = [0.0]
    for p1, p2 in zip(xy, xy[1:]):
        cumulative.append(cumulative[-1] + math.hypot(p2[0] - p1[0], p2[1] - p1[1]))
    return xy, cumulative, cumulative[-1], origin_lat, origin_lon


POLYLINE_XY, CUMULATIVE_M, POLYLINE_LENGTH_M, ORIGIN_LAT, ORIGIN_LON = build_polyline_xy()


def project_to_corridor(lat: float, lon: float) -> CorridorPosition:
    px, py = local_xy_m(lat, lon, ORIGIN_LAT, ORIGIN_LON)
    best_distance = float("inf")
    best_along = 0.0

    for index, (a, b) in enumerate(zip(POLYLINE_XY, POLYLINE_XY[1:])):
        vx, vy = b[0] - a[0], b[1] - a[1]
        length_sq = vx * vx + vy * vy
        if length_sq <= 0:
            continue
        t = ((px - a[0]) * vx + (py - a[1]) * vy) / length_sq
        t = max(0.0, min(1.0, t))
        qx, qy = a[0] + t * vx, a[1] + t * vy
        distance = math.hypot(px - qx, py - qy)
        if distance < best_distance:
            segment_length = math.sqrt(length_sq)
            best_distance = distance
            best_along = CUMULATIVE_M[index] + t * segment_length

    progress = best_along / POLYLINE_LENGTH_M if POLYLINE_LENGTH_M else 0.0
    return CorridorPosition(best_distance, best_along, progress)


def iter_vehicle_tracks(db_path: Path) -> Iterable[tuple[str, list[TrackPoint]]]:
    with sqlite3.connect(db_path) as connection:
        plates = [row[0] for row in connection.execute("SELECT DISTINCT plate FROM points ORDER BY plate")]
        for index, plate in enumerate(plates, start=1):
            rows = connection.execute(
                """
                SELECT plate, ts, lat, lon, speed, address
                FROM points
                WHERE plate=?
                ORDER BY ts
                """,
                (plate,),
            )
            track = [
                TrackPoint(
                    plate=row[0],
                    timestamp=datetime.fromisoformat(row[1]),
                    lat=float(row[2]),
                    lon=float(row[3]),
                    speed=float(row[4]) if row[4] is not None else None,
                    address=row[5] or "",
                )
                for row in rows
            ]
            if track:
                yield plate, track
            if index % 100 == 0 or index == len(plates):
                print(f"Обработано автомобилей: {index}/{len(plates)}")


def detect_violations(
    track: list[TrackPoint],
    width_m: float,
    max_sequence_seconds: int,
    cooldown_seconds: int,
) -> list[Violation]:
    half_width = width_m / 2.0
    violations: list[Violation] = []
    active: dict[str, object] | None = None
    last_violation_time: datetime | None = None

    for point in track:
        position = project_to_corridor(point.lat, point.lon)
        inside = position.distance_m <= half_width

        if last_violation_time and (point.timestamp - last_violation_time).total_seconds() < cooldown_seconds:
            continue

        if active is not None:
            start_point = active["start_point"]
            assert isinstance(start_point, TrackPoint)
            elapsed = (point.timestamp - start_point.timestamp).total_seconds()
            if elapsed <= 0 or elapsed > max_sequence_seconds:
                active = None

        if not inside:
            continue

        # Старт фиксируется только у правой (восточной) части коридора.
        if active is None:
            if position.progress <= START_PROGRESS_MAX:
                active = {
                    "start_point": point,
                    "start_position": position,
                    "last_along": position.along_m,
                    "max_progress": position.progress,
                    "min_distance": position.distance_m,
                    "point_count": 1,
                }
            continue

        last_along = float(active["last_along"])
        # Существенное движение обратно вправо сбрасывает кандидат.
        if position.along_m < last_along - MAX_BACKTRACK_M:
            active = None
            if position.progress <= START_PROGRESS_MAX:
                active = {
                    "start_point": point,
                    "start_position": position,
                    "last_along": position.along_m,
                    "max_progress": position.progress,
                    "min_distance": position.distance_m,
                    "point_count": 1,
                }
            continue

        active["last_along"] = max(last_along, position.along_m)
        active["max_progress"] = max(float(active["max_progress"]), position.progress)
        active["min_distance"] = min(float(active["min_distance"]), position.distance_m)
        active["point_count"] = int(active["point_count"]) + 1

        if position.progress < FINISH_PROGRESS_MIN:
            continue

        start_point = active["start_point"]
        start_position = active["start_position"]
        assert isinstance(start_point, TrackPoint)
        assert isinstance(start_position, CorridorPosition)
        violations.append(
            Violation(
                plate=point.plate,
                start_point=start_point,
                finish_point=point,
                start_position=start_position,
                finish_position=position,
                min_distance_m=float(active["min_distance"]),
                max_progress=float(active["max_progress"]),
                point_count=int(active["point_count"]),
            )
        )
        last_violation_time = point.timestamp
        active = None

    return violations


def style_sheet(sheet) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.showGridLines = False
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="C00000")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[1].height = 45
    for column in range(1, sheet.max_column + 1):
        width = min(
            max(len(str(sheet.cell(row, column).value or "")) for row in range(1, min(sheet.max_row, 300) + 1)) + 2,
            34,
        )
        sheet.column_dimensions[get_column_letter(column)].width = max(width, 12)


def save_report(path: Path, violations: list[Violation], source: Path, width_m: float, max_seconds: int) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Нарушения"
    sheet.append([
        "№", "Госномер", "Дата", "Начало прохода", "Окончание прохода",
        "Продолжительность", "Скорость в начале", "Скорость в конце",
        "Прогресс начала, %", "Прогресс конца, %", "Мин. отклонение от линии, м",
        "Точек в коридоре", "Координаты начала", "Координаты окончания",
        "Адрес начала", "Адрес окончания",
    ])

    for number, item in enumerate(violations, start=1):
        sheet.append([
            number,
            item.plate,
            item.start.date(),
            item.start,
            item.finish,
            item.duration_seconds / 86400.0,
            item.start_point.speed,
            item.finish_point.speed,
            item.start_position.progress,
            item.finish_position.progress,
            round(item.min_distance_m, 1),
            item.point_count,
            f"{item.start_point.lat:.7f}, {item.start_point.lon:.7f}",
            f"{item.finish_point.lat:.7f}, {item.finish_point.lon:.7f}",
            item.start_point.address,
            item.finish_point.address,
        ])

    for row in sheet.iter_rows(min_row=2):
        row[2].number_format = "dd.mm.yyyy"
        row[3].number_format = "dd.mm.yyyy hh:mm:ss"
        row[4].number_format = "dd.mm.yyyy hh:mm:ss"
        row[5].number_format = "[h]:mm:ss"
        row[6].number_format = "0.0"
        row[7].number_format = "0.0"
        row[8].number_format = "0.0%"
        row[9].number_format = "0.0%"
        row[10].number_format = "0.0"
    style_sheet(sheet)

    settings = workbook.create_sheet("Параметры")
    settings.append(["Параметр", "Значение"])
    settings.append(["Исходный файл", str(source)])
    settings.append(["Направление", "справа налево (точка 1 → точка 4)"])
    settings.append(["Ширина линейной геозоны, м", width_m])
    settings.append(["Максимальное время прохода", max_seconds / 86400.0])
    settings.cell(5, 2).number_format = "[h]:mm:ss"
    settings.append(["Длина осевой линии, м", round(POLYLINE_LENGTH_M, 1)])
    settings.append(["Начальная часть коридора", f"0–{START_PROGRESS_MAX * 100:.0f}%"])
    settings.append(["Конечная часть коридора", f"{FINISH_PROGRESS_MIN * 100:.0f}–100%"])
    for index, (lat, lon) in enumerate(CORRIDOR_POINTS, start=1):
        settings.append([f"Точка осевой линии {index}", f"{lat}, {lon}"])
    settings.append(["Количество нарушений", len(violations)])
    style_sheet(settings)
    workbook.save(path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Отчёт о запрещённом повороте налево")
    parser.add_argument("source", nargs="?", help="XLSX, XLSM или CSV выгрузка Arvento")
    parser.add_argument("output", nargs="?", help="Путь итогового XLSX")
    parser.add_argument("--width", type=float, default=DEFAULT_WIDTH_M, help="Ширина линейной геозоны, м")
    parser.add_argument("--max-minutes", type=float, default=DEFAULT_MAX_SEQUENCE_SECONDS / 60.0)
    parser.add_argument("--cooldown-minutes", type=float, default=DEFAULT_COOLDOWN_SECONDS / 60.0)
    args = parser.parse_args()

    source = Path(args.source).expanduser().resolve() if args.source else choose_source()
    if not source.exists():
        raise SystemExit(f"Файл не найден: {source}")
    output = Path(args.output).expanduser().resolve() if args.output else source.with_name(source.stem + "_запрещенный_поворот.xlsx")

    temp_dir = Path(tempfile.mkdtemp(prefix="arvento_left_turn_"))
    db_path = temp_dir / "points.sqlite3"
    try:
        print(f"Исходный файл: {source}")
        print("Импорт данных во временную SQLite-базу...")
        stats = import_source_to_sqlite(source, db_path)
        violations: list[Violation] = []
        for _, track in iter_vehicle_tracks(db_path):
            violations.extend(detect_violations(
                track,
                width_m=max(1.0, args.width),
                max_sequence_seconds=max(1, int(args.max_minutes * 60)),
                cooldown_seconds=max(0, int(args.cooldown_minutes * 60)),
            ))
        violations.sort(key=lambda item: (item.start, item.plate))
        save_report(output, violations, source, max(1.0, args.width), max(1, int(args.max_minutes * 60)))
        print(f"Готово: {output}")
        print(f"Загружено GPS-точек: {stats['loaded']}")
        print(f"Найдено нарушений: {len(violations)}")
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)


if __name__ == "__main__":
    main()
