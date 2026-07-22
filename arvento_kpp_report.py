#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""Формирование итогового отчёта по KPP 4/KPP 5 из выгрузки Arvento."""

from __future__ import annotations

import math
import statistics
import sys
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

KPP4_REGION = "KPP 4 TEST"
KPP5_REGION = "KPP 5 TEST"
KPP5_POINT_1 = (36.15886644084745, 33.549775396083724)
KPP5_POINT_2 = (36.159167964778234, 33.55101996362877)

SIDE_HYSTERESIS_M = 18.0
GATE_MARGIN_M = 45.0
MAX_CROSSING_WINDOW_SECONDS = 15 * 60
EVENT_COOLDOWN_SECONDS = 120
MAX_ODOMETER_DELTA_KM = 20.0
MAX_REASONABLE_SPEED_KMH = 180.0

# Скорость не выше этого значения считается стоянкой.
STOP_SPEED_KMH = 1.0

# Большие разрывы сообщений не включаются во время движения/стоянки,
# чтобы пропавшая связь не считалась длительной стоянкой.
MAX_TIME_ACCOUNT_GAP_SECONDS = 5 * 60

ALIASES = {
    "plate": ["номерной знак", "госномер", "plaka", "license plate"],
    "time": ["дата / время", "дата/время", "tarih / saat", "date / time"],
    "odometer": ["одометр", "odometer"],
    "distance": ["расстояние (км)", "расстояние", "distance"],
    "speed": ["скорость", "speed"],
    "region": ["область", "регион", "region"],
    "address": ["адрес", "address"],
    "lat": ["latitudine", "широта", "latitude", "enlem"],
    "lon": ["долгота", "longitude", "boylam"],
}


@dataclass
class Point:
    plate: str
    time: datetime
    lat: float
    lon: float
    odometer: Optional[float] = None
    source_distance: Optional[float] = None
    speed: Optional[float] = None
    region: str = ""
    address: str = ""


@dataclass
class Gate:
    name: str
    p1: tuple[float, float]
    p2: tuple[float, float]
    sample_count: int


def norm(value: Any) -> str:
    return (
        " ".join(str(value or "").replace("\n", " ").split())
        .strip()
        .lower()
        .replace("ё", "е")
    )


def as_float(value: Any) -> Optional[float]:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip().replace(" ", "").replace(",", "."))
    except ValueError:
        return None


def as_datetime(value: Any) -> Optional[datetime]:
    if isinstance(value, datetime):
        return value
    if value in (None, ""):
        return None
    text = str(value).strip()
    for fmt in (
        "%d.%m.%Y %H:%M:%S",
        "%d.%m.%Y %H:%M",
        "%Y-%m-%d %H:%M:%S",
        "%m/%d/%Y %H:%M:%S",
    ):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def local_xy(
    lat: float,
    lon: float,
    origin_lat: float,
    origin_lon: float,
) -> tuple[float, float]:
    y = (lat - origin_lat) * 111_320.0
    x = (lon - origin_lon) * 111_320.0 * math.cos(math.radians(origin_lat))
    return x, y


def gate_geometry(gate: Gate) -> tuple[float, float, float, float, float]:
    mid_lat = (gate.p1[0] + gate.p2[0]) / 2.0
    mid_lon = (gate.p1[1] + gate.p2[1]) / 2.0
    x1, y1 = local_xy(*gate.p1, mid_lat, mid_lon)
    x2, y2 = local_xy(*gate.p2, mid_lat, mid_lon)
    vx, vy = x2 - x1, y2 - y1
    length = math.hypot(vx, vy)
    if length < 0.1:
        vx, vy, length = 20.0, 0.0, 20.0
    return mid_lat, mid_lon, vx, vy, length


def gate_coords(point: Point, gate: Gate) -> tuple[float, float]:
    mid_lat, mid_lon, vx, vy, length = gate_geometry(gate)
    px, py = local_xy(point.lat, point.lon, mid_lat, mid_lon)
    ux, uy = vx / length, vy / length
    nx, ny = -uy, ux
    along = px * ux + py * uy
    across = px * nx + py * ny
    if ny < 0:
        across = -across
    return along, across


def side(point: Point, gate: Gate) -> int:
    _, across = gate_coords(point, gate)
    if across > SIDE_HYSTERESIS_M:
        return 1
    if across < -SIDE_HYSTERESIS_M:
        return -1
    return 0


def near_gate(point: Point, gate: Gate) -> bool:
    along, across = gate_coords(point, gate)
    length = gate_geometry(gate)[4]
    return (
        abs(across) <= GATE_MARGIN_M
        and abs(along) <= length / 2.0 + GATE_MARGIN_M
    )


def crossing_fraction(p1: Point, p2: Point, gate: Gate) -> float:
    a1 = gate_coords(p1, gate)[1]
    a2 = gate_coords(p2, gate)[1]
    delta = a2 - a1
    if abs(delta) < 1e-12:
        return 0.5
    return max(0.0, min(1.0, -a1 / delta))


def haversine_km(p1: Point, p2: Point) -> float:
    radius = 6371.0088
    lat1, lat2 = math.radians(p1.lat), math.radians(p2.lat)
    dlat = math.radians(p2.lat - p1.lat)
    dlon = math.radians(p2.lon - p1.lon)
    value = (
        math.sin(dlat / 2.0) ** 2
        + math.cos(lat1) * math.cos(lat2) * math.sin(dlon / 2.0) ** 2
    )
    return radius * 2.0 * math.atan2(math.sqrt(value), math.sqrt(1.0 - value))


def segment_distance(p1: Point, p2: Point) -> float:
    gap = (p2.time - p1.time).total_seconds()
    if gap <= 0:
        return 0.0

    if p1.odometer is not None and p2.odometer is not None:
        delta = p2.odometer - p1.odometer
        if 0 <= delta <= MAX_ODOMETER_DELTA_KM:
            return delta

    if (
        p2.source_distance is not None
        and 0 <= p2.source_distance <= MAX_ODOMETER_DELTA_KM
    ):
        return p2.source_distance

    gps = haversine_km(p1, p2)
    calculated_speed = gps / (gap / 3600.0)
    if calculated_speed <= MAX_REASONABLE_SPEED_KMH:
        return gps
    return 0.0


def segment_speed_kmh(p1: Point, p2: Point, distance_km: float) -> float:
    speeds = [value for value in (p1.speed, p2.speed) if value is not None]
    if speeds:
        return max(speeds)

    gap = (p2.time - p1.time).total_seconds()
    if gap <= 0:
        return 0.0
    return distance_km / (gap / 3600.0)


def choose_file() -> Path:
    if len(sys.argv) > 1:
        return Path(sys.argv[1]).expanduser().resolve()

    from tkinter import Tk, filedialog

    root = Tk()
    root.withdraw()
    selected = filedialog.askopenfilename(
        title="Выберите выгрузку Arvento",
        filetypes=[("Excel", "*.xlsx *.xlsm"), ("Все файлы", "*.*")],
    )
    root.destroy()
    if not selected:
        raise SystemExit("Файл не выбран")
    return Path(selected).resolve()


def load_points(path: Path) -> tuple[list[Point], dict[str, int]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    rows = sheet.iter_rows(values_only=True)

    try:
        headers = [norm(value) for value in next(rows)]
    except StopIteration as exc:
        workbook.close()
        raise ValueError("Первый лист пуст") from exc

    columns: dict[str, int] = {}
    for key, aliases in ALIASES.items():
        for index, header in enumerate(headers):
            if any(norm(alias) == header or norm(alias) in header for alias in aliases):
                columns[key] = index
                break

    missing = [key for key in ("plate", "time", "lat", "lon") if key not in columns]
    if missing:
        workbook.close()
        raise ValueError("Не найдены колонки: " + ", ".join(missing))

    points: list[Point] = []
    skipped = 0

    for row in rows:
        plate = str(row[columns["plate"]] or "").strip()
        timestamp = as_datetime(row[columns["time"]])
        lat = as_float(row[columns["lat"]])
        lon = as_float(row[columns["lon"]])

        if not plate or timestamp is None or lat is None or lon is None:
            skipped += 1
            continue

        points.append(
            Point(
                plate=plate,
                time=timestamp,
                lat=lat,
                lon=lon,
                odometer=(
                    as_float(row[columns["odometer"]])
                    if "odometer" in columns
                    else None
                ),
                source_distance=(
                    as_float(row[columns["distance"]])
                    if "distance" in columns
                    else None
                ),
                speed=(
                    as_float(row[columns["speed"]])
                    if "speed" in columns
                    else None
                ),
                region=(
                    str(row[columns["region"]] or "").strip()
                    if "region" in columns
                    else ""
                ),
                address=(
                    str(row[columns["address"]] or "").strip()
                    if "address" in columns
                    else ""
                ),
            )
        )

    workbook.close()
    return points, {"loaded": len(points), "skipped": skipped}


def build_gates(points: list[Point]) -> list[Gate]:
    kpp4_points = [
        point
        for point in points
        if KPP4_REGION.lower() in point.region.lower()
    ]
    if not kpp4_points:
        raise ValueError(f"Не найдена геозона {KPP4_REGION}")

    lat = statistics.median(point.lat for point in kpp4_points)
    lon = statistics.median(point.lon for point in kpp4_points)
    lon_delta = 25.0 / (111_320.0 * math.cos(math.radians(lat)))

    gate4 = Gate(
        name=KPP4_REGION,
        p1=(lat, lon - lon_delta / 2.0),
        p2=(lat, lon + lon_delta / 2.0),
        sample_count=len(kpp4_points),
    )

    kpp5_count = sum(
        KPP5_REGION.lower() in point.region.lower()
        for point in points
    )
    gate5 = Gate(
        name=KPP5_REGION,
        p1=KPP5_POINT_1,
        p2=KPP5_POINT_2,
        sample_count=kpp5_count,
    )
    return [gate4, gate5]


def add_inside_time(
    seconds: float,
    speed_kmh: float,
    totals: dict[str, float],
) -> None:
    if seconds <= 0 or seconds > MAX_TIME_ACCOUNT_GAP_SECONDS:
        return
    totals["inside_seconds"] += seconds
    if speed_kmh <= STOP_SPEED_KMH:
        totals["stopped_seconds"] += seconds
    else:
        totals["moving_seconds"] += seconds


def analyze(points: list[Point], gates: list[Gate]) -> list[dict[str, Any]]:
    grouped: dict[str, list[Point]] = defaultdict(list)
    for point in points:
        grouped[point.plate].append(point)

    summaries: list[dict[str, Any]] = []

    for plate in sorted(grouped):
        vehicle_points = sorted(grouped[plate], key=lambda point: point.time)
        if len(vehicle_points) < 2:
            continue

        stable_side = {gate.name: side(vehicle_points[0], gate) for gate in gates}
        stable_time = {gate.name: vehicle_points[0].time for gate in gates}

        inside = False
        first_event: Optional[str] = None
        entries = 0
        exits = 0
        inside_km = 0.0
        outside_km = 0.0
        total_km = 0.0
        last_event_time: Optional[datetime] = None
        time_totals = defaultdict(float)

        for p1, p2 in zip(vehicle_points, vehicle_points[1:]):
            gap_seconds = (p2.time - p1.time).total_seconds()
            if gap_seconds <= 0:
                continue

            distance_km = segment_distance(p1, p2)
            speed_kmh = segment_speed_kmh(p1, p2, distance_km)
            total_km += distance_km

            candidates = []
            for gate in gates:
                side2 = side(p2, gate)
                previous_side = stable_side[gate.name]

                if side2 != 0 and previous_side != 0 and side2 != previous_side:
                    elapsed = (p2.time - stable_time[gate.name]).total_seconds()
                    if (
                        elapsed <= MAX_CROSSING_WINDOW_SECONDS
                        and (near_gate(p1, gate) or near_gate(p2, gate))
                    ):
                        kind = (
                            "Въезд"
                            if previous_side == 1 and side2 == -1
                            else "Выезд"
                        )
                        candidates.append(
                            (crossing_fraction(p1, p2, gate), gate, kind)
                        )

                if side2 != 0:
                    stable_side[gate.name] = side2
                    stable_time[gate.name] = p2.time

            candidate = min(candidates, key=lambda item: item[0]) if candidates else None

            if candidate:
                fraction, gate, kind = candidate
                event_time = p1.time + (p2.time - p1.time) * fraction
                if (
                    last_event_time is not None
                    and (event_time - last_event_time).total_seconds()
                    < EVENT_COOLDOWN_SECONDS
                ):
                    candidate = None

            if candidate is None:
                if inside:
                    inside_km += distance_km
                    add_inside_time(gap_seconds, speed_kmh, time_totals)
                else:
                    outside_km += distance_km
                continue

            fraction, gate, kind = candidate
            event_time = p1.time + (p2.time - p1.time) * fraction
            before_km = distance_km * fraction
            after_km = distance_km * (1.0 - fraction)
            before_seconds = gap_seconds * fraction
            after_seconds = gap_seconds * (1.0 - fraction)

            if first_event is None:
                first_event = kind
                if kind == "Выезд":
                    inside = True

            if kind == "Въезд":
                if inside:
                    inside_km += distance_km
                    add_inside_time(gap_seconds, speed_kmh, time_totals)
                else:
                    outside_km += before_km
                    inside_km += after_km
                    add_inside_time(after_seconds, speed_kmh, time_totals)
                    inside = True
                    entries += 1
                    last_event_time = event_time
            else:
                if inside:
                    inside_km += before_km
                    outside_km += after_km
                    add_inside_time(before_seconds, speed_kmh, time_totals)
                    inside = False
                    exits += 1
                    last_event_time = event_time
                else:
                    outside_km += distance_km

        inside_percent = inside_km / total_km if total_km > 0 else 0.0
        outside_percent = outside_km / total_km if total_km > 0 else 0.0
        inside_seconds = time_totals["inside_seconds"]
        stopped_percent = (
            time_totals["stopped_seconds"] / inside_seconds
            if inside_seconds > 0
            else 0.0
        )
        moving_percent = (
            time_totals["moving_seconds"] / inside_seconds
            if inside_seconds > 0
            else 0.0
        )

        summaries.append(
            {
                "plate": plate,
                "first_time": vehicle_points[0].time,
                "last_time": vehicle_points[-1].time,
                "start_state": "Внутри" if first_event == "Выезд" else "Снаружи",
                "end_state": "Внутри" if inside else "Снаружи",
                "entries": entries,
                "exits": exits,
                "inside_km": inside_km,
                "outside_km": outside_km,
                "total_km": total_km,
                "inside_percent": inside_percent,
                "outside_percent": outside_percent,
                "inside_seconds": inside_seconds,
                "moving_seconds": time_totals["moving_seconds"],
                "stopped_seconds": time_totals["stopped_seconds"],
                "moving_percent": moving_percent,
                "stopped_percent": stopped_percent,
            }
        )

    return summaries


def style_sheet(sheet) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.showGridLines = False

    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    sheet.row_dimensions[1].height = 45

    for column in range(1, sheet.max_column + 1):
        width = min(
            max(
                len(str(sheet.cell(row, column).value or ""))
                for row in range(1, min(sheet.max_row, 300) + 1)
            )
            + 2,
            32,
        )
        sheet.column_dimensions[get_column_letter(column)].width = max(width, 11)


def save_report(path: Path, summaries: list[dict[str, Any]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Итоговый отчёт"

    headers = [
        "Госномер",
        "Начало периода",
        "Конец периода",
        "Состояние в начале",
        "Состояние в конце",
        "Въездов",
        "Выездов",
        "Пробег внутри, км",
        "Пробег снаружи, км",
        "Общий пробег, км",
        "Пробег внутри, %",
        "Пробег снаружи, %",
        "Время на территории",
        "Время в движении на территории",
        "Время стоянки на территории",
        "Движение на территории, %",
        "Стоянка на территории, %",
    ]
    sheet.append(headers)

    for item in summaries:
        sheet.append(
            [
                item["plate"],
                item["first_time"],
                item["last_time"],
                item["start_state"],
                item["end_state"],
                item["entries"],
                item["exits"],
                item["inside_km"],
                item["outside_km"],
                item["total_km"],
                item["inside_percent"],
                item["outside_percent"],
                item["inside_seconds"] / 86400.0,
                item["moving_seconds"] / 86400.0,
                item["stopped_seconds"] / 86400.0,
                item["moving_percent"],
                item["stopped_percent"],
            ]
        )

    for row in sheet.iter_rows(min_row=2):
        row[1].number_format = "dd.mm.yyyy hh:mm:ss"
        row[2].number_format = "dd.mm.yyyy hh:mm:ss"

        for index in range(7, 10):
            row[index].number_format = "0.000"

        row[10].number_format = "0.0%"
        row[11].number_format = "0.0%"

        for index in range(12, 15):
            row[index].number_format = "[h]:mm:ss"

        row[15].number_format = "0.0%"
        row[16].number_format = "0.0%"

    style_sheet(sheet)
    workbook.save(path)


def main() -> None:
    source = choose_file()
    if not source.exists():
        raise SystemExit(f"Файл не найден: {source}")

    print(f"Чтение: {source}")
    points, stats = load_points(source)
    if not points:
        raise SystemExit("В файле нет пригодных координатных сообщений")

    gates = build_gates(points)
    for gate in gates:
        print(
            f"{gate.name}: {gate.p1} -> {gate.p2}; "
            f"точек зоны: {gate.sample_count}"
        )

    summaries = analyze(points, gates)
    output = source.with_name(source.stem + "_КПП_отчет.xlsx")
    save_report(output, summaries)

    print(f"Готово: {output}")
    print(
        f"Автомобилей: {len(summaries)}; "
        f"загружено точек: {stats['loaded']}; "
        f"пропущено строк: {stats['skipped']}"
    )


if __name__ == "__main__":
    main()
