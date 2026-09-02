#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Standalone consolidated daily vehicle report.

This module is intentionally not connected to the web portal. It reads GPS
points either from PostgreSQL or from an Arvento CSV/XLSX export, enriches them
with one roster workbook and writes an XLSX file for validation.
"""
from __future__ import annotations

import argparse
import math
import os
import xml.etree.ElementTree as ET
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from pathlib import Path
from types import SimpleNamespace
from typing import Any, Iterator, Sequence
from zoneinfo import ZoneInfo

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from excel_formatting import save_report_workbook

from arvento_analysis import segment_distance
from arvento_io import Point, load_points
from business_rules import (
    ENTRY_EXIT_TIME_FROM,
    ENTRY_EXIT_TIME_TO,
    GEOFENCE_VIOLATION_KM,
    NIGHT_END,
    NIGHT_START,
    MIN_SPEED_EVENT_DURATION_SECONDS,
    PERSONAL_USE_DISTANCE_DIFF_KM,
    PERSONAL_USE_PERCENT_DIFF,
    TIMEZONE_NAME,
)
from geozone_registry import (
    Geozone,
    find_site_boundary,
    load_registry,
    point_in_polygon,
    suppress_speed_in_exclusions,
)
from roster_registry import normalize_plate, parse_date_value
from speed_violation_report import (
    MAX_SPEED_EVENT_GAP_SECONDS,
    MAX_VALID_GPS_SPEED_KMH,
    _event_is_smooth,
)

TZ = ZoneInfo(TIMEZONE_NAME)
APP_DIR = Path(__file__).resolve().parent
DEFAULT_ROUTE_KML = APP_DIR / "route_akkuyu_tasucu.kml"
DEFAULT_GEOZONES = APP_DIR / "geozones.json"

MOVEMENT_SPEED_KMH = 3.0
MOVEMENT_SEGMENT_KM = 0.02
POSITION_OUTLIER_SPEED_KMH = 300.0
MAX_REPORT_DAYS = 31

COMPANY_ALIASES = (
    "Компания или фирма", "Компания", "Фирма", "FİRMA", "FIRMA", "ŞİRKET",
    "SIRKET", "COMPANY", "ALT YÜKLENİCİ", "ALT YUKLENICI", "TAŞERON",
    "TASERON", "ПОДРЯДЧИК",
)
ROSTER_ALIASES = {
    "plate": ("Гос рег знак", "PLAKA", "Госномер", "Номерной знак"),
    "company": COMPANY_ALIASES,
    "user": ("ПОЛЬЗОВАТЕЛЬ", "KULLANICI", "Пользователь", "Водитель"),
    "grade": ("Грейд", "SCALA", "Grade"),
}

HEADERS = [
    "Дата",
    "Компания или фирма",
    "Госномер / Plaka",
    "ПОЛЬЗОВАТЕЛЬ / KULLANICI",
    "Грейд / SCALA",
    "Максимальная скорость / Maximum Hız",
    "Максимальная скорость только на маршруте Аккую - Ташуджу / Maximum Hız Rota Akkuyu-Taşucu",
    "Максимальная скорость на АЭС Аккую",
    "Пробег общий, км",
    "Пробег внутри АЭС АККУЮ, км",
    "Пробег вне площадки, км",
    "Разница между пробегами, км",
    "Пробег внутри АЭС АККУЮ, %",
    "Пробег вне площадки, %",
    "Разница между пробегами, %",
    "Дата выезда",
    "Дата прибытия",
    "День недели",
    "Прибыл / Giriş",
    "Убыл / Çıkış",
    "Всего отработано часов",
    "Нарушение геозоны / Sınır ihlali",
    "Использование ТС в личных целях",
    "Hafta sonu çalışmaları / Работа в выходные дни",
    "Gece vardiyasında çalışmalar / Работа ночью",
]

WEEKDAYS_RU = (
    "Понедельник", "Вторник", "Среда", "Четверг",
    "Пятница", "Суббота", "Воскресенье",
)


@dataclass(frozen=True, slots=True)
class RosterVehicle:
    plate: str
    company: str = ""
    user: str = ""
    grade: str = ""


@dataclass(frozen=True, slots=True)
class ReportRow:
    day: date
    company: str
    plate: str
    user: str
    grade: str
    max_speed: float | None
    route_max_speed: float | None
    site_max_speed: float | None
    total_km: float
    inside_km: float
    outside_km: float
    distance_difference_km: float
    inside_percent: float
    outside_percent: float
    percent_difference: float
    departure: datetime | None
    arrival: datetime | None
    weekday: str
    entry_time: time | None
    exit_time: time | None
    worked_hours: float
    boundary_violation: int
    personal_use: int
    weekend_work: int
    night_work: int
    in_roster: bool
    raw_points: int
    retained_points: int
    valid_speed_points: int
    max_distance_from_site_km: float


def clean(value: Any) -> str:
    return " ".join(str(value or "").replace("\n", " ").split()).strip()


def normalized_text(value: Any) -> str:
    return clean(value).casefold().replace("ё", "е").replace("ı", "i")


def find_column(headers: Sequence[Any], aliases: Sequence[str]) -> int | None:
    normalized = [normalized_text(value) for value in headers]
    for alias in aliases:
        needle = normalized_text(alias)
        for index, header in enumerate(normalized):
            if header == needle or needle in header:
                return index
    return None


def load_roster(path: Path) -> tuple[dict[str, RosterVehicle], date | None]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        roster_date = parse_date_value(path.stem)
        if roster_date is None:
            for sheet in workbook.worksheets[:3]:
                for row in sheet.iter_rows(min_row=1, max_row=12, max_col=30, values_only=True):
                    for value in row:
                        roster_date = parse_date_value(value)
                        if roster_date is not None:
                            break
                    if roster_date is not None:
                        break
                if roster_date is not None:
                    break

        vehicles: dict[str, RosterVehicle] = {}
        for sheet in workbook.worksheets:
            iterator = iter(sheet.iter_rows(values_only=True))
            columns: dict[str, int | None] | None = None
            for _ in range(80):
                try:
                    row = next(iterator)
                except StopIteration:
                    break
                candidate = {
                    key: find_column(row, aliases)
                    for key, aliases in ROSTER_ALIASES.items()
                }
                if candidate["plate"] is not None:
                    columns = candidate
                    break
            if columns is None:
                continue

            for row in iterator:
                plate_index = columns["plate"]
                if plate_index is None or plate_index >= len(row):
                    continue
                plate = normalize_plate(row[plate_index])
                if not plate:
                    continue

                def value(key: str) -> str:
                    index = columns.get(key)
                    return clean(row[index]) if index is not None and index < len(row) else ""

                item = RosterVehicle(
                    plate=plate,
                    company=value("company"),
                    user=value("user"),
                    grade=value("grade"),
                )
                old = vehicles.get(plate)
                old_score = (
                    sum(bool(getattr(old, field)) for field in ("company", "user", "grade"))
                    if old else -1
                )
                new_score = sum(
                    bool(getattr(item, field)) for field in ("company", "user", "grade")
                )
                if new_score > old_score:
                    vehicles[plate] = item

        if not vehicles:
            raise ValueError("В разнарядке не найден столбец с госномерами или данные автомобилей")
        return vehicles, roster_date
    finally:
        workbook.close()


def load_kml_polygon(path: Path) -> list[tuple[float, float]]:
    if not path.exists():
        raise ValueError(f"Не найден KML маршрута: {path}")
    root = ET.parse(path).getroot()
    coordinates_text = None
    for element in root.iter():
        if element.tag.rsplit("}", 1)[-1] == "coordinates" and element.text:
            coordinates_text = element.text
            break
    if not coordinates_text:
        raise ValueError("В KML не найден полигон с координатами")

    points: list[tuple[float, float]] = []
    for token in coordinates_text.split():
        parts = token.split(",")
        if len(parts) < 2:
            continue
        lon = float(parts[0])
        lat = float(parts[1])
        points.append((lat, lon))
    if len(points) < 3:
        raise ValueError("В KML маршрута недостаточно точек полигона")
    if points[0] == points[-1]:
        points.pop()
    return points


def valid_speed(value: Any) -> float | None:
    try:
        speed = float(value)
    except (TypeError, ValueError):
        return None
    if not math.isfinite(speed) or speed < 0 or speed > MAX_VALID_GPS_SPEED_KMH:
        return None
    return speed


def haversine_km_coords(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    radius = 6371.0088
    p1 = math.radians(lat1)
    p2 = math.radians(lat2)
    dp = math.radians(lat2 - lat1)
    dl = math.radians(lon2 - lon1)
    value = math.sin(dp / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dl / 2) ** 2
    return radius * 2 * math.atan2(math.sqrt(value), math.sqrt(max(0.0, 1.0 - value)))


def implied_speed_kmh(p1: Point, p2: Point) -> float:
    seconds = (p2.time - p1.time).total_seconds()
    if seconds <= 0:
        return math.inf
    return haversine_km_coords(p1.lat, p1.lon, p2.lat, p2.lon) / (seconds / 3600.0)


def sanitize_position_outliers(points: Sequence[Point]) -> list[Point]:
    """Remove isolated impossible coordinate jumps while preserving real trips."""
    source = [
        point for point in sorted(points, key=lambda item: item.time)
        if math.isfinite(point.lat) and math.isfinite(point.lon)
        and -90 <= point.lat <= 90 and -180 <= point.lon <= 180
    ]
    if len(source) < 3:
        return source

    rejected: set[int] = set()
    for index in range(1, len(source) - 1):
        previous = source[index - 1]
        current = source[index]
        following = source[index + 1]
        before = implied_speed_kmh(previous, current)
        after = implied_speed_kmh(current, following)
        direct = implied_speed_kmh(previous, following)
        if (
            before > POSITION_OUTLIER_SPEED_KMH
            and after > POSITION_OUTLIER_SPEED_KMH
            and direct <= MAX_VALID_GPS_SPEED_KMH
        ):
            rejected.add(index)
    return [point for index, point in enumerate(source) if index not in rejected]


def validated_speed_indices(points: Sequence[Point]) -> set[int]:
    """Return point indexes participating in at least one sustained smooth window."""
    valid: set[int] = set()
    chunk: list[tuple[int, Point]] = []

    def close_chunk() -> None:
        nonlocal chunk
        if len(chunk) < 3:
            chunk = []
            return
        for start in range(len(chunk)):
            for finish in range(start + 2, min(len(chunk), start + 7)):
                window = chunk[start:finish + 1]
                samples = [
                    SimpleNamespace(timestamp=point.time, speed=float(point.speed))
                    for _, point in window
                ]
                if _event_is_smooth(samples):
                    valid.update(index for index, _ in window)
        chunk = []

    previous_time: datetime | None = None
    for index, point in enumerate(points):
        speed = valid_speed(point.speed)
        gap = (point.time - previous_time).total_seconds() if previous_time is not None else None
        if speed is None or (gap is not None and (gap <= 0 or gap > MAX_SPEED_EVENT_GAP_SECONDS)):
            close_chunk()
        if speed is not None:
            chunk.append((index, point))
        previous_time = point.time
    close_chunk()
    return valid


def smooth_boolean_states(raw_states: Sequence[bool]) -> list[bool]:
    states = list(raw_states)
    for index in range(1, len(states) - 1):
        if raw_states[index - 1] == raw_states[index + 1] != raw_states[index]:
            states[index] = raw_states[index - 1]
    return states


def polygon_crossing_fraction(p1: Point, p2: Point, polygon: list[tuple[float, float]]) -> float:
    """Binary-search the first site-boundary crossing on a segment."""
    start_state = point_in_polygon(p1.lat, p1.lon, polygon)
    low, high = 0.0, 1.0
    for _ in range(32):
        middle = (low + high) / 2.0
        lat = p1.lat + (p2.lat - p1.lat) * middle
        lon = p1.lon + (p2.lon - p1.lon) * middle
        if point_in_polygon(lat, lon, polygon) == start_state:
            low = middle
        else:
            high = middle
    return (low + high) / 2.0


def interval_overlaps_night(start: datetime, finish: datetime) -> bool:
    if finish <= start:
        return False
    day = start.date() - timedelta(days=1)
    final_day = finish.date()
    while day <= final_day:
        night_start = datetime.combine(day, NIGHT_START)
        night_finish = datetime.combine(day + timedelta(days=1), NIGHT_END)
        if max(start, night_start) < min(finish, night_finish):
            return True
        day += timedelta(days=1)
    return False


def distance_point_to_segment_m(
    point_lat: float,
    point_lon: float,
    a_lat: float,
    a_lon: float,
    b_lat: float,
    b_lon: float,
) -> float:
    cos_lat = math.cos(math.radians(point_lat))
    ax = (a_lon - point_lon) * 111_320.0 * cos_lat
    ay = (a_lat - point_lat) * 111_320.0
    bx = (b_lon - point_lon) * 111_320.0 * cos_lat
    by = (b_lat - point_lat) * 111_320.0
    dx, dy = bx - ax, by - ay
    if dx == 0 and dy == 0:
        return math.hypot(ax, ay)
    projection = max(0.0, min(1.0, -(ax * dx + ay * dy) / (dx * dx + dy * dy)))
    qx = ax + projection * dx
    qy = ay + projection * dy
    return math.hypot(qx, qy)


def distance_to_polygon_km(lat: float, lon: float, polygon: list[tuple[float, float]]) -> float:
    if point_in_polygon(lat, lon, polygon):
        return 0.0
    distances = []
    for index, (a_lat, a_lon) in enumerate(polygon):
        b_lat, b_lon = polygon[(index + 1) % len(polygon)]
        distances.append(distance_point_to_segment_m(lat, lon, a_lat, a_lon, b_lat, b_lon))
    return min(distances) / 1000.0 if distances else math.inf


def allowed_entry_exit_time(value: datetime | None) -> time | None:
    if value is None:
        return None
    clock = value.time().replace(tzinfo=None)
    return clock if ENTRY_EXIT_TIME_FROM <= clock <= ENTRY_EXIT_TIME_TO else None


def analyze_track(
    day: date,
    display_plate: str,
    points: Sequence[Point],
    roster: dict[str, RosterVehicle],
    site_polygon: list[tuple[float, float]],
    route_polygon: list[tuple[float, float]],
) -> ReportRow | None:
    normalized_plate = normalize_plate(display_plate)
    raw_count = len(points)
    track = sanitize_position_outliers(points)
    if len(track) < 2:
        return None

    roster_info = roster.get(normalized_plate)
    speed_indexes = validated_speed_indices(track)
    site_states = smooth_boolean_states([
        point_in_polygon(point.lat, point.lon, site_polygon) for point in track
    ])

    max_speed = None
    route_max_speed = None
    site_max_speed = None
    for index in speed_indexes:
        point = track[index]
        speed = valid_speed(point.speed)
        if speed is None:
            continue
        max_speed = speed if max_speed is None else max(max_speed, speed)
        if point_in_polygon(point.lat, point.lon, route_polygon):
            route_max_speed = speed if route_max_speed is None else max(route_max_speed, speed)
        if site_states[index]:
            site_max_speed = speed if site_max_speed is None else max(site_max_speed, speed)

    total_km = inside_km = outside_km = inside_seconds = 0.0
    departure: datetime | None = None
    arrival: datetime | None = None
    first_entry: datetime | None = None
    last_exit: datetime | None = None
    night_work = 0

    for index, (p1, p2) in enumerate(zip(track, track[1:])):
        gap = (p2.time - p1.time).total_seconds()
        if gap <= 0:
            continue
        distance = segment_distance(p1, p2)
        if not math.isfinite(distance) or distance < 0:
            distance = 0.0
        total_km += distance

        p1_inside = site_states[index]
        p2_inside = site_states[index + 1]
        if p1_inside == p2_inside:
            portions = [(p1_inside, 0.0, 1.0)]
        else:
            fraction = polygon_crossing_fraction(p1, p2, site_polygon)
            crossing_time = p1.time + (p2.time - p1.time) * fraction
            if not p1_inside and p2_inside:
                if first_entry is None:
                    first_entry = crossing_time
                portions = [(False, 0.0, fraction), (True, fraction, 1.0)]
            else:
                last_exit = crossing_time
                portions = [(True, 0.0, fraction), (False, fraction, 1.0)]

        for is_inside, start_fraction, finish_fraction in portions:
            fraction = max(0.0, finish_fraction - start_fraction)
            part_distance = distance * fraction
            if is_inside:
                inside_km += part_distance
                inside_seconds += gap * fraction
                if part_distance > 0:
                    part_start = p1.time + (p2.time - p1.time) * start_fraction
                    part_finish = p1.time + (p2.time - p1.time) * finish_fraction
                    if interval_overlaps_night(part_start, part_finish):
                        night_work = 1
            else:
                outside_km += part_distance

        speed1 = valid_speed(p1.speed) or 0.0
        speed2 = valid_speed(p2.speed) or 0.0
        moving = distance >= MOVEMENT_SEGMENT_KM or max(speed1, speed2) >= MOVEMENT_SPEED_KMH
        if moving:
            if departure is None:
                departure = p1.time
            arrival = p2.time

    inside_percent = inside_km / total_km if total_km else 0.0
    outside_percent = outside_km / total_km if total_km else 0.0
    distance_difference = outside_km - inside_km
    percent_difference = outside_percent - inside_percent

    max_distance = max(
        (distance_to_polygon_km(point.lat, point.lon, site_polygon) for point in track),
        default=0.0,
    )
    boundary_violation = int(max_distance > GEOFENCE_VIOLATION_KM)
    personal_use = int(
        distance_difference > PERSONAL_USE_DISTANCE_DIFF_KM
        or percent_difference > PERSONAL_USE_PERCENT_DIFF
    )
    weekend_work = int(day.weekday() >= 5 and inside_km > 0.01)

    return ReportRow(
        day=day,
        company=roster_info.company if roster_info else "",
        plate=display_plate,
        user=roster_info.user if roster_info else "",
        grade=roster_info.grade if roster_info else "",
        max_speed=max_speed,
        route_max_speed=route_max_speed,
        site_max_speed=site_max_speed,
        total_km=total_km,
        inside_km=inside_km,
        outside_km=outside_km,
        distance_difference_km=distance_difference,
        inside_percent=inside_percent,
        outside_percent=outside_percent,
        percent_difference=percent_difference,
        departure=departure,
        arrival=arrival,
        weekday=WEEKDAYS_RU[day.weekday()],
        entry_time=allowed_entry_exit_time(first_entry),
        exit_time=allowed_entry_exit_time(last_exit),
        worked_hours=inside_seconds / 3600.0,
        boundary_violation=boundary_violation,
        personal_use=personal_use,
        weekend_work=weekend_work,
        night_work=night_work,
        in_roster=roster_info is not None,
        raw_points=raw_count,
        retained_points=len(track),
        valid_speed_points=len(speed_indexes),
        max_distance_from_site_km=max_distance,
    )


def iter_database_tracks(
    database_url: str,
    start_day: date,
    end_day: date,
) -> Iterator[tuple[date, str, list[Point]]]:
    try:
        import psycopg
    except ModuleNotFoundError as exc:
        raise RuntimeError("Для чтения PostgreSQL требуется пакет psycopg") from exc

    start = datetime.combine(start_day, time.min, tzinfo=TZ)
    finish = datetime.combine(end_day + timedelta(days=1), time.min, tzinfo=TZ)
    query = """
        SELECT
            (event_time AT TIME ZONE 'Europe/Istanbul')::date AS local_day,
            COALESCE(NULLIF(plate, ''), normalized_plate) AS display_plate,
            event_time AT TIME ZONE 'Europe/Istanbul' AS local_time,
            latitude,
            longitude,
            speed_kmh,
            distance_km,
            COALESCE(address, '')
        FROM gps_points
        WHERE event_time >= %s AND event_time < %s
        ORDER BY local_day, normalized_plate, event_time
    """
    with psycopg.connect(database_url) as connection:
        with connection.cursor(name="consolidated_report_points") as cursor:
            cursor.itersize = 20_000
            cursor.execute(query, (start, finish))
            key = None
            display_plate = ""
            points: list[Point] = []
            for row in cursor:
                day_value, row_plate, local_time, lat, lon, speed, distance, address = row
                row_key = (day_value, normalize_plate(row_plate))
                if key is not None and row_key != key:
                    yield key[0], display_plate, points
                    points = []
                key = row_key
                display_plate = clean(row_plate)
                points.append(
                    Point(
                        plate=display_plate,
                        time=local_time,
                        lat=float(lat),
                        lon=float(lon),
                        speed=float(speed) if speed is not None else None,
                        source_distance=float(distance) if distance is not None else None,
                        address=clean(address),
                    )
                )
            if key is not None and points:
                yield key[0], display_plate, points


def iter_source_tracks(
    path: Path,
    start_day: date,
    end_day: date,
) -> Iterator[tuple[date, str, list[Point]]]:
    points, _ = load_points(path)
    grouped: dict[tuple[date, str], list[Point]] = defaultdict(list)
    display_plates: dict[tuple[date, str], str] = {}
    for point in points:
        day = point.time.date()
        if not start_day <= day <= end_day:
            continue
        key = (day, normalize_plate(point.plate))
        grouped[key].append(point)
        display_plates[key] = clean(point.plate)
    for key in sorted(grouped):
        yield key[0], display_plates[key], grouped[key]


def report_row_values(row: ReportRow) -> list[Any]:
    def one_decimal(value: float | None) -> float | None:
        return round(value, 1) if value is not None else None

    return [
        row.day,
        row.company,
        row.plate,
        row.user,
        row.grade,
        one_decimal(row.max_speed),
        one_decimal(row.route_max_speed),
        one_decimal(row.site_max_speed),
        one_decimal(row.total_km),
        one_decimal(row.inside_km),
        one_decimal(row.outside_km),
        one_decimal(row.distance_difference_km),
        round(row.inside_percent, 3),
        round(row.outside_percent, 3),
        round(row.percent_difference, 3),
        row.departure,
        row.arrival,
        row.weekday,
        row.entry_time,
        row.exit_time,
        one_decimal(row.worked_hours),
        row.boundary_violation,
        row.personal_use,
        row.weekend_work,
        row.night_work,
    ]


def style_header(sheet) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[1].height = 72
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.showGridLines = False


def save_report(
    output_path: Path,
    rows: Sequence[ReportRow],
    roster_path: Path,
    roster_date: date | None,
    route_path: Path,
    site_zone: Geozone,
    start_day: date,
    end_day: date,
    source_description: str,
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Сводный отчет"
    sheet.append(HEADERS)

    missing_fill = PatternFill("solid", fgColor="FECACA")
    flag_fill = PatternFill("solid", fgColor="FDE68A")
    for item in sorted(rows, key=lambda row: (row.day, row.company, normalize_plate(row.plate))):
        sheet.append(report_row_values(item))
        excel_row = sheet.max_row
        if not item.in_roster:
            cell = sheet.cell(excel_row, 3)
            cell.fill = missing_fill
            cell.font = Font(bold=True, color="9C0006")
            cell.comment = Comment("Автомобиль отсутствует в загруженной разнарядке", "Arvento")
        for column in (22, 23, 24, 25):
            if sheet.cell(excel_row, column).value == 1:
                sheet.cell(excel_row, column).fill = flag_fill

    style_header(sheet)
    for row in sheet.iter_rows(min_row=2):
        row[0].number_format = "dd.mm.yyyy"
        for index in (5, 6, 7, 8, 9, 10, 11, 20):
            row[index].number_format = "0.0"
        for index in (12, 13, 14):
            row[index].number_format = "0.0%"
        for index in (15, 16):
            row[index].number_format = "dd.mm.yyyy hh:mm:ss"
        for index in (18, 19):
            row[index].number_format = "hh:mm:ss"
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    widths = [
        12, 24, 18, 34, 12, 18, 38, 22, 16, 20, 20, 20, 20, 20, 20,
        22, 22, 16, 14, 14, 20, 18, 22, 22, 22,
    ]
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    diagnostics = workbook.create_sheet("Диагностика")
    diagnostics.append([
        "Дата", "Госномер", "В разнарядке", "Исходных GPS-точек",
        "Точек после удаления координатных выбросов", "Точек для расчёта максимальной скорости",
        "Максимальное удаление от АЭС, км",
    ])
    for item in sorted(rows, key=lambda row: (row.day, normalize_plate(row.plate))):
        diagnostics.append([
            item.day,
            item.plate,
            "да" if item.in_roster else "нет",
            item.raw_points,
            item.retained_points,
            item.valid_speed_points,
            round(item.max_distance_from_site_km, 1),
        ])
    style_header(diagnostics)
    for row in diagnostics.iter_rows(min_row=2):
        row[0].number_format = "dd.mm.yyyy"
        row[6].number_format = "0.0"
    for index, width in enumerate((12, 18, 16, 22, 34, 30, 28), 1):
        diagnostics.column_dimensions[get_column_letter(index)].width = width

    parameters = workbook.create_sheet("Параметры")
    parameter_rows = [
        ("Параметр", "Значение"),
        ("Период", start_day.strftime("%d.%m.%Y") if start_day == end_day else f"{start_day:%d.%m.%Y}–{end_day:%d.%m.%Y}"),
        ("Источник GPS", source_description),
        ("Разнарядка", str(roster_path)),
        ("Дата разнарядки", roster_date.strftime("%d.%m.%Y") if roster_date else "не определена"),
        ("Граница АЭС", site_zone.name),
        ("Полигон маршрута", str(route_path)),
        ("Максимально допустимая GPS-скорость", f"{MAX_VALID_GPS_SPEED_KMH:g} км/ч"),
        (
            "Проверка максимальной скорости",
            f"минимум 3 плавные точки и не менее {MIN_SPEED_EVENT_DURATION_SECONDS} секунд; "
            "одиночные скачки исключаются",
        ),
        ("Начало движения", f"скорость ≥ {MOVEMENT_SPEED_KMH:g} км/ч или сегмент ≥ {MOVEMENT_SEGMENT_KM:g} км"),
        ("Допустимое время Прибыл/Убыл", "05:00–23:00"),
        ("Нарушение геозоны", f"удаление от границы АЭС более {GEOFENCE_VIOLATION_KM:g} км"),
        (
            "Личное использование",
            f"внешний пробег больше внутреннего >{PERSONAL_USE_DISTANCE_DIFF_KM:g} км "
            f"или >{PERSONAL_USE_PERCENT_DIFF * 100:g} процентных пунктов",
        ),
        ("Разница пробегов", "пробег вне площадки минус пробег внутри площадки"),
        ("Разница процентов", "процент вне площадки минус процент внутри площадки"),
        ("Работа ночью", "есть пробег внутри АЭС в интервале 22:00–05:00"),
        ("Работа в выходные", "есть пробег внутри АЭС в субботу или воскресенье"),
    ]
    for row in parameter_rows:
        parameters.append(row)
    parameters["A1"].font = Font(bold=True, color="FFFFFF")
    parameters["B1"].font = Font(bold=True, color="FFFFFF")
    parameters["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    parameters["B1"].fill = PatternFill("solid", fgColor="1F4E78")
    parameters.column_dimensions["A"].width = 42
    parameters.column_dimensions["B"].width = 105
    for row in parameters.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    save_report_workbook(workbook, output_path)


def parse_day(value: str) -> date:
    try:
        return date.fromisoformat(value)
    except ValueError as exc:
        raise argparse.ArgumentTypeError("Дата должна быть в формате YYYY-MM-DD") from exc


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Сводный ежедневный отчёт по служебным автомобилям")
    period = parser.add_mutually_exclusive_group(required=True)
    period.add_argument("--date", type=parse_day, help="Один день, YYYY-MM-DD")
    period.add_argument("--date-from", type=parse_day, help="Начало периода, YYYY-MM-DD")
    parser.add_argument("--date-to", type=parse_day, help="Конец периода, YYYY-MM-DD")
    parser.add_argument("--roster", type=Path, required=True, help="Разнарядка XLSX/XLSM")
    parser.add_argument("--output", type=Path, required=True, help="Выходной XLSX")
    parser.add_argument("--source", type=Path, help="Необязательная выгрузка Arvento CSV/XLSX вместо PostgreSQL")
    parser.add_argument("--database-url", default=os.environ.get("DATABASE_URL", ""))
    parser.add_argument("--route-kml", type=Path, default=DEFAULT_ROUTE_KML)
    parser.add_argument("--geozones", type=Path, default=DEFAULT_GEOZONES)
    return parser


def main() -> None:
    args = build_parser().parse_args()
    if args.date:
        start_day = end_day = args.date
    else:
        start_day = args.date_from
        end_day = args.date_to or start_day
    if start_day is None or end_day is None:
        raise ValueError("Не задан период")
    if end_day < start_day:
        raise ValueError("Дата окончания раньше даты начала")
    if (end_day - start_day).days + 1 > MAX_REPORT_DAYS:
        raise ValueError(f"Период не должен превышать {MAX_REPORT_DAYS} дней")
    if not args.roster.exists():
        raise ValueError(f"Разнарядка не найдена: {args.roster}")

    roster, roster_date = load_roster(args.roster)
    registry = load_registry(args.geozones)
    site_zone = find_site_boundary(registry)
    site_polygon = list(site_zone.points or [])
    route_polygon = registry.route_polygon or load_kml_polygon(args.route_kml)

    if args.source:
        if not args.source.exists():
            raise ValueError(f"Выгрузка Arvento не найдена: {args.source}")
        tracks = iter_source_tracks(args.source, start_day, end_day)
        source_description = str(args.source)
    else:
        if not args.database_url:
            raise ValueError("DATABASE_URL не задан и --source не указан")
        tracks = iter_database_tracks(args.database_url, start_day, end_day)
        source_description = "PostgreSQL gps_points"

    result: list[ReportRow] = []
    processed = 0
    for day, plate, points in tracks:
        analysis_points = suppress_speed_in_exclusions(points, registry)
        item = analyze_track(
            day, plate, analysis_points, roster, site_polygon, route_polygon
        )
        if item is not None:
            result.append(item)
        processed += 1
        if processed % 100 == 0:
            print(f"Обработано автомобилей-дней: {processed}")

    if not result:
        raise ValueError("За выбранный период нет пригодных данных для отчёта")

    save_report(
        args.output,
        result,
        args.roster,
        roster_date,
        args.route_kml,
        site_zone,
        start_day,
        end_day,
        source_description,
    )
    missing = sum(not row.in_roster for row in result)
    print(f"Готово: {args.output}")
    print(f"Строк: {len(result)}; отсутствуют в разнарядке: {missing}")


if __name__ == "__main__":
    main()
