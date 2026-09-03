#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Offline regression checks for the optimized consolidated XLSX export path."""
from __future__ import annotations

import tempfile
from datetime import date, datetime, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook

import consolidated_cache_portal as cache_portal
from consolidated_export_layout import EXPORT_HEADERS, REPORT_SHEET, RosterExportDetails
from consolidated_export_optimization import (
    _add_cached_fuel_and_parameters,
    _apply_mileage_review,
    _apply_roster_layout,
    _preview_loaded,
    apply_cached_workbook_optimization,
    write_cached_workbook_fast,
)
from excel_formatting import save_report_workbook
from fuel_enriched_consolidated_report import FUEL_HEADER
from mileage_review_policy import (
    MileageReviewCandidate,
    REVIEW_HEADER,
    REVIEW_SHEET,
    REVIEW_VALUE,
)


REPORT_DAY = date(2026, 9, 2)
PLATE = "34 ABC 123"
NORMALIZED = "34ABC123"


def main() -> int:
    with tempfile.TemporaryDirectory(prefix="verify_fast_export_") as temp_name:
        output_path = Path(temp_name) / "report.xlsx"
        workbook = Workbook()
        report = workbook.active
        report.title = REPORT_SHEET
        report.append([
            "Дата",
            "Компания или фирма",
            "Госномер / Plaka",
            "ПОЛЬЗОВАТЕЛЬ / KULLANICI",
            "Грейд / SCALA",
            "Пробег общий, км",
        ])
        report.append([
            REPORT_DAY,
            "OLD COMPANY",
            PLATE,
            "Старый водитель",
            "7a",
            123.456,
        ])
        parameters = workbook.create_sheet("Параметры")
        parameters.append(["Параметр", "Значение"])
        workbook.create_sheet("Диагностика")
        workbook.save(output_path)
        workbook.close()

        workbook = load_workbook(output_path)
        try:
            fuel_total, fuel_rows = _add_cached_fuel_and_parameters(
                workbook,
                {(REPORT_DAY, NORMALIZED): 12.3},
                datetime(2026, 9, 3, 0, 30, tzinfo=timezone.utc),
            )
            layout = _apply_roster_layout(
                workbook,
                {
                    (REPORT_DAY, NORMALIZED): RosterExportDetails(
                        driver_name="Актуальный водитель",
                        grade="9b",
                        directorate="Дирекция",
                        responsible="Ответственный",
                    )
                },
            )
            review = _apply_mileage_review(
                workbook,
                {
                    (REPORT_DAY, NORMALIZED): MileageReviewCandidate(
                        report_day=REPORT_DAY,
                        normalized_plate=NORMALIZED,
                        authoritative_km=123.456,
                        coordinate_km=100.0,
                        gap_km=23.456,
                        gap_percent=23.456,
                        reason="Тест",
                    )
                },
            )
            save_report_workbook(workbook, output_path)
            columns, rows, total_rows = _preview_loaded(workbook)
        finally:
            workbook.close()

        assert fuel_total == 12.3
        assert fuel_rows == 1
        assert layout["enriched_rows"] == 1
        assert review["candidates"] == 1
        assert review["flagged_rows"] == 1
        assert total_rows == 1
        assert rows and rows[0][columns.index(FUEL_HEADER)] == 12.3
        assert rows[0][columns.index(REVIEW_HEADER)] == REVIEW_VALUE

        result = load_workbook(output_path, read_only=True, data_only=True)
        try:
            assert result.sheetnames == [REPORT_SHEET, REVIEW_SHEET], result.sheetnames
            sheet = result[REPORT_SHEET]
            headers = [str(cell.value or "") for cell in sheet[1]]
            assert FUEL_HEADER in headers
            assert tuple(headers[-5:-1]) == EXPORT_HEADERS, headers
            assert headers[-1] == REVIEW_HEADER, headers
            assert sheet.cell(2, headers.index("Имя водителя") + 1).value == "Актуальный водитель"
            assert sheet.cell(2, headers.index("Грейд") + 1).value == "9b"
            assert sheet.cell(2, headers.index("Дирекция") + 1).value == "Дирекция"
            assert sheet.cell(2, headers.index("Ответственный") + 1).value == "Ответственный"
            assert result[REVIEW_SHEET].cell(2, 9).value == REVIEW_VALUE
        finally:
            result.close()

    apply_cached_workbook_optimization()
    assert cache_portal.write_cached_workbook is write_cached_workbook_fast

    print(
        "OK: cached export preserves fuel/roster/review contract while using the optimized writer"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
