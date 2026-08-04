#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Persist and restore the subdivision responsible person from rosters.

The established roster store already keeps driver, grade, position and
Directorate. This patch adds one optional ``responsible`` field sourced only
from ``Ответственный по подразделению / Birim sorumlusu``. The neighboring
``Ответственный / Sorumlu`` column is intentionally not used.
"""
from __future__ import annotations

import io
import tempfile
from collections import defaultdict
from datetime import date
from pathlib import Path
from typing import Any, Sequence

import psycopg
from openpyxl import Workbook, load_workbook

import consolidated_cache as cache
import extended_roster_fields as extended
from consolidated_multi_report import load_rosters
from roster_registry import normalize_plate

_BASE_ENSURE_SCHEMA = extended.ensure_schema
_BASE_SAVE_ROSTER_UPLOADS = extended.save_roster_uploads
_PATCHED = False

PLATE_ALIASES = (
    "Гос рег знак / PLAKA",
    "Гос рег знак",
    "PLAKA",
    "Госномер",
    "Номерной знак",
    "License Plate",
)
RESPONSIBLE_ALIASES = (
    "Ответственный по подразделению / Birim sorumlusu",
    "Ответственный по подразделению",
    "Birim sorumlusu",
)


def _clean(value: Any) -> str:
    return " ".join(str(value or "").replace("\n", " ").split()).strip()


def _normalized(value: Any) -> str:
    return _clean(value).casefold().replace("ё", "е").replace("ı", "i")


def _find_column(values: Sequence[Any], aliases: Sequence[str]) -> int | None:
    headers = [_normalized(value) for value in values]
    needles = [_normalized(alias) for alias in aliases]

    # Exact bilingual/monolingual headers have priority. This prevents the
    # adjacent generic ``Ответственный / Sorumlu`` column from being selected.
    for needle in needles:
        for index, header in enumerate(headers):
            if header == needle:
                return index

    # Accept harmless suffixes/prefixes occasionally added to the same field.
    for needle in needles:
        for index, header in enumerate(headers):
            if needle and needle in header:
                return index
    return None


def _load_responsible_values(path: Path) -> dict[str, str]:
    """Read subdivision responsible persons keyed by normalized plate."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        result: dict[str, str] = {}
        for sheet in workbook.worksheets:
            iterator = iter(sheet.iter_rows(values_only=True))
            plate_column: int | None = None
            responsible_column: int | None = None

            for _ in range(80):
                try:
                    row = next(iterator)
                except StopIteration:
                    break
                candidate_plate = _find_column(row, PLATE_ALIASES)
                candidate_responsible = _find_column(row, RESPONSIBLE_ALIASES)
                if candidate_plate is not None:
                    plate_column = candidate_plate
                    responsible_column = candidate_responsible
                    break

            if plate_column is None or responsible_column is None:
                continue

            for row in iterator:
                if plate_column >= len(row):
                    continue
                normalized_plate = normalize_plate(row[plate_column])
                if not normalized_plate:
                    continue
                value = (
                    _clean(row[responsible_column])
                    if responsible_column < len(row)
                    else ""
                )
                if value:
                    result[normalized_plate] = value
        return result
    finally:
        workbook.close()


def ensure_schema(connection: psycopg.Connection) -> None:
    """Create the existing central roster schema and add ``responsible``."""
    _BASE_ENSURE_SCHEMA(connection)
    with connection.cursor() as cursor:
        cursor.execute(
            "ALTER TABLE consolidated_roster_entries "
            "ADD COLUMN IF NOT EXISTS responsible TEXT NOT NULL DEFAULT ''"
        )


def _write_roster_workbook(rows: Sequence[Sequence[str]]) -> bytes:
    """Write a complete stored roster including subdivision responsible."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Разнарядка"
    sheet.append([
        "Гос рег знак",
        "Компания или фирма",
        "Марка, модель",
        "Грейд",
        "ПОЛЬЗОВАТЕЛЬ",
        "Должность",
        "Дирекция",
        "Ответственный по подразделению / Birim sorumlusu",
    ])
    for source_row in rows:
        row = list(source_row)
        if len(row) < 8:
            row.extend([""] * (8 - len(row)))
        sheet.append(row[:8])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column, width in {
        "A": 18,
        "B": 28,
        "C": 26,
        "D": 14,
        "E": 38,
        "F": 38,
        "G": 52,
        "H": 38,
    }.items():
        sheet.column_dimensions[column].width = width
    stream = io.BytesIO()
    workbook.save(stream)
    workbook.close()
    return stream.getvalue()


def save_roster_uploads(
    database_url: str,
    uploads: Sequence[tuple[str, bytes]],
) -> int:
    """Save established fields, then persist subdivision responsible persons."""
    if not uploads:
        return 0

    saved = _BASE_SAVE_ROSTER_UPLOADS(database_url, uploads)

    with tempfile.TemporaryDirectory(prefix="arvento_responsible_roster_") as temp_name:
        temp_dir = Path(temp_name)
        paths: list[Path] = []
        responsible_by_path: dict[Path, dict[str, str]] = {}

        for index, (filename, content) in enumerate(uploads, start=1):
            path = temp_dir / f"{index:02d}_{Path(filename).name}"
            path.write_bytes(content)
            paths.append(path)
            responsible_by_path[path.resolve()] = _load_responsible_values(path)

        # load_rosters applies the same dated-roster and duplicate-date rules as
        # the established central store. The last upload for a duplicate date is
        # authoritative.
        rosters = load_rosters(paths)

        with psycopg.connect(database_url) as connection:
            ensure_schema(connection)
            connection.commit()
            with connection.cursor() as cursor:
                for roster in rosters:
                    values = responsible_by_path.get(roster.path.resolve(), {})
                    for normalized_plate, responsible in values.items():
                        cursor.execute(
                            """
                            UPDATE consolidated_roster_entries
                            SET responsible=%s
                            WHERE roster_day=%s
                              AND normalized_plate=%s
                            """,
                            (responsible, roster.day, normalized_plate),
                        )
            connection.commit()

    return saved


def _load_all_rows(
    database_url: str,
) -> dict[date, list[tuple[str, str, str, str, str, str, str, str]]]:
    with psycopg.connect(database_url) as connection:
        ensure_schema(connection)
        connection.commit()
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT
                    roster_day, plate, company, model, grade,
                    user_name, position, directorate, responsible
                FROM consolidated_roster_entries
                ORDER BY roster_day, normalized_plate
                """
            )
            rows = cursor.fetchall()

    grouped: defaultdict[
        date,
        list[tuple[str, str, str, str, str, str, str, str]],
    ] = defaultdict(list)
    for (
        roster_day,
        plate,
        company,
        model,
        grade,
        user_name,
        position,
        directorate,
        responsible,
    ) in rows:
        grouped[roster_day].append(
            (
                plate or "",
                company or "",
                model or "",
                grade or "",
                user_name or "",
                position or "",
                directorate or "",
                responsible or "",
            )
        )
    return grouped


def export_stored_rosters(database_url: str, target_dir: Path) -> list[Path]:
    """Export stored snapshots without discarding the responsible field."""
    grouped = _load_all_rows(database_url)
    if not grouped:
        raise ValueError(
            "В базе нет сохранённых разнарядок. Сначала загрузите разнарядку на странице «Разнарядки»."
        )

    result: list[Path] = []
    for roster_day in sorted(grouped):
        path = target_dir / f"roster_{roster_day.isoformat()}.xlsx"
        path.write_bytes(_write_roster_workbook(grouped[roster_day]))
        result.append(path)
    return result


def build_roster_download(roster_day: date) -> tuple[bytes, str]:
    """Download one stored roster with its subdivision responsible values."""
    import consolidated_portal as portal

    database_url = portal.implementation.db_url()
    with psycopg.connect(database_url) as connection:
        ensure_schema(connection)
        connection.commit()
        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT 1 FROM consolidated_roster_snapshots WHERE roster_day=%s",
                (roster_day,),
            )
            if cursor.fetchone() is None:
                raise ValueError("Разнарядка за выбранную дату не найдена")
            cursor.execute(
                """
                SELECT plate, company, model, grade,
                       user_name, position, directorate, responsible
                FROM consolidated_roster_entries
                WHERE roster_day=%s
                ORDER BY normalized_plate
                """,
                (roster_day,),
            )
            rows = [
                tuple(value or "" for value in row)
                for row in cursor.fetchall()
            ]

    return _write_roster_workbook(rows), f"Разнарядка_{roster_day.isoformat()}.xlsx"


def apply_responsible_roster_fields() -> None:
    """Install responsible storage and lossless stored-roster exports."""
    global _PATCHED
    if _PATCHED:
        return

    extended.ensure_schema = ensure_schema
    extended.save_roster_uploads = save_roster_uploads
    extended._write_roster_workbook = _write_roster_workbook
    extended._load_all_rows = _load_all_rows
    extended.export_stored_rosters = export_stored_rosters
    extended.build_roster_download = build_roster_download

    cache.ensure_schema = ensure_schema
    cache.save_roster_uploads = save_roster_uploads
    cache.export_stored_rosters = export_stored_rosters

    try:
        import consolidated_cache_portal as cache_portal

        cache_portal.save_roster_uploads = save_roster_uploads
    except ImportError:
        pass

    try:
        import roster_management_portal as roster_portal

        roster_portal.ensure_schema = ensure_schema
        roster_portal.save_roster_uploads = save_roster_uploads
        roster_portal.build_roster_download = build_roster_download
    except ImportError:
        pass

    _PATCHED = True


__all__ = [
    "RESPONSIBLE_ALIASES",
    "apply_responsible_roster_fields",
    "build_roster_download",
    "ensure_schema",
    "export_stored_rosters",
    "save_roster_uploads",
]
